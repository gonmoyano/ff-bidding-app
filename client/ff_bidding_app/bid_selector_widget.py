"""
Bid Selector Widget
Reusable widget component for selecting and managing Bids (CustomEntity06).
"""

import random
import string

from PySide6 import QtWidgets, QtCore, QtGui
from PySide6.QtSvg import QSvgRenderer

# File cog icon (from Material Design Icons - https://pictogrammers.com/library/mdi/icon/file-cog/)
FILE_COG_SVG_PATH = "M6 2C4.89 2 4 2.9 4 4V20C4 21.11 4.89 22 6 22H12.68A7 7 0 0 1 12 19A7 7 0 0 1 19 12A7 7 0 0 1 22 12.68V8L15 2H6M13 3.5L19.5 10H13V3.5M18 14C17.87 14 17.76 14.09 17.74 14.21L17.55 15.53C17.25 15.66 16.96 15.82 16.7 16L15.46 15.5C15.35 15.5 15.22 15.5 15.15 15.63L14.15 17.36C14.09 17.47 14.11 17.6 14.21 17.68L15.27 18.5C15.25 18.67 15.24 18.83 15.24 19C15.24 19.17 15.25 19.33 15.27 19.5L14.21 20.32C14.12 20.4 14.09 20.53 14.15 20.64L15.15 22.37C15.21 22.5 15.34 22.5 15.46 22.5L16.7 22C16.96 22.18 17.24 22.35 17.55 22.47L17.74 23.79C17.76 23.91 17.86 24 18 24H20C20.11 24 20.22 23.91 20.24 23.79L20.43 22.47C20.73 22.34 21 22.18 21.27 22L22.5 22.5C22.63 22.5 22.76 22.5 22.83 22.37L23.83 20.64C23.89 20.53 23.86 20.4 23.77 20.32L22.7 19.5C22.72 19.33 22.74 19.17 22.74 19C22.74 18.83 22.73 18.67 22.7 18.5L23.76 17.68C23.85 17.6 23.88 17.47 23.82 17.36L22.82 15.63C22.76 15.5 22.63 15.5 22.5 15.5L21.27 16C21 15.82 20.73 15.65 20.42 15.53L20.23 14.21C20.22 14.09 20.11 14 20 14H18M19 17.5C19.83 17.5 20.5 18.17 20.5 19C20.5 19.83 19.83 20.5 19 20.5C18.16 20.5 17.5 19.83 17.5 19C17.5 18.17 18.17 17.5 19 17.5Z"


def create_icon_from_svg_path(svg_path, size=24, color="#e0e0e0"):
    """Create a QIcon from an SVG path string."""
    svg_content = f'''<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="{size}" height="{size}">
        <path fill="{color}" d="{svg_path}"/>
    </svg>'''
    renderer = QSvgRenderer(svg_content.encode('utf-8'))
    pixmap = QtGui.QPixmap(size, size)
    pixmap.fill(QtCore.Qt.transparent)
    painter = QtGui.QPainter(pixmap)
    renderer.render(painter)
    painter.end()
    return QtGui.QIcon(pixmap)

try:
    from .logger import logger
    from .settings import AppSettings
except ImportError:
    import logging
    logger = logging.getLogger("FFPackageManager")
    from settings import AppSettings


class CollapsibleGroupBox(QtWidgets.QWidget):
    """A collapsible group box that hides/shows content without disabling it."""

    def __init__(self, title="", parent=None):
        """Initialize the collapsible group box.

        Args:
            title: The title for the group
            parent: Parent widget
        """
        super().__init__(parent)
        self.is_collapsed = False
        self.base_title = title
        self.additional_info = ""

        # Main layout
        main_layout = QtWidgets.QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # Create a custom title bar widget with label for rich text support
        self.title_bar = QtWidgets.QWidget()
        self.title_bar.setStyleSheet("""
            QWidget {
                border: 1px solid #555555;
                border-radius: 4px;
                padding: 3px 5px;
                background-color: #2b2b2b;
            }
        """)
        self.title_bar.setCursor(QtCore.Qt.PointingHandCursor)
        self.title_bar.mousePressEvent = lambda event: self._on_toggle()

        title_layout = QtWidgets.QHBoxLayout(self.title_bar)
        title_layout.setContentsMargins(2, 2, 2, 2)

        # Label with rich text support
        self.title_label = QtWidgets.QLabel()
        self.title_label.setStyleSheet("color: #e0e0e0; font-weight: bold; font-size: 11px; border: none; background: transparent;")
        self.title_label.setTextFormat(QtCore.Qt.RichText)
        title_layout.addWidget(self.title_label)

        self._update_button_text()
        main_layout.addWidget(self.title_bar)

        # Content frame
        self.content_frame = QtWidgets.QFrame()
        self.content_frame.setObjectName("collapsibleContent")
        self.content_frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.content_frame.setStyleSheet("""
            QFrame#collapsibleContent {
                border: 1px solid #555555;
                border-radius: 4px;
            }
        """)
        self.content_layout = QtWidgets.QVBoxLayout(self.content_frame)
        main_layout.addWidget(self.content_frame)

    def _update_button_text(self):
        """Update button text with arrow indicator and additional info."""
        arrow = "▼" if not self.is_collapsed else "▶"
        # Show additional info when collapsed with blue color
        if self.is_collapsed and self.additional_info:
            # Use HTML to color the additional info in blue
            html_text = f'{arrow} {self.base_title} | <span style="color:#6b9bd1;">{self.additional_info}</span>'
            self.title_label.setText(html_text)
        else:
            self.title_label.setText(f"{arrow} {self.base_title}")

    def _on_toggle(self):
        """Toggle the visibility of the content."""
        self.is_collapsed = not self.is_collapsed
        self.content_frame.setVisible(not self.is_collapsed)
        self._update_button_text()

    def setTitle(self, title):
        """Set the title of the group box."""
        self.base_title = title
        self._update_button_text()

    def setAdditionalInfo(self, info):
        """Set additional information to display in the title bar.

        Args:
            info: Additional info string to display after the title
        """
        self.additional_info = info
        self._update_button_text()

    def set_collapsed(self, collapsed):
        """Set the collapsed state of the group.

        Args:
            collapsed: True to collapse, False to expand
        """
        if self.is_collapsed != collapsed:
            self._on_toggle()

    def addWidget(self, widget):
        """Add a widget to the content area."""
        self.content_layout.addWidget(widget)

    def addLayout(self, layout):
        """Add a layout to the content area."""
        self.content_layout.addLayout(layout)


class TickCheckBox(QtWidgets.QCheckBox):
    """Custom checkbox that draws a tick mark instead of a solid indicator when checked."""

    def __init__(self, text="", parent=None):
        """Initialize the tick checkbox.

        Args:
            text: Label text for the checkbox
            parent: Parent widget
        """
        super().__init__(text, parent)

    def paintEvent(self, event):
        """Custom paint to draw tick mark when checked."""
        super().paintEvent(event)

        if self.isChecked():
            painter = QtGui.QPainter(self)
            painter.setRenderHint(QtGui.QPainter.Antialiasing)

            # Get the indicator rect (checkbox square area)
            style = self.style()
            opt = QtWidgets.QStyleOptionButton()
            self.initStyleOption(opt)
            indicator_rect = style.subElementRect(
                QtWidgets.QStyle.SE_CheckBoxIndicator, opt, self
            )

            # Draw the tick mark
            pen = QtGui.QPen(QtGui.QColor("#4a9eff"))
            pen.setWidth(2)
            pen.setCapStyle(QtCore.Qt.RoundCap)
            painter.setPen(pen)

            # Calculate tick mark points within the indicator
            x = indicator_rect.x()
            y = indicator_rect.y()
            w = indicator_rect.width()
            h = indicator_rect.height()

            # Tick mark coordinates (checkmark shape)
            start_point = QtCore.QPointF(x + w * 0.2, y + h * 0.5)
            mid_point = QtCore.QPointF(x + w * 0.4, y + h * 0.7)
            end_point = QtCore.QPointF(x + w * 0.8, y + h * 0.3)

            painter.drawLine(start_point, mid_point)
            painter.drawLine(mid_point, end_point)

            painter.end()


class AddBidDialog(QtWidgets.QDialog):
    """Dialog for adding a new Bid with option to copy from existing."""

    def __init__(self, sg_session, project_id, rfq_id, parent=None):
        """Initialize the dialog.

        Args:
            sg_session: ShotgridClient instance
            project_id: ID of the current project
            rfq_id: ID of the current RFQ
            parent: Parent widget
        """
        super().__init__(parent)
        self.sg_session = sg_session
        self.project_id = project_id
        self.rfq_id = rfq_id
        self.existing_bids = []

        self.setWindowTitle("Add New Bid")
        self.setModal(True)
        self.setMinimumWidth(500)

        self._load_existing_bids()
        self._build_ui()

    def _load_existing_bids(self):
        """Load existing Bids for the copy option."""
        try:
            if self.project_id:
                # Get all bids for this project (not just this RFQ)
                self.existing_bids = self.sg_session.get_bids(
                    self.project_id,
                    fields=["id", "code", "sg_bid_type", "sg_vfx_breakdown", "sg_bid_assets", "sg_price_list"]
                )
                # Sort by name
                self.existing_bids.sort(key=lambda x: x.get("code", "").lower())
        except Exception as e:
            logger.error(f"Failed to load existing Bids: {e}")
            self.existing_bids = []

    def _build_ui(self):
        """Build the dialog UI."""
        layout = QtWidgets.QVBoxLayout(self)

        # Bid name field
        name_layout = QtWidgets.QHBoxLayout()
        name_label = QtWidgets.QLabel("Bid Name:")
        name_layout.addWidget(name_label)

        self.name_field = QtWidgets.QLineEdit()
        self.name_field.setPlaceholderText("Enter bid name...")
        name_layout.addWidget(self.name_field, stretch=1)

        layout.addLayout(name_layout)

        # Bid type selection
        type_layout = QtWidgets.QHBoxLayout()
        type_label = QtWidgets.QLabel("Bid Type:")
        type_layout.addWidget(type_label)

        self.type_combo = QtWidgets.QComboBox()
        self.type_combo.addItem("Early Bid")
        self.type_combo.addItem("Turnover Bid")
        type_layout.addWidget(self.type_combo, stretch=1)

        layout.addLayout(type_layout)

        # Description field
        desc_layout = QtWidgets.QHBoxLayout()
        desc_label = QtWidgets.QLabel("Description:")
        desc_label.setAlignment(QtCore.Qt.AlignTop)
        desc_layout.addWidget(desc_label)

        self.description_field = QtWidgets.QPlainTextEdit()
        self.description_field.setPlaceholderText("Enter bid description...")
        self.description_field.setMaximumHeight(60)
        desc_layout.addWidget(self.description_field, stretch=1)

        layout.addLayout(desc_layout)

        # Creation options group
        options_group = QtWidgets.QGroupBox("Creation Options")
        options_layout = QtWidgets.QVBoxLayout(options_group)

        # Custom stylesheet for radio buttons (radial style with inner circle)
        radio_style = """
            QRadioButton {
                spacing: 8px;
                font-weight: normal;
            }
            QRadioButton::indicator {
                width: 16px;
                height: 16px;
                border-radius: 9px;
                border: 2px solid #666666;
                background-color: #2b2b2b;
            }
            QRadioButton::indicator:checked {
                border: 2px solid #4a9eff;
                background-color: qradialgradient(
                    cx:0.5, cy:0.5, radius:0.4,
                    fx:0.5, fy:0.5,
                    stop:0 #4a9eff, stop:0.6 #4a9eff, stop:0.7 #2b2b2b, stop:1 #2b2b2b
                );
            }
            QRadioButton::indicator:hover {
                border: 2px solid #888888;
            }
            QRadioButton::indicator:checked:hover {
                border: 2px solid #6ab0ff;
            }
        """

        # Radio button: Create new (with child entities)
        self.create_new_radio = QtWidgets.QRadioButton("Create new")
        self.create_new_radio.setStyleSheet(radio_style)
        self.create_new_radio.setToolTip(
            "Create a new Bid with VFX Breakdown, Bid Assets, and Price List"
        )
        self.create_new_radio.setChecked(True)
        self.create_new_radio.toggled.connect(self._on_radio_toggled)
        options_layout.addWidget(self.create_new_radio)

        # Radio button: Copy from existing
        self.copy_from_radio = QtWidgets.QRadioButton("Copy from existing Bid")
        self.copy_from_radio.setStyleSheet(radio_style)
        self.copy_from_radio.setToolTip(
            "Copy VFX Breakdown, Bid Assets, and Price List from an existing Bid"
        )
        self.copy_from_radio.toggled.connect(self._on_radio_toggled)
        options_layout.addWidget(self.copy_from_radio)

        # Container widget for copy options (for smooth show/hide)
        self.copy_options_container = QtWidgets.QWidget()
        copy_container_layout = QtWidgets.QVBoxLayout(self.copy_options_container)
        copy_container_layout.setContentsMargins(0, 0, 0, 0)
        copy_container_layout.setSpacing(5)

        # ComboBox for selecting source Bid (only visible when copy option selected)
        self.source_combo_layout = QtWidgets.QHBoxLayout()
        self.source_combo_layout.setContentsMargins(20, 5, 0, 0)  # Indent

        self.source_label = QtWidgets.QLabel("Source Bid:")
        self.source_combo_layout.addWidget(self.source_label)

        self.source_combo = QtWidgets.QComboBox()
        self.source_combo.setMinimumWidth(300)
        for bid in self.existing_bids:
            bid_type = bid.get("sg_bid_type", "")
            display_text = f"{bid.get('code', 'Unnamed')}"
            if bid_type:
                display_text += f" ({bid_type})"
            self.source_combo.addItem(display_text, bid["id"])
        self.source_combo_layout.addWidget(self.source_combo, stretch=1)

        copy_container_layout.addLayout(self.source_combo_layout)

        # Custom stylesheet for checkboxes (tick icon style - transparent with tick mark)
        checkbox_style = """
            QCheckBox {
                spacing: 8px;
                font-weight: normal;
            }
            QCheckBox::indicator {
                width: 16px;
                height: 16px;
                border-radius: 3px;
                border: 2px solid #666666;
                background-color: transparent;
            }
            QCheckBox::indicator:unchecked {
                border: 2px solid #666666;
                background-color: transparent;
            }
            QCheckBox::indicator:checked {
                border: 2px solid #4a9eff;
                background-color: transparent;
            }
            QCheckBox::indicator:hover {
                border: 2px solid #888888;
            }
            QCheckBox::indicator:checked:hover {
                border: 2px solid #6ab0ff;
            }
        """

        # Checkboxes for what to copy (only visible when copy option selected)
        self.copy_options_layout = QtWidgets.QVBoxLayout()
        self.copy_options_layout.setContentsMargins(20, 5, 0, 0)  # Indent

        # VFX Breakdown checkbox
        self.copy_vfx_breakdown_checkbox = TickCheckBox("Copy VFX Breakdown (with Bidding Scenes)")
        self.copy_vfx_breakdown_checkbox.setStyleSheet(checkbox_style)
        self.copy_vfx_breakdown_checkbox.setChecked(True)
        self.copy_options_layout.addWidget(self.copy_vfx_breakdown_checkbox)

        # Bid Assets checkbox
        self.copy_bid_assets_checkbox = TickCheckBox("Copy Bid Assets (with Asset Items)")
        self.copy_bid_assets_checkbox.setStyleSheet(checkbox_style)
        self.copy_bid_assets_checkbox.setChecked(True)
        self.copy_options_layout.addWidget(self.copy_bid_assets_checkbox)

        # Price List checkbox
        self.copy_price_list_checkbox = TickCheckBox("Copy Price List (with Line Items)")
        self.copy_price_list_checkbox.setStyleSheet(checkbox_style)
        self.copy_price_list_checkbox.setChecked(True)
        self.copy_options_layout.addWidget(self.copy_price_list_checkbox)

        copy_container_layout.addLayout(self.copy_options_layout)

        options_layout.addWidget(self.copy_options_container)

        # Initially hide copy options container since "Create new" is selected
        self.copy_options_container.setVisible(False)

        # Disable copy option if no existing Bids
        if not self.existing_bids:
            self.copy_from_radio.setEnabled(False)
            self.copy_from_radio.setToolTip("No existing Bids to copy from")

        layout.addWidget(options_group)

        # Buttons
        button_layout = QtWidgets.QHBoxLayout()
        button_layout.addStretch()

        self.ok_button = QtWidgets.QPushButton("Create")
        self.ok_button.clicked.connect(self.accept)
        self.ok_button.setDefault(True)
        button_layout.addWidget(self.ok_button)

        self.cancel_button = QtWidgets.QPushButton("Cancel")
        self.cancel_button.clicked.connect(self.reject)
        button_layout.addWidget(self.cancel_button)

        layout.addLayout(button_layout)

    def _on_radio_toggled(self, checked):
        """Handle radio button toggle."""
        is_copy = self.copy_from_radio.isChecked()
        self.copy_options_container.setVisible(is_copy)

        # Force layout to recalculate
        self.layout().invalidate()
        self.layout().activate()

        # Adjust dialog size when toggling
        if is_copy:
            # Expand to fit copy options
            self.adjustSize()
        else:
            # Reset to minimum size when "Create new" is selected
            # Set size constraint to allow shrinking
            self.setMinimumHeight(0)
            self.setMaximumHeight(16777215)  # QWIDGETSIZE_MAX

            # Process pending layout events
            QtWidgets.QApplication.processEvents()

            # Get the new minimum size hint and resize
            new_size = self.sizeHint()
            self.resize(self.width(), new_size.height())

    def get_bid_name(self):
        """Get the bid name from the dialog."""
        return self.name_field.text().strip()

    def get_bid_type(self):
        """Get the selected bid type."""
        return self.type_combo.currentText()

    def get_description(self):
        """Get the bid description from the dialog."""
        return self.description_field.toPlainText().strip()

    def is_copy_mode(self):
        """Check if copy mode is selected."""
        return self.copy_from_radio.isChecked()

    def get_source_bid_id(self):
        """Get the selected source Bid ID for copying."""
        if self.is_copy_mode():
            return self.source_combo.currentData()
        return None

    def get_source_bid_data(self):
        """Get the full data for the selected source Bid."""
        source_id = self.get_source_bid_id()
        if source_id:
            for bid in self.existing_bids:
                if bid["id"] == source_id:
                    return bid
        return None

    def should_copy_vfx_breakdown(self):
        """Check if VFX Breakdown should be copied."""
        return self.is_copy_mode() and self.copy_vfx_breakdown_checkbox.isChecked()

    def should_copy_bid_assets(self):
        """Check if Bid Assets should be copied."""
        return self.is_copy_mode() and self.copy_bid_assets_checkbox.isChecked()

    def should_copy_price_list(self):
        """Check if Price List should be copied."""
        return self.is_copy_mode() and self.copy_price_list_checkbox.isChecked()


class RenameBidDialog(QtWidgets.QDialog):
    """Dialog for renaming a Bid."""

    def __init__(self, current_name, parent=None):
        """Initialize the dialog."""
        super().__init__(parent)
        self.current_name = current_name
        self.setWindowTitle("Rename Bid")
        self.setModal(True)
        self.setMinimumWidth(400)

        self._build_ui()

    def _build_ui(self):
        """Build the dialog UI."""
        layout = QtWidgets.QVBoxLayout(self)

        # Name field
        name_layout = QtWidgets.QHBoxLayout()
        name_label = QtWidgets.QLabel("New Name:")
        name_layout.addWidget(name_label)

        self.name_field = QtWidgets.QLineEdit()
        self.name_field.setText(self.current_name)
        self.name_field.selectAll()  # Select all text for easy editing
        self.name_field.setPlaceholderText("Enter new bid name...")
        name_layout.addWidget(self.name_field, stretch=1)

        layout.addLayout(name_layout)

        # Buttons
        button_layout = QtWidgets.QHBoxLayout()
        button_layout.addStretch()

        self.ok_button = QtWidgets.QPushButton("OK")
        self.ok_button.clicked.connect(self.accept)
        self.ok_button.setDefault(True)
        button_layout.addWidget(self.ok_button)

        self.cancel_button = QtWidgets.QPushButton("Cancel")
        self.cancel_button.clicked.connect(self.reject)
        button_layout.addWidget(self.cancel_button)

        layout.addLayout(button_layout)

    def get_new_name(self):
        """Get the new name from the dialog."""
        return self.name_field.text().strip()


class ConfigBidDialog(QtWidgets.QDialog):
    """Dialog for configuring Bid properties and children entities."""

    # Signal emitted when bid is deleted (bid_id)
    bidDeleted = QtCore.Signal(int)

    def __init__(self, sg_session, project_id, bid_id, bid_data, parent=None):
        """Initialize the dialog.

        Args:
            sg_session: ShotgridClient instance
            project_id: ID of the current project
            bid_id: ID of the current Bid
            bid_data: Current bid data dict
            parent: Parent widget
        """
        super().__init__(parent)
        self.sg_session = sg_session
        self.project_id = project_id
        self.bid_id = bid_id
        self.bid_data = bid_data
        self._bid_was_deleted = False

        # Generate random confirmation string for delete
        self.confirmation_string = ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))

        self.setWindowTitle("Configure Bid")
        self.setModal(True)
        self.setMinimumWidth(500)

        self._load_available_items()
        self._build_ui()

    def _load_available_items(self):
        """Load available items for each dropdown."""
        try:
            # Load VFX Breakdowns for this bid
            self.vfx_breakdowns = self.sg_session.sg.find(
                "CustomEntity01",
                [
                    ["project", "is", {"type": "Project", "id": int(self.project_id)}],
                    ["sg_parent_bid", "is", {"type": "CustomEntity06", "id": int(self.bid_id)}]
                ],
                ["id", "code", "name"]
            )
            self.vfx_breakdowns.sort(key=lambda x: x.get("code", "").lower())
        except Exception as e:
            logger.error(f"Failed to load VFX Breakdowns: {e}")
            self.vfx_breakdowns = []

        try:
            # Load Bid Assets for this bid
            self.bid_assets = self.sg_session.sg.find(
                "CustomEntity08",
                [
                    ["project", "is", {"type": "Project", "id": int(self.project_id)}],
                    ["sg_parent_bid", "is", {"type": "CustomEntity06", "id": int(self.bid_id)}]
                ],
                ["id", "code", "name"]
            )
            self.bid_assets.sort(key=lambda x: x.get("code", "").lower())
        except Exception as e:
            logger.error(f"Failed to load Bid Assets: {e}")
            self.bid_assets = []

        try:
            # Load Price Lists for this bid
            self.price_lists = self.sg_session.sg.find(
                "CustomEntity10",
                [
                    ["project", "is", {"type": "Project", "id": int(self.project_id)}],
                    ["sg_parent_bid", "is", {"type": "CustomEntity06", "id": int(self.bid_id)}]
                ],
                ["id", "code", "name", "sg_rate_card"]
            )
            self.price_lists.sort(key=lambda x: x.get("code", "").lower())
        except Exception as e:
            logger.error(f"Failed to load Price Lists: {e}")
            self.price_lists = []

        try:
            # Load Rate Cards (non-project entity)
            self.rate_cards = self.sg_session.sg.find(
                "CustomNonProjectEntity01",
                [],
                ["id", "code", "name"]
            )
            self.rate_cards.sort(key=lambda x: x.get("code", "").lower())
        except Exception as e:
            logger.error(f"Failed to load Rate Cards: {e}")
            self.rate_cards = []

        try:
            # Load all existing bids for duplicate name validation
            self.existing_bids = self.sg_session.sg.find(
                "CustomEntity06",
                [["project", "is", {"type": "Project", "id": int(self.project_id)}]],
                ["id", "code"]
            )
        except Exception as e:
            logger.error(f"Failed to load existing bids: {e}")
            self.existing_bids = []

    def _build_ui(self):
        """Build the dialog UI."""
        layout = QtWidgets.QVBoxLayout(self)

        # Bid Name field
        name_layout = QtWidgets.QHBoxLayout()
        name_label = QtWidgets.QLabel("Bid Name:")
        name_label.setMinimumWidth(100)
        name_layout.addWidget(name_label)

        self.name_field = QtWidgets.QLineEdit()
        self.name_field.setText(self.bid_data.get("code", ""))
        self.name_field.setPlaceholderText("Enter bid name")
        name_layout.addWidget(self.name_field, stretch=1)
        layout.addLayout(name_layout)

        # Description field
        desc_layout = QtWidgets.QHBoxLayout()
        desc_label = QtWidgets.QLabel("Description:")
        desc_label.setMinimumWidth(100)
        desc_label.setAlignment(QtCore.Qt.AlignTop)
        desc_layout.addWidget(desc_label)

        self.description_field = QtWidgets.QTextEdit()
        self.description_field.setPlainText(self.bid_data.get("description", "") or "")
        self.description_field.setPlaceholderText("Enter bid description")
        self.description_field.setMaximumHeight(60)
        desc_layout.addWidget(self.description_field, stretch=1)
        layout.addLayout(desc_layout)

        # Separator
        layout.addSpacing(10)
        separator = QtWidgets.QFrame()
        separator.setFrameShape(QtWidgets.QFrame.HLine)
        separator.setFrameShadow(QtWidgets.QFrame.Sunken)
        layout.addWidget(separator)
        layout.addSpacing(5)

        # VFX Breakdown dropdown
        vfx_layout = QtWidgets.QHBoxLayout()
        vfx_label = QtWidgets.QLabel("VFX Breakdown:")
        vfx_label.setMinimumWidth(100)
        vfx_layout.addWidget(vfx_label)

        self.vfx_combo = QtWidgets.QComboBox()
        self.vfx_combo.addItem("-- None --", None)
        for breakdown in self.vfx_breakdowns:
            name = breakdown.get("code") or breakdown.get("name") or f"ID {breakdown['id']}"
            self.vfx_combo.addItem(name, breakdown["id"])
        vfx_layout.addWidget(self.vfx_combo, stretch=1)
        layout.addLayout(vfx_layout)

        # Pre-select current VFX Breakdown
        current_vfx = self.bid_data.get("sg_vfx_breakdown")
        if current_vfx and isinstance(current_vfx, dict):
            current_vfx_id = current_vfx.get("id")
            for i in range(self.vfx_combo.count()):
                if self.vfx_combo.itemData(i) == current_vfx_id:
                    self.vfx_combo.setCurrentIndex(i)
                    break

        # Bid Assets dropdown
        assets_layout = QtWidgets.QHBoxLayout()
        assets_label = QtWidgets.QLabel("Bid Assets:")
        assets_label.setMinimumWidth(100)
        assets_layout.addWidget(assets_label)

        self.assets_combo = QtWidgets.QComboBox()
        self.assets_combo.addItem("-- None --", None)
        for assets in self.bid_assets:
            name = assets.get("code") or assets.get("name") or f"ID {assets['id']}"
            self.assets_combo.addItem(name, assets["id"])
        assets_layout.addWidget(self.assets_combo, stretch=1)
        layout.addLayout(assets_layout)

        # Pre-select current Bid Assets
        current_assets = self.bid_data.get("sg_bid_assets")
        if current_assets and isinstance(current_assets, dict):
            current_assets_id = current_assets.get("id")
            for i in range(self.assets_combo.count()):
                if self.assets_combo.itemData(i) == current_assets_id:
                    self.assets_combo.setCurrentIndex(i)
                    break

        # Price List dropdown
        price_layout = QtWidgets.QHBoxLayout()
        price_label = QtWidgets.QLabel("Price List:")
        price_label.setMinimumWidth(100)
        price_layout.addWidget(price_label)

        self.price_combo = QtWidgets.QComboBox()
        self.price_combo.addItem("-- None --", None)
        for price_list in self.price_lists:
            name = price_list.get("code") or price_list.get("name") or f"ID {price_list['id']}"
            self.price_combo.addItem(name, price_list["id"])
        price_layout.addWidget(self.price_combo, stretch=1)
        layout.addLayout(price_layout)

        # Pre-select current Price List
        current_price = self.bid_data.get("sg_price_list")
        if current_price and isinstance(current_price, dict):
            current_price_id = current_price.get("id")
            for i in range(self.price_combo.count()):
                if self.price_combo.itemData(i) == current_price_id:
                    self.price_combo.setCurrentIndex(i)
                    break

        # Rate Card dropdown
        rate_layout = QtWidgets.QHBoxLayout()
        rate_label = QtWidgets.QLabel("Rate Card:")
        rate_label.setMinimumWidth(100)
        rate_layout.addWidget(rate_label)

        self.rate_combo = QtWidgets.QComboBox()
        self.rate_combo.addItem("-- None --", None)
        base_rates_index = 0
        for idx, rate_card in enumerate(self.rate_cards):
            name = rate_card.get("code") or rate_card.get("name") or f"ID {rate_card['id']}"
            self.rate_combo.addItem(name, rate_card["id"])
            if name.lower() == "base rates":
                base_rates_index = idx + 1

        rate_layout.addWidget(self.rate_combo, stretch=1)
        layout.addLayout(rate_layout)

        # Pre-select current Rate Card from Price List, or default to Base Rates
        self._select_rate_card(base_rates_index)

        # Connect price list change signal now that rate_combo exists
        self.price_combo.currentIndexChanged.connect(self._on_price_list_changed)

        # Add spacing before buttons
        layout.addSpacing(15)

        # Buttons row
        button_layout = QtWidgets.QHBoxLayout()

        # Remove button on the left
        self.remove_button = QtWidgets.QPushButton("Remove Bid...")
        self.remove_button.setStyleSheet("color: #ff6666;")
        self.remove_button.clicked.connect(self._on_remove_bid)
        button_layout.addWidget(self.remove_button)

        button_layout.addStretch()

        self.ok_button = QtWidgets.QPushButton("Save")
        self.ok_button.clicked.connect(self._validate_and_accept)
        self.ok_button.setDefault(True)
        button_layout.addWidget(self.ok_button)

        self.cancel_button = QtWidgets.QPushButton("Cancel")
        self.cancel_button.clicked.connect(self.reject)
        button_layout.addWidget(self.cancel_button)

        layout.addLayout(button_layout)

    def _validate_and_accept(self):
        """Validate bid name and accept if valid."""
        new_name = self.name_field.text().strip()
        current_name = self.bid_data.get("code", "")

        # Check if name is empty
        if not new_name:
            QtWidgets.QMessageBox.warning(
                self,
                "Validation Error",
                "Bid name cannot be empty."
            )
            self.name_field.setFocus()
            return

        # Check for duplicate name (only if name changed)
        if new_name.lower() != current_name.lower():
            for bid in self.existing_bids:
                if bid["id"] != self.bid_id and bid.get("code", "").lower() == new_name.lower():
                    QtWidgets.QMessageBox.warning(
                        self,
                        "Duplicate Name",
                        f"A bid with the name '{bid.get('code')}' already exists.\n\n"
                        "Please choose a different name."
                    )
                    self.name_field.setFocus()
                    self.name_field.selectAll()
                    return

        self.accept()

    def _on_price_list_changed(self, index):
        """Handle price list selection change to update rate card."""
        price_list_id = self.price_combo.currentData()
        if price_list_id:
            for price_list in self.price_lists:
                if price_list["id"] == price_list_id:
                    rate_card = price_list.get("sg_rate_card")
                    if rate_card and isinstance(rate_card, dict):
                        rate_card_id = rate_card.get("id")
                        for i in range(self.rate_combo.count()):
                            if self.rate_combo.itemData(i) == rate_card_id:
                                self.rate_combo.setCurrentIndex(i)
                                return
                    break

    def _select_rate_card(self, default_index=0):
        """Select the rate card based on current price list or default."""
        current_price = self.bid_data.get("sg_price_list")
        if current_price and isinstance(current_price, dict):
            current_price_id = current_price.get("id")
            for price_list in self.price_lists:
                if price_list["id"] == current_price_id:
                    rate_card = price_list.get("sg_rate_card")
                    if rate_card and isinstance(rate_card, dict):
                        rate_card_id = rate_card.get("id")
                        for i in range(self.rate_combo.count()):
                            if self.rate_combo.itemData(i) == rate_card_id:
                                self.rate_combo.setCurrentIndex(i)
                                return
                    break
        if default_index > 0:
            self.rate_combo.setCurrentIndex(default_index)

    def _on_remove_bid(self):
        """Handle Remove Bid button click - show confirmation dialog."""
        bid_name = self.bid_data.get("code", f"Bid {self.bid_id}")

        # Build summary of children that will be deleted
        children_summary = []
        if self.vfx_breakdowns:
            names = [b.get("code") or b.get("name") or f"ID {b['id']}" for b in self.vfx_breakdowns]
            children_summary.append(f"• {len(self.vfx_breakdowns)} VFX Breakdown(s): {', '.join(names)}")
        if self.bid_assets:
            names = [a.get("code") or a.get("name") or f"ID {a['id']}" for a in self.bid_assets]
            children_summary.append(f"• {len(self.bid_assets)} Bid Assets: {', '.join(names)}")
        if self.price_lists:
            names = [p.get("code") or p.get("name") or f"ID {p['id']}" for p in self.price_lists]
            children_summary.append(f"• {len(self.price_lists)} Price List(s): {', '.join(names)}")

        # Create confirmation dialog
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("Remove Bid")
        dialog.setModal(True)
        dialog.setMinimumWidth(450)

        layout = QtWidgets.QVBoxLayout(dialog)

        # Warning message
        warning_label = QtWidgets.QLabel(
            f"⚠️ WARNING: This action cannot be undone!\n\n"
            f"Deleting bid '{bid_name}' will also delete all linked children:"
        )
        warning_label.setStyleSheet("color: #ff6666; font-weight: bold; padding: 10px;")
        warning_label.setWordWrap(True)
        layout.addWidget(warning_label)

        # Children summary
        if children_summary:
            summary_text = "\n".join(children_summary)
            summary_label = QtWidgets.QLabel(summary_text)
            summary_label.setStyleSheet("padding: 10px; background-color: #2a2a2a; border-radius: 4px;")
            summary_label.setWordWrap(True)
            layout.addWidget(summary_label)
        else:
            no_children_label = QtWidgets.QLabel("(No linked children found)")
            no_children_label.setStyleSheet("padding: 10px; color: #888888;")
            layout.addWidget(no_children_label)

        # Confirmation section
        layout.addSpacing(20)

        confirm_label = QtWidgets.QLabel(
            f"To confirm deletion, type the following string:\n\n{self.confirmation_string}"
        )
        confirm_label.setStyleSheet("font-weight: bold; padding: 10px;")
        confirm_label.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(confirm_label)

        # Confirmation input
        confirm_layout = QtWidgets.QHBoxLayout()
        confirm_input_label = QtWidgets.QLabel("Confirmation:")
        confirm_layout.addWidget(confirm_input_label)

        confirmation_field = QtWidgets.QLineEdit()
        confirmation_field.setPlaceholderText("Type confirmation string here...")
        confirm_layout.addWidget(confirmation_field, stretch=1)
        layout.addLayout(confirm_layout)

        # Buttons
        layout.addSpacing(20)
        button_layout = QtWidgets.QHBoxLayout()
        button_layout.addStretch()

        delete_button = QtWidgets.QPushButton("Delete")
        delete_button.setEnabled(False)
        delete_button.setStyleSheet("background-color: #ff6666; color: white; font-weight: bold;")
        button_layout.addWidget(delete_button)

        cancel_button = QtWidgets.QPushButton("Cancel")
        button_layout.addWidget(cancel_button)

        layout.addLayout(button_layout)

        # Connect signals
        def on_confirmation_changed(text):
            delete_button.setEnabled(text == self.confirmation_string)

        confirmation_field.textChanged.connect(on_confirmation_changed)
        delete_button.clicked.connect(dialog.accept)
        cancel_button.clicked.connect(dialog.reject)

        # Show dialog
        if dialog.exec_() != QtWidgets.QDialog.Accepted:
            return

        # Perform deletion
        try:
            # Delete children first
            for breakdown in self.vfx_breakdowns:
                # Also delete bidding scenes for each breakdown
                bidding_scenes = self.sg_session.sg.find(
                    "CustomEntity02",
                    [["sg_parent", "is", {"type": "CustomEntity01", "id": breakdown["id"]}]],
                    ["id"]
                )
                for scene in bidding_scenes:
                    self.sg_session.sg.delete("CustomEntity02", scene["id"])
                self.sg_session.sg.delete("CustomEntity01", breakdown["id"])
                logger.info(f"Deleted VFX Breakdown {breakdown['id']} with {len(bidding_scenes)} bidding scenes")

            for assets in self.bid_assets:
                # Also delete asset items
                asset_items = self.sg_session.sg.find(
                    "CustomEntity07",
                    [["sg_bid_assets", "is", {"type": "CustomEntity08", "id": assets["id"]}]],
                    ["id"]
                )
                for item in asset_items:
                    self.sg_session.sg.delete("CustomEntity07", item["id"])
                self.sg_session.sg.delete("CustomEntity08", assets["id"])
                logger.info(f"Deleted Bid Assets {assets['id']} with {len(asset_items)} asset items")

            for price_list in self.price_lists:
                # Also delete line items
                line_items = self.sg_session.sg.find(
                    "CustomEntity03",
                    [["sg_parent_pricelist", "is", {"type": "CustomEntity10", "id": price_list["id"]}]],
                    ["id"]
                )
                for item in line_items:
                    self.sg_session.sg.delete("CustomEntity03", item["id"])
                self.sg_session.sg.delete("CustomEntity10", price_list["id"])
                logger.info(f"Deleted Price List {price_list['id']} with {len(line_items)} line items")

            # Delete the bid itself
            self.sg_session.sg.delete("CustomEntity06", self.bid_id)
            logger.info(f"Deleted Bid {self.bid_id}")

            self._bid_was_deleted = True
            self.bidDeleted.emit(self.bid_id)

            QtWidgets.QMessageBox.information(
                self,
                "Success",
                f"Bid '{bid_name}' and all linked children have been deleted."
            )

            # Close the dialog
            self.reject()

        except Exception as e:
            logger.error(f"Failed to delete bid: {e}", exc_info=True)
            QtWidgets.QMessageBox.critical(
                self,
                "Error",
                f"Failed to delete bid:\n{str(e)}"
            )

    def was_bid_deleted(self):
        """Check if the bid was deleted during this dialog session."""
        return self._bid_was_deleted

    def get_bid_name(self):
        """Get the bid name from the text field."""
        return self.name_field.text().strip()

    def get_description(self):
        """Get the description from the text field."""
        return self.description_field.toPlainText().strip()

    def get_vfx_breakdown_id(self):
        """Get selected VFX Breakdown ID."""
        return self.vfx_combo.currentData()

    def get_bid_assets_id(self):
        """Get selected Bid Assets ID."""
        return self.assets_combo.currentData()

    def get_price_list_id(self):
        """Get selected Price List ID."""
        return self.price_combo.currentData()

    def get_rate_card_id(self):
        """Get selected Rate Card ID."""
        return self.rate_combo.currentData()


class CreateBidDialog(QtWidgets.QDialog):
    """Dialog for creating a new Bid during import."""

    def __init__(self, parent=None):
        """Initialize the dialog."""
        super().__init__(parent)
        self.setWindowTitle("Create Bid")
        self.setModal(True)
        self.setMinimumWidth(400)

        self._build_ui()

    def _build_ui(self):
        """Build the dialog UI."""
        layout = QtWidgets.QVBoxLayout(self)

        # Instructions
        instructions = QtWidgets.QLabel(
            "Before importing, you need to create a Bid to contain the data. "
            "Please enter a name for the new Bid:"
        )
        instructions.setWordWrap(True)
        instructions.setStyleSheet("padding: 10px; background-color: #2b2b2b; border-radius: 4px;")
        layout.addWidget(instructions)

        # Name field
        name_layout = QtWidgets.QHBoxLayout()
        name_label = QtWidgets.QLabel("Bid Name:")
        name_layout.addWidget(name_label)

        self.name_field = QtWidgets.QLineEdit()
        self.name_field.setPlaceholderText("Enter bid name...")
        name_layout.addWidget(self.name_field, stretch=1)

        layout.addLayout(name_layout)

        # Bid type selection
        type_layout = QtWidgets.QHBoxLayout()
        type_label = QtWidgets.QLabel("Bid Type:")
        type_layout.addWidget(type_label)

        self.type_combo = QtWidgets.QComboBox()
        self.type_combo.addItem("Early Bid")
        self.type_combo.addItem("Turnover Bid")
        type_layout.addWidget(self.type_combo, stretch=1)

        layout.addLayout(type_layout)

        # Buttons
        button_layout = QtWidgets.QHBoxLayout()
        button_layout.addStretch()

        self.ok_button = QtWidgets.QPushButton("Create")
        self.ok_button.clicked.connect(self.accept)
        self.ok_button.setDefault(True)
        button_layout.addWidget(self.ok_button)

        self.cancel_button = QtWidgets.QPushButton("Cancel")
        self.cancel_button.clicked.connect(self.reject)
        button_layout.addWidget(self.cancel_button)

        layout.addLayout(button_layout)

    def get_bid_name(self):
        """Get the bid name from the dialog."""
        return self.name_field.text().strip()

    def get_bid_type(self):
        """Get the selected bid type."""
        return self.type_combo.currentText()


class CreateVFXBreakdownDialog(QtWidgets.QDialog):
    """Simple dialog for creating a new VFX Breakdown during import."""

    def __init__(self, parent=None):
        """Initialize the dialog."""
        super().__init__(parent)
        self.setWindowTitle("Create VFX Breakdown")
        self.setModal(True)
        self.setMinimumWidth(400)

        self._build_ui()

    def _build_ui(self):
        """Build the dialog UI."""
        layout = QtWidgets.QVBoxLayout(self)

        # Instructions
        instructions = QtWidgets.QLabel(
            "Before importing breakdown items, you need to create a VFX Breakdown "
            "to contain them. Please enter a name for the new VFX Breakdown:"
        )
        instructions.setWordWrap(True)
        instructions.setStyleSheet("padding: 10px; background-color: #2b2b2b; border-radius: 4px;")
        layout.addWidget(instructions)

        # Name field
        name_layout = QtWidgets.QHBoxLayout()
        name_label = QtWidgets.QLabel("Name:")
        name_layout.addWidget(name_label)

        self.name_field = QtWidgets.QLineEdit()
        self.name_field.setPlaceholderText("Enter VFX Breakdown name...")
        name_layout.addWidget(self.name_field, stretch=1)

        layout.addLayout(name_layout)

        # Buttons
        button_layout = QtWidgets.QHBoxLayout()
        button_layout.addStretch()

        self.ok_button = QtWidgets.QPushButton("Create")
        self.ok_button.clicked.connect(self.accept)
        self.ok_button.setDefault(True)
        button_layout.addWidget(self.ok_button)

        self.cancel_button = QtWidgets.QPushButton("Cancel")
        self.cancel_button.clicked.connect(self.reject)
        button_layout.addWidget(self.cancel_button)

        layout.addLayout(button_layout)

    def get_breakdown_name(self):
        """Get the VFX Breakdown name from the dialog."""
        return self.name_field.text().strip()


class SelectBidDialog(QtWidgets.QDialog):
    """Dialog for selecting Bid and entity types to import."""

    # Custom return code for Back button
    BACK_CLICKED = QtWidgets.QDialog.Rejected + 1

    def __init__(self, existing_bids, available_sheets, parent=None):
        """Initialize the dialog.

        Args:
            existing_bids: List of existing bids [{"id": 1, "code": "BidName"}, ...]
            available_sheets: Dict of sheet names that have data {"breakdown": True, "assets": False, ...}
            parent: Parent widget
        """
        super().__init__(parent)
        self.setWindowTitle("Select Bid and Import Options")
        self.setModal(True)
        self.setMinimumWidth(500)

        self.existing_bids = existing_bids
        self.available_sheets = available_sheets

        self._build_ui()

    def _build_ui(self):
        """Build the dialog UI."""
        layout = QtWidgets.QVBoxLayout(self)

        # Instructions
        instructions = QtWidgets.QLabel(
            "Select whether to add data to an existing Bid or create a new one, "
            "then choose which entity types to import."
        )
        instructions.setWordWrap(True)
        instructions.setStyleSheet("padding: 10px; background-color: #2b2b2b; border-radius: 4px;")
        layout.addWidget(instructions)

        # Bid selection section
        bid_group = QtWidgets.QGroupBox("Bid Selection")
        bid_layout = QtWidgets.QVBoxLayout()

        # Existing bid option
        self.existing_bid_radio = QtWidgets.QRadioButton("Add to existing Bid")
        self.existing_bid_radio.setChecked(True)
        self.existing_bid_radio.toggled.connect(self._on_bid_option_changed)
        bid_layout.addWidget(self.existing_bid_radio)

        existing_bid_container = QtWidgets.QHBoxLayout()
        existing_bid_container.addSpacing(30)
        bid_label = QtWidgets.QLabel("Select Bid:")
        existing_bid_container.addWidget(bid_label)

        self.bid_combo = QtWidgets.QComboBox()
        self.bid_combo.addItem("-- Select Bid --", None)
        for bid in self.existing_bids:
            self.bid_combo.addItem(bid.get("code", "Unknown"), bid.get("id"))
        existing_bid_container.addWidget(self.bid_combo, stretch=1)
        bid_layout.addLayout(existing_bid_container)

        bid_layout.addSpacing(10)

        # New bid option
        self.new_bid_radio = QtWidgets.QRadioButton("Create new Bid")
        self.new_bid_radio.toggled.connect(self._on_bid_option_changed)
        bid_layout.addWidget(self.new_bid_radio)

        new_bid_container = QtWidgets.QVBoxLayout()
        new_bid_container.setContentsMargins(30, 0, 0, 0)

        name_layout = QtWidgets.QHBoxLayout()
        name_label = QtWidgets.QLabel("Bid Name:")
        name_layout.addWidget(name_label)
        self.name_field = QtWidgets.QLineEdit()
        self.name_field.setPlaceholderText("Enter bid name...")
        name_layout.addWidget(self.name_field, stretch=1)
        new_bid_container.addLayout(name_layout)

        type_layout = QtWidgets.QHBoxLayout()
        type_label = QtWidgets.QLabel("Bid Type:")
        type_layout.addWidget(type_label)
        self.type_combo = QtWidgets.QComboBox()
        self.type_combo.addItem("Early Bid")
        self.type_combo.addItem("Turnover Bid")
        type_layout.addWidget(self.type_combo, stretch=1)
        new_bid_container.addLayout(type_layout)

        bid_layout.addLayout(new_bid_container)

        bid_group.setLayout(bid_layout)
        layout.addWidget(bid_group)

        # Entity selection section
        entity_group = QtWidgets.QGroupBox("What to Import")
        entity_layout = QtWidgets.QVBoxLayout()

        entity_instructions = QtWidgets.QLabel(
            "Select which entity types to import from the Excel sheets:"
        )
        entity_instructions.setWordWrap(True)
        entity_layout.addWidget(entity_instructions)

        entity_layout.addSpacing(5)

        # Get DPI scale factor from parent's app_settings if available
        dpi_scale = 1.0
        if self.parent() and hasattr(self.parent(), 'app_settings'):
            try:
                dpi_scale = self.parent().app_settings.get_dpi_scale()
            except:
                dpi_scale = 1.0

        # Scale checkbox dimensions
        indicator_size = int(20 * dpi_scale)
        border_radius = max(2, int(3 * dpi_scale))
        border_width = max(1, int(2 * dpi_scale))
        spacing = int(8 * dpi_scale)
        checkmark_size = max(10, int(16 * dpi_scale))

        # Checkbox stylesheet to match VFX Breakdown table style
        checkbox_style = f"""
            QCheckBox {{
                spacing: {spacing}px;
            }}
            QCheckBox::indicator {{
                width: {indicator_size}px;
                height: {indicator_size}px;
                border-radius: {border_radius}px;
                background-color: #2b2b2b;
                border: {border_width}px solid #555555;
            }}
            QCheckBox::indicator:checked {{
                border: {border_width}px solid #0078d4;
                background-color: #2b2b2b;
                image: none;
            }}
            QCheckBox::indicator:disabled {{
                background-color: #1a1a1a;
                border-color: #333333;
            }}
        """

        # Create a custom paint function for checkmarks
        def create_checkbox_with_checkmark(text):
            """Create a checkbox with custom checkmark rendering."""
            checkbox = QtWidgets.QCheckBox(text)
            checkbox.setStyleSheet(checkbox_style)

            # Override paint event to draw custom checkmark
            original_paint = checkbox.paintEvent

            def custom_paint(event):
                original_paint(event)
                if checkbox.isChecked():
                    painter = QtGui.QPainter(checkbox)
                    painter.setRenderHint(QtGui.QPainter.Antialiasing)

                    # Get indicator rect (scaled size, positioned at left side)
                    indicator_rect = QtCore.QRect(0, (checkbox.height() - indicator_size) // 2, indicator_size, indicator_size)

                    # Draw checkmark
                    painter.setPen(QtGui.QPen(QtGui.QColor("#0078d4"), border_width))
                    font = painter.font()
                    font.setPixelSize(checkmark_size)
                    font.setBold(True)
                    painter.setFont(font)
                    painter.drawText(indicator_rect, QtCore.Qt.AlignCenter, "✓")
                    painter.end()

            checkbox.paintEvent = custom_paint
            return checkbox

        self.breakdown_checkbox = create_checkbox_with_checkmark("Breakdown")
        self.breakdown_checkbox.setChecked(self.available_sheets.get("breakdown", False))
        self.breakdown_checkbox.setEnabled(self.available_sheets.get("breakdown", False))
        entity_layout.addWidget(self.breakdown_checkbox)

        self.assets_checkbox = create_checkbox_with_checkmark("Assets")
        self.assets_checkbox.setChecked(self.available_sheets.get("assets", False))
        self.assets_checkbox.setEnabled(self.available_sheets.get("assets", False))
        entity_layout.addWidget(self.assets_checkbox)

        self.scenes_checkbox = create_checkbox_with_checkmark("Scenes")
        self.scenes_checkbox.setChecked(self.available_sheets.get("scenes", False))
        self.scenes_checkbox.setEnabled(self.available_sheets.get("scenes", False))
        entity_layout.addWidget(self.scenes_checkbox)

        self.rates_checkbox = create_checkbox_with_checkmark("Rates")
        self.rates_checkbox.setChecked(self.available_sheets.get("rates", False))
        self.rates_checkbox.setEnabled(self.available_sheets.get("rates", False))
        entity_layout.addWidget(self.rates_checkbox)

        entity_group.setLayout(entity_layout)
        layout.addWidget(entity_group)

        # Buttons
        button_layout = QtWidgets.QHBoxLayout()

        self.back_button = QtWidgets.QPushButton("Back")
        self.back_button.clicked.connect(self._on_back_clicked)
        button_layout.addWidget(self.back_button)

        button_layout.addStretch()

        self.import_button = QtWidgets.QPushButton("Import")
        self.import_button.clicked.connect(self._on_import_clicked)
        self.import_button.setDefault(True)
        button_layout.addWidget(self.import_button)

        self.cancel_button = QtWidgets.QPushButton("Cancel")
        self.cancel_button.clicked.connect(self.reject)
        button_layout.addWidget(self.cancel_button)

        layout.addLayout(button_layout)

        # Initialize UI state
        self._on_bid_option_changed()

    def _on_bid_option_changed(self):
        """Handle bid option radio button changes."""
        is_existing = self.existing_bid_radio.isChecked()

        # Enable/disable existing bid controls
        self.bid_combo.setEnabled(is_existing)

        # Enable/disable new bid controls
        self.name_field.setEnabled(not is_existing)
        self.type_combo.setEnabled(not is_existing)

    def _on_back_clicked(self):
        """Handle Back button click."""
        self.done(self.BACK_CLICKED)

    def _on_import_clicked(self):
        """Handle Import button click with validation."""
        # Validate bid selection
        if self.existing_bid_radio.isChecked():
            if self.bid_combo.currentData() is None:
                QtWidgets.QMessageBox.warning(
                    self,
                    "No Bid Selected",
                    "Please select a Bid from the dropdown."
                )
                return
        else:
            bid_name = self.name_field.text().strip()
            if not bid_name:
                QtWidgets.QMessageBox.warning(
                    self,
                    "Invalid Name",
                    "Please enter a name for the new Bid."
                )
                return

        # Validate at least one entity type selected
        if not any([
            self.breakdown_checkbox.isChecked(),
            self.assets_checkbox.isChecked(),
            self.scenes_checkbox.isChecked(),
            self.rates_checkbox.isChecked()
        ]):
            QtWidgets.QMessageBox.warning(
                self,
                "No Entity Types Selected",
                "Please select at least one entity type to import."
            )
            return

        # All validation passed
        self.accept()

    def is_creating_new_bid(self):
        """Check if creating a new bid."""
        return self.new_bid_radio.isChecked()

    def get_selected_bid_id(self):
        """Get the selected existing bid ID."""
        return self.bid_combo.currentData()

    def get_new_bid_name(self):
        """Get the new bid name."""
        return self.name_field.text().strip()

    def get_new_bid_type(self):
        """Get the new bid type."""
        return self.type_combo.currentText()

    def get_selected_entity_types(self):
        """Get dict of selected entity types."""
        return {
            "breakdown": self.breakdown_checkbox.isChecked(),
            "assets": self.assets_checkbox.isChecked(),
            "scenes": self.scenes_checkbox.isChecked(),
            "rates": self.rates_checkbox.isChecked()
        }


class ColumnMappingDialog(QtWidgets.QDialog):
    """Dialog for mapping Excel columns to ShotGrid fields."""

    # Custom return code for Back button
    BACK_CLICKED = QtWidgets.QDialog.Rejected + 1

    # ShotGrid field definitions for VFX Breakdown (CustomEntity02 - Bidding Scenes)
    BREAKDOWN_ITEM_REQUIRED_FIELDS = {
        "code": "text",
        "sg_bid_assets": "entity",
        "sg_sequence_code": "text",
        "sg_interior_exterior": "list",
        "sg_number_of_shots": "number",
        "sg_on_set_vfx_needs": "text",
        "sg_page_eights": "text",
        "sg_previs": "checkbox",
        "sg_script_excerpt": "text",
        "sg_set": "text",
        "sg_sim": "checkbox",
        "sg_sorting_priority": "float",
        "sg_team_notes": "text",
        "sg_time_of_day": "list",
        "sg_unit": "text",
        "sg_vfx_assumptions": "text",
        "sg_vfx_breakdown_scene": "text",
        "sg_vfx_description": "text",
        "sg_vfx_questions": "text",
        "sg_vfx_supervisor_notes": "text",
        "sg_vfx_type": "text",
    }

    # ShotGrid field definitions for Assets (CustomEntity07 - Asset items)
    ASSET_ITEM_REQUIRED_FIELDS = {
        "code": "text",
        "sg_bid_asset_type": "list",
        "sg_bidding_notes": "text",
    }

    # ShotGrid field definitions for Scenes (placeholder - customize as needed)
    SCENE_REQUIRED_FIELDS = {
        "code": "text",
        "description": "text",
    }

    # ShotGrid field definitions for Line Items (CustomEntity03)
    # Note: Actual import auto-discovers all fields ending with "_mandays"
    # These are common fields shown in the column mapping dialog
    RATE_REQUIRED_FIELDS = {
        "code": "text",
        "sg_entity": "entity",
        "sg_type": "text",
        "sg_complexity_price": "float",
        "sg_model_mandays": "float",
        "sg_tex_mandays": "float",
        "sg_lookdev_mandays": "float",
        "sg_rig_mandays": "float",
        "sg_mm_mandays": "float",
        "sg_prep_mandays": "float",
        "sg_gen_mandays": "float",
        "sg_anim_mandays": "float",
        "sg_lgt_mandays": "float",
        "sg_fx_mandays": "float",
        "sg_cmp_mandays": "float",
    }

    # Entity type mapping
    ENTITY_CONFIGS = {
        "breakdown": {
            "name": "Breakdown",
            "entity_type": "CustomEntity02",
            "fields": BREAKDOWN_ITEM_REQUIRED_FIELDS,
            "mapping_key": "vfx_breakdown"
        },
        "assets": {
            "name": "Assets",
            "entity_type": "CustomEntity07",
            "fields": ASSET_ITEM_REQUIRED_FIELDS,
            "mapping_key": "assets"
        },
        "scenes": {
            "name": "Scenes",
            "entity_type": "Scene",
            "fields": SCENE_REQUIRED_FIELDS,
            "mapping_key": "scenes"
        },
        "rates": {
            "name": "Rates",
            "entity_type": "CustomEntity03",  # Line Items
            "fields": RATE_REQUIRED_FIELDS,
            "mapping_key": "rates"
        }
    }

    def __init__(self, excel_columns, sg_session, project_id, parent=None):
        """Initialize the column mapping dialog.

        Args:
            excel_columns: Either a list of column headers (uses same columns for all tabs)
                          or a dict mapping entity type to column list {"breakdown": [...], "assets": [...]}
            sg_session: ShotGrid session for API access
            project_id: Project ID for field schema lookup
            parent: Parent widget
        """
        super().__init__(parent)
        self.setWindowTitle("Map Columns to ShotGrid Fields")
        self.setModal(True)
        self.setMinimumSize(900, 700)

        # Convert excel_columns to dict format if it's a list (backward compatibility)
        if isinstance(excel_columns, list):
            # Use same columns for all entity types
            self.excel_columns_dict = {
                "breakdown": excel_columns,
                "assets": excel_columns,
                "scenes": excel_columns,
                "rates": excel_columns
            }
        else:
            # Already a dict, use as-is (with defaults for missing keys)
            self.excel_columns_dict = {}
            for config_key in self.ENTITY_CONFIGS.keys():
                self.excel_columns_dict[config_key] = excel_columns.get(config_key, [])

        self.sg_session = sg_session
        self.project_id = project_id

        # Storage for mapping widgets - keyed by entity config key (breakdown, assets, etc.)
        self.mapping_combos = {}  # config_key -> {sg_field -> {"excel": combo, "sg": combo}}
        self.sg_display_names = {}  # config_key -> {sg_field -> display_name}

        # App settings for saving/loading mappings
        self.app_settings = AppSettings()

        self._fetch_all_display_names()
        self._build_ui()
        self._load_all_saved_mappings()
        self._auto_map_all_columns()

    def _fetch_all_display_names(self):
        """Fetch human-readable display names for all entity types."""
        for config_key, config in self.ENTITY_CONFIGS.items():
            entity_type = config["entity_type"]
            field_names = list(config["fields"].keys())

            try:
                display_names = self._get_field_display_names(entity_type, field_names, self.project_id)
                self.sg_display_names[config_key] = display_names
                logger.info(f"Fetched {len(display_names)} display names for {config['name']}")
            except Exception as e:
                logger.error(f"Error fetching display names for {config['name']}: {e}", exc_info=True)
                # Fallback to field names
                self.sg_display_names[config_key] = {field: field for field in field_names}

    def _get_field_display_names(self, entity_name, field_names, project_id=None):
        """Get display names for multiple fields of an entity.

        Args:
            entity_name (str): The entity type
            field_names (list): List of field names
            project_id (int, optional): The project ID for project-specific fields

        Returns:
            dict: Dictionary with field_name: display_name pairs
        """
        try:
            # Get ALL fields for the entity
            if project_id:
                fields = self.sg_session.sg.schema_field_read(
                    entity_name,
                    project_entity={'type': 'Project', 'id': project_id}
                )
            else:
                fields = self.sg_session.sg.schema_field_read(entity_name)

            # Build the result dictionary with only the requested fields
            result = {}
            for field_name in field_names:
                if field_name in fields:
                    result[field_name] = fields[field_name]['name']['value']
                else:
                    result[field_name] = field_name  # Fallback to field name

            return result

        except Exception as e:
            logger.error(f"Error retrieving field display names: {e}", exc_info=True)
            return {field: field for field in field_names}

    def _build_ui(self):
        """Build the mapping dialog UI with tabs."""
        layout = QtWidgets.QVBoxLayout(self)

        # Instructions
        instructions = QtWidgets.QLabel(
            "Map Excel columns to ShotGrid fields. Use the dropdowns to select which "
            "Excel column corresponds to each ShotGrid field. Leave unmapped if the column "
            "doesn't exist in your Excel file."
        )
        instructions.setWordWrap(True)
        instructions.setStyleSheet("padding: 10px; background-color: #2b2b2b; border-radius: 4px;")
        layout.addWidget(instructions)

        # Create tab widget
        self.tab_widget = QtWidgets.QTabWidget()

        # Create a tab for each entity type
        for config_key, config in self.ENTITY_CONFIGS.items():
            tab = self._create_mapping_tab(config_key, config)
            self.tab_widget.addTab(tab, config["name"])

        layout.addWidget(self.tab_widget, stretch=1)

        # Buttons
        button_layout = QtWidgets.QHBoxLayout()

        self.back_button = QtWidgets.QPushButton("Back")
        self.back_button.clicked.connect(self._on_back_clicked)
        button_layout.addWidget(self.back_button)

        button_layout.addStretch()

        self.next_button = QtWidgets.QPushButton("Next")
        self.next_button.clicked.connect(self.accept)
        self.next_button.setDefault(True)
        button_layout.addWidget(self.next_button)

        self.cancel_button = QtWidgets.QPushButton("Cancel")
        self.cancel_button.clicked.connect(self.reject)
        button_layout.addWidget(self.cancel_button)

        layout.addLayout(button_layout)

    def _on_back_clicked(self):
        """Handle Back button click."""
        self.done(self.BACK_CLICKED)

    def _create_mapping_tab(self, config_key, config):
        """Create a tab for a specific entity type.

        Args:
            config_key: Key for this entity configuration (e.g., "breakdown", "assets")
            config: Entity configuration dictionary

        Returns:
            QWidget: Tab widget
        """
        tab_widget = QtWidgets.QWidget()
        tab_layout = QtWidgets.QVBoxLayout(tab_widget)

        # Scroll area for mappings
        scroll = QtWidgets.QScrollArea()
        scroll.setWidgetResizable(True)
        scroll_widget = QtWidgets.QWidget()
        scroll_layout = QtWidgets.QVBoxLayout(scroll_widget)

        # Initialize mapping combos storage for this entity type
        self.mapping_combos[config_key] = {}

        # Create mapping rows for this entity's fields
        for sg_field, field_type in config["fields"].items():
            row = self._create_mapping_row(config_key, sg_field, field_type)
            scroll_layout.addLayout(row)

        scroll_layout.addStretch()
        scroll.setWidget(scroll_widget)
        tab_layout.addWidget(scroll)

        return tab_widget

    def _create_mapping_row(self, config_key, sg_field, field_type):
        """Create a mapping row with two dropdowns.

        Args:
            config_key: Entity configuration key (e.g., "breakdown", "assets")
            sg_field: ShotGrid field name
            field_type: Field type (text, list, number, etc.)

        Returns:
            QHBoxLayout: Layout for the mapping row
        """
        row_layout = QtWidgets.QHBoxLayout()

        # Excel column dropdown (left) - use columns specific to this entity type
        excel_combo = QtWidgets.QComboBox()
        excel_combo.addItem("-- Not Mapped --", None)
        entity_columns = self.excel_columns_dict.get(config_key, [])
        for col in entity_columns:
            excel_combo.addItem(col)
        excel_combo.setMinimumWidth(200)
        row_layout.addWidget(excel_combo)

        # Arrow
        arrow_label = QtWidgets.QLabel("→")
        arrow_label.setStyleSheet("font-size: 16px; padding: 0 10px;")
        row_layout.addWidget(arrow_label)

        # SG field dropdown (right) - read-only display
        display_names = self.sg_display_names.get(config_key, {})
        sg_display_name = display_names.get(sg_field, sg_field)
        sg_label = QtWidgets.QLabel(f"{sg_display_name}")
        sg_label.setStyleSheet("font-weight: bold;")
        sg_label.setMinimumWidth(200)
        row_layout.addWidget(sg_label)

        # Field type indicator
        type_label = QtWidgets.QLabel(f"({field_type})")
        type_label.setStyleSheet("color: #888888; font-size: 10px;")
        row_layout.addWidget(type_label)

        row_layout.addStretch()

        # Store reference
        self.mapping_combos[config_key][sg_field] = {
            "excel": excel_combo,
            "sg_display": sg_display_name
        }

        return row_layout

    def _auto_map_all_columns(self):
        """Automatically map Excel columns to SG fields for all entity types."""
        for config_key, config in self.ENTITY_CONFIGS.items():
            self._auto_map_columns_for_entity(config_key, config)

    def _auto_map_columns_for_entity(self, config_key, config):
        """Automatically map Excel columns to SG fields using fuzzy matching.

        Only maps fields that don't already have a mapping (i.e., saved mappings take precedence).

        Args:
            config_key: Entity configuration key
            config: Entity configuration dictionary
        """
        entity_mappings = self.mapping_combos.get(config_key, {})

        for sg_field in config["fields"].keys():
            if sg_field not in entity_mappings:
                continue

            combo = entity_mappings[sg_field]["excel"]

            # Skip if this field already has a mapping (from saved settings)
            current_text = combo.currentText()
            if current_text and current_text != "-- Not Mapped --":
                logger.info(f"Skipping auto-map for '{config_key}.{sg_field}' - already mapped to '{current_text}'")
                continue

            # No saved mapping, try fuzzy matching
            best_match = self._find_best_column_match(config_key, sg_field)
            if best_match:
                index = combo.findText(best_match)
                if index >= 0:
                    combo.setCurrentIndex(index)
                    logger.info(f"Auto-mapped '{best_match}' -> '{config_key}.{sg_field}'")

    def _find_best_column_match(self, config_key, sg_field):
        """Find the best matching Excel column for a SG field using fuzzy logic.

        Args:
            config_key: Entity configuration key
            sg_field: ShotGrid field name

        Returns:
            str: Best matching Excel column name or None
        """
        # Get entity-specific columns
        entity_columns = self.excel_columns_dict.get(config_key, [])
        if not entity_columns:
            return None

        # Get display name for comparison
        display_names = self.sg_display_names.get(config_key, {})
        sg_display_name = display_names.get(sg_field, sg_field)

        # Clean field name (remove sg_ prefix)
        sg_clean = sg_field.replace("sg_", "").replace("_", " ").lower()
        sg_display_clean = sg_display_name.replace("_", " ").lower()

        best_match = None
        best_score = 0

        for excel_col in entity_columns:
            excel_clean = excel_col.replace("_", " ").lower()
            score = 0

            # Exact match (case insensitive)
            if excel_clean == sg_clean or excel_clean == sg_display_clean:
                return excel_col

            # Check if field name is in column or vice versa
            if sg_clean in excel_clean:
                score += len(sg_clean) * 2
            elif excel_clean in sg_clean:
                score += len(excel_clean) * 2

            # Check display name match
            if sg_display_clean in excel_clean:
                score += len(sg_display_clean) * 2
            elif excel_clean in sg_display_clean:
                score += len(excel_clean) * 2

            # Word matching
            sg_words = sg_clean.split()
            excel_words = excel_clean.split()

            for sg_word in sg_words:
                for excel_word in excel_words:
                    if sg_word == excel_word:
                        score += len(sg_word) * 3
                    elif sg_word in excel_word or excel_word in sg_word:
                        score += len(min(sg_word, excel_word, key=len))

            if score > best_score:
                best_score = score
                best_match = excel_col

        # Only return if score is meaningful
        return best_match if best_score > 3 else None

    def _load_all_saved_mappings(self):
        """Load previously saved column mappings from settings for all entity types."""
        for config_key, config in self.ENTITY_CONFIGS.items():
            self._load_saved_mappings_for_entity(config_key, config)

    def _load_saved_mappings_for_entity(self, config_key, config):
        """Load previously saved column mappings from settings.

        Args:
            config_key: Entity configuration key
            config: Entity configuration dictionary
        """
        mapping_key = config["mapping_key"]
        saved_mapping = self.app_settings.get_column_mapping(mapping_key)

        if not saved_mapping:
            logger.info(f"No saved column mappings found for '{config_key}'")
            return

        # Apply saved mappings (only if column exists in current Excel file)
        applied_count = 0
        entity_mappings = self.mapping_combos.get(config_key, {})
        entity_columns = self.excel_columns_dict.get(config_key, [])

        for sg_field, excel_col in saved_mapping.items():
            if sg_field not in entity_mappings:
                continue

            if excel_col and excel_col in entity_columns:
                combo = entity_mappings[sg_field]["excel"]
                index = combo.findText(excel_col)
                if index >= 0:
                    combo.setCurrentIndex(index)
                    applied_count += 1
                    logger.info(f"Applied saved mapping for '{config_key}': '{excel_col}' -> '{sg_field}'")

        logger.info(f"Applied {applied_count} saved column mappings for '{config_key}'")

    def accept(self):
        """Override accept to save mappings before closing."""
        # Save mappings only for entity types that have columns
        # This prevents overwriting existing saved mappings for tabs that weren't used
        for config_key, config in self.ENTITY_CONFIGS.items():
            entity_columns = self.excel_columns_dict.get(config_key, [])

            # Only save if this entity type has columns (i.e., was actually used)
            if entity_columns:
                mapping = self.get_column_mapping_for_entity(config_key)
                mapping_key = config["mapping_key"]
                self.app_settings.set_column_mapping(mapping_key, mapping)
                logger.info(f"Saved column mappings for '{config_key}' with key '{mapping_key}'")
            else:
                logger.info(f"Skipping save for '{config_key}' - no columns provided")

        # Call parent accept
        super().accept()

    def get_column_mapping(self):
        """Get the column mapping results for the Breakdown tab (for backward compatibility).

        Returns:
            dict: Mapping of sg_field -> excel_column_name (or None if not mapped)
        """
        return self.get_column_mapping_for_entity("breakdown")

    def get_column_mapping_for_entity(self, config_key):
        """Get the column mapping results for a specific entity type.

        Args:
            config_key: Entity configuration key (e.g., "breakdown", "assets")

        Returns:
            dict: Mapping of sg_field -> excel_column_name (or None if not mapped)
        """
        mapping = {}
        entity_mappings = self.mapping_combos.get(config_key, {})

        for sg_field, widgets in entity_mappings.items():
            excel_combo = widgets["excel"]
            excel_col = excel_combo.currentText()
            if excel_col and excel_col != "-- Not Mapped --":
                mapping[sg_field] = excel_col
            else:
                mapping[sg_field] = None
        return mapping

    def get_all_column_mappings(self):
        """Get all column mappings for all entity types.

        Returns:
            dict: Mapping of config_key -> {sg_field -> excel_column_name}
        """
        all_mappings = {}
        for config_key in self.ENTITY_CONFIGS.keys():
            all_mappings[config_key] = self.get_column_mapping_for_entity(config_key)
        return all_mappings


class ImportBidDialog(QtWidgets.QDialog):
    """Dialog for importing bid data from an Excel file with tabs for different data types."""

    def __init__(self, sg_session, project_id, parent=None):
        """Initialize the dialog.

        Args:
            sg_session: ShotGrid session for API access
            project_id: Project ID for field schema lookup and record creation
            parent: Parent widget
        """
        super().__init__(parent)
        self.setWindowTitle("Import Bid Data")
        self.setModal(True)
        self.setMinimumSize(900, 700)

        # Enable maximize button
        self.setWindowFlags(self.windowFlags() | QtCore.Qt.WindowMaximizeButtonHint)

        # ShotGrid connection
        self.sg_session = sg_session
        self.project_id = project_id

        # Data storage
        self.excel_file_path = None
        self.sheet_names = []

        # Tab configuration: tab_name -> (sheet_match_keywords, data)
        self.tab_config = {
            "VFX Breakdown": (["vfx", "breakdown", "break"], None),
            "Assets": (["asset", "assets"], None),
            "Scene": (["scene", "scenes"], None),
            "Rates": (["rate", "rates", "pricing", "price"], None)
        }

        # UI components for each tab
        self.tab_widgets = {}  # tab_name -> {"combo": combo, "table": table}

        self._build_ui()

    def _build_ui(self):
        """Build the dialog UI."""
        layout = QtWidgets.QVBoxLayout(self)

        # Drag and drop area
        self.drop_area = DragDropArea()
        self.drop_area.fileDropped.connect(self._on_file_dropped)
        layout.addWidget(self.drop_area)

        # Tab widget (initially hidden)
        self.tab_widget = QtWidgets.QTabWidget()
        self.tab_widget.hide()

        # Create tabs
        for tab_name in self.tab_config.keys():
            tab = self._create_tab(tab_name)
            self.tab_widget.addTab(tab, tab_name)

        layout.addWidget(self.tab_widget, stretch=1)

        # Status label
        self.status_label = QtWidgets.QLabel("")
        self.status_label.setStyleSheet("color: #a0a0a0; padding: 5px;")
        layout.addWidget(self.status_label)

        # Buttons
        button_layout = QtWidgets.QHBoxLayout()
        button_layout.addStretch()

        self.next_button = QtWidgets.QPushButton("Next")
        self.next_button.setMinimumHeight(40)
        self.next_button.setMinimumWidth(120)
        self.next_button.setStyleSheet("""
            QPushButton {
                font-size: 14px;
                font-weight: bold;
            }
        """)
        self.next_button.clicked.connect(self.accept)
        self.next_button.setEnabled(False)
        button_layout.addWidget(self.next_button)

        self.cancel_button = QtWidgets.QPushButton("Cancel")
        self.cancel_button.clicked.connect(self.reject)
        button_layout.addWidget(self.cancel_button)

        layout.addLayout(button_layout)

    def _create_tab(self, tab_name):
        """Create a tab with sheet selector and table."""
        tab_widget = QtWidgets.QWidget()
        tab_layout = QtWidgets.QVBoxLayout(tab_widget)

        # Sheet selection row
        sheet_layout = QtWidgets.QHBoxLayout()
        sheet_label = QtWidgets.QLabel("Select Sheet:")
        sheet_layout.addWidget(sheet_label)

        sheet_combo = QtWidgets.QComboBox()
        sheet_combo.currentIndexChanged.connect(lambda idx, tn=tab_name: self._on_sheet_changed(tn, idx))
        sheet_layout.addWidget(sheet_combo, stretch=1)

        # Row selection buttons
        select_all_btn = QtWidgets.QPushButton("Select All")
        select_all_btn.clicked.connect(lambda checked, tn=tab_name: self._select_all_rows(tn, True))
        sheet_layout.addWidget(select_all_btn)

        deselect_all_btn = QtWidgets.QPushButton("Deselect All")
        deselect_all_btn.clicked.connect(lambda checked, tn=tab_name: self._select_all_rows(tn, False))
        sheet_layout.addWidget(deselect_all_btn)

        tab_layout.addLayout(sheet_layout)

        # Table for displaying data
        table = QtWidgets.QTableWidget()
        table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectItems)
        table.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        table.setAlternatingRowColors(False)
        table.setWordWrap(True)
        table.horizontalHeader().setStretchLastSection(False)
        table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Interactive)
        table.verticalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Interactive)
        table.itemChanged.connect(lambda item, tn=tab_name: self._on_table_item_changed(tn, item))

        tab_layout.addWidget(table, stretch=1)

        # Store references
        self.tab_widgets[tab_name] = {
            "combo": sheet_combo,
            "table": table
        }

        return tab_widget

    def _find_best_match_sheet(self, tab_name):
        """Find the best matching sheet name for a tab using fuzzy matching.

        Args:
            tab_name: The tab name to match against

        Returns:
            str: Best matching sheet name or None
        """
        if not self.sheet_names:
            return None

        keywords, _ = self.tab_config[tab_name]

        # Try exact match first (case insensitive)
        tab_lower = tab_name.lower()
        for sheet in self.sheet_names:
            sheet_lower = sheet.lower()
            if sheet_lower == tab_lower:
                return sheet

        # Try keyword matching
        best_match = None
        best_score = 0

        for sheet in self.sheet_names:
            sheet_lower = sheet.lower()
            score = 0

            # Check if any keyword is in the sheet name
            for keyword in keywords:
                if keyword in sheet_lower:
                    score += len(keyword)

            # Check if sheet name is in tab name
            if sheet_lower in tab_lower or tab_lower in sheet_lower:
                score += 10

            if score > best_score:
                best_score = score
                best_match = sheet

        return best_match if best_score > 0 else None

    def _on_file_dropped(self, file_path):
        """Handle file drop event."""
        if not file_path.lower().endswith(('.xlsx', '.xls')):
            self._set_status("Error: Please drop an Excel file (.xlsx or .xls)", is_error=True)
            return

        self.excel_file_path = file_path
        self._set_status(f"Loading file: {file_path}")

        try:
            # Try to import pandas
            import pandas as pd

            # Create progress dialog
            progress = QtWidgets.QProgressDialog("Loading Excel file...", None, 0, 100, self)
            progress.setWindowModality(QtCore.Qt.WindowModal)
            progress.setMinimumDuration(0)
            progress.setValue(0)
            QtWidgets.QApplication.processEvents()

            # Read Excel file to get sheet names
            progress.setLabelText("Reading Excel file...")
            progress.setValue(10)
            QtWidgets.QApplication.processEvents()

            # Use openpyxl to read sheet names (consistent with how we read data)
            from openpyxl import load_workbook
            wb = load_workbook(file_path, data_only=False, read_only=True)
            self.sheet_names = wb.sheetnames
            wb.close()

            progress.setValue(20)
            QtWidgets.QApplication.processEvents()

            # Populate all sheet combos and auto-select best matches
            tabs_to_load = []
            progress.setLabelText("Setting up sheet selections...")
            for tab_name, widgets in self.tab_widgets.items():
                combo = widgets["combo"]
                combo.blockSignals(True)
                combo.clear()
                combo.addItem("-- Select Sheet --", None)
                for sheet in self.sheet_names:
                    combo.addItem(sheet)
                combo.blockSignals(False)

                # Auto-select best match
                best_match = self._find_best_match_sheet(tab_name)
                if best_match:
                    index = combo.findText(best_match)
                    if index >= 0:
                        combo.setCurrentIndex(index)
                        tabs_to_load.append((tab_name, best_match))

            progress.setValue(30)
            QtWidgets.QApplication.processEvents()

            # Load sheets with progress updates
            if tabs_to_load:
                progress_step = 70 / len(tabs_to_load)
                for i, (tab_name, sheet_name) in enumerate(tabs_to_load):
                    progress.setLabelText(f"Loading sheet: {sheet_name}...")
                    progress.setValue(30 + int(i * progress_step))
                    QtWidgets.QApplication.processEvents()
                    self._load_sheet_for_tab(tab_name, sheet_name)

            progress.setValue(100)
            QtWidgets.QApplication.processEvents()
            progress.close()

            # Show tabs and hide drop area
            self.tab_widget.show()
            self.drop_area.hide()
            self.next_button.setEnabled(True)

            # Auto-adjust dialog size to fit content
            self._adjust_dialog_size()

            self._set_status(f"Loaded: {file_path} ({len(self.sheet_names)} sheets)")

        except ImportError:
            self._set_status("Error: pandas library not installed. Please install with: pip install pandas openpyxl", is_error=True)
        except Exception as e:
            self._set_status(f"Error loading file: {str(e)}", is_error=True)
            logger.error(f"Error loading Excel file: {e}", exc_info=True)

    def _on_sheet_changed(self, tab_name, index):
        """Handle sheet selection change for a specific tab."""
        widgets = self.tab_widgets.get(tab_name)
        if not widgets:
            return

        combo = widgets["combo"]
        sheet_name = combo.currentText()

        if sheet_name and sheet_name != "-- Select Sheet --":
            self._load_sheet_for_tab(tab_name, sheet_name)

    def _load_sheet_for_tab(self, tab_name, sheet_name):
        """Load and display a specific sheet in a specific tab.

        Reads cells using openpyxl to get the exact displayed text value.
        """
        try:
            import pandas as pd
            from openpyxl import load_workbook
            from fractions import Fraction
            from datetime import datetime

            # Use openpyxl to read cells and preserve their display format
            wb = load_workbook(self.excel_file_path, data_only=True)

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

            ws = wb[sheet_name]

            # Extract all data as displayed text
            data = []
            headers = None
            row_idx = 0

            for row in ws.iter_rows(values_only=False):
                row_values = []
                col_idx = 0
                for cell in row:
                    if cell.value is None:
                        row_values.append("")
                    elif cell.data_type == 'b':  # Boolean
                        converted = "TRUE" if cell.value else "FALSE"
                        row_values.append(converted)
                        if row_idx > 0:  # Log data rows, not header
                            logger.info(f"Excel R{row_idx}C{col_idx}: Boolean type, value={cell.value} -> '{converted}'")
                    elif isinstance(cell.value, (int, float)):
                        # Numeric value - check if it has fraction formatting
                        if cell.number_format and '?' in cell.number_format and '/' in cell.number_format:
                            # It's a fraction format - reconstruct the fraction
                            try:
                                frac = Fraction(cell.value).limit_denominator(8)
                                whole = frac.numerator // frac.denominator
                                remainder = frac.numerator % frac.denominator
                                if whole > 0 and remainder > 0:
                                    row_values.append(f"{whole} {remainder}/{frac.denominator}")
                                elif remainder > 0:
                                    row_values.append(f"{remainder}/{frac.denominator}")
                                else:
                                    row_values.append(str(whole))
                            except:
                                row_values.append(str(cell.value))
                        elif cell.value in (0, 1):
                            # Could be a boolean represented as 0/1
                            # Keep as string "0" or "1" for later detection
                            converted = str(int(cell.value))
                            row_values.append(converted)
                            if row_idx > 0:
                                logger.info(f"Excel R{row_idx}C{col_idx}: Numeric 0/1, value={cell.value} -> '{converted}'")
                        else:
                            # Regular number - just convert to string
                            row_values.append(str(cell.value))
                    elif isinstance(cell.value, datetime):
                        # Date/datetime value - check if it might be a misinterpreted fraction
                        # Common date formats that might be fractions: m/d, m/d/yy, d/m, etc.
                        if cell.number_format and '/' in cell.number_format:
                            # Check if it's a simple date format without year (likely a fraction)
                            # Examples: "m/d", "d/m" (these are often misinterpreted fractions)
                            format_lower = cell.number_format.lower()
                            if ('y' not in format_lower and
                                format_lower.count('/') == 1 and
                                any(fmt in format_lower for fmt in ['m/d', 'd/m', 'm"/"d', 'd"/"m'])):
                                # This looks like a fraction that Excel interpreted as a date
                                # Format it as "month/day" to restore the original appearance
                                row_values.append(f"{cell.value.month}/{cell.value.day}")
                            else:
                                # It's a real date with year or more complex format
                                row_values.append(cell.value.strftime('%Y-%m-%d'))
                        else:
                            # No format info, treat as date
                            row_values.append(cell.value.strftime('%Y-%m-%d'))
                    else:
                        # Text or other - convert to string
                        row_values.append(str(cell.value))

                    col_idx += 1

                # First row is headers
                if headers is None:
                    # Create initial headers
                    headers = [str(v) if v else f"Column{i}" for i, v in enumerate(row_values)]

                    # Handle duplicate column names by making them unique
                    seen = {}
                    unique_headers = []
                    for i, header in enumerate(headers):
                        if header in seen:
                            # Duplicate found - append suffix
                            seen[header] += 1
                            new_header = f"{header}_{seen[header]}"
                            unique_headers.append(new_header)
                            logger.warning(f"Duplicate column name '{header}' at position {i}, renamed to '{new_header}'")
                        else:
                            seen[header] = 1
                            unique_headers.append(header)

                    headers = unique_headers
                    logger.info(f"Headers: {headers}")
                else:
                    # Only add rows that have at least one non-empty value
                    if any(v and v != "" and v != "None" for v in row_values):
                        data.append(row_values)

                row_idx += 1

            wb.close()

            # Create DataFrame
            try:
                df = pd.DataFrame(data, columns=headers)
                logger.info(f"DataFrame created successfully with {len(df)} rows and {len(df.columns)} columns")
            except Exception as e:
                error_msg = f"Failed to create DataFrame: {str(e)}"
                logger.error(error_msg)
                logger.error(f"Headers: {headers}")
                logger.error(f"Data rows: {len(data)}")
                raise ValueError(error_msg)

            # Log sample of DataFrame values for debugging
            for col_name in df.columns:
                unique_vals = df[col_name].dropna().unique()[:5]  # First 5 unique values
                logger.info(f"Column '{col_name}': sample values = {list(unique_vals)}")

            # Ensure all values are strings
            df = df.fillna("")
            for col in df.columns:
                df[col] = df[col].astype(str)

            # Store data
            keywords, _ = self.tab_config[tab_name]
            self.tab_config[tab_name] = (keywords, df)

            # Display in table
            widgets = self.tab_widgets[tab_name]
            self._populate_table(widgets["table"], df)

            logger.info(f"Loaded sheet '{sheet_name}' for tab '{tab_name}' with preserved formatting: {len(df)} rows, {len(df.columns)} columns")

        except Exception as e:
            self._set_status(f"Error loading sheet: {str(e)}", is_error=True)
            logger.error(f"Error loading sheet {sheet_name} for tab {tab_name}: {e}", exc_info=True)

    def _populate_table(self, table, df):
        """Populate a table widget with dataframe data, with checkbox column for selection.

        Boolean columns (TRUE/FALSE) are displayed as checkboxes.
        """
        import pandas as pd

        # Block signals while populating to avoid triggering itemChanged
        table.blockSignals(True)

        # Set dimensions - add 1 column for checkbox
        table.setRowCount(len(df))
        table.setColumnCount(len(df.columns) + 1)

        # Set headers - checkbox column first
        headers = ["Import"] + [str(col) for col in df.columns]
        table.setHorizontalHeaderLabels(headers)

        # Identify boolean columns (columns with boolean-like values)
        # Check ALL columns for TRUE/FALSE, 1/0, Yes/No, X/"" values to display as checkboxes
        boolean_columns = set()
        for j, col_name in enumerate(df.columns):
            # Check if all non-empty values in the column are boolean-like
            unique_values = df.iloc[:, j].dropna().unique()
            # Filter out empty strings and normalize values
            non_empty_values = [str(v).strip().upper() for v in unique_values if str(v).strip() != '']

            # Check if all values match common boolean representations
            if non_empty_values:
                is_boolean = all(v in ['TRUE', 'FALSE', '1', '0', 'YES', 'NO', 'Y', 'N', 'X'] for v in non_empty_values)
                if is_boolean:
                    boolean_columns.add(j)
                    logger.info(f"Column {j} ({col_name}) identified as boolean with values: {non_empty_values}")

        # Populate data
        for i in range(len(df)):
            # Add checkbox in first column (Import column)
            # This checkbox controls whether the row gets imported
            # Checked (default) = row will be imported
            # Unchecked = row will be skipped during import
            checkbox_item = QtWidgets.QTableWidgetItem()
            checkbox_item.setFlags(QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled)
            checkbox_item.setCheckState(QtCore.Qt.Checked)  # All rows checked by default
            table.setItem(i, 0, checkbox_item)

            # Add data in remaining columns
            for j in range(len(df.columns)):
                value = df.iloc[i, j]
                # Handle None/NaN values
                if pd.isna(value):
                    value = ""
                else:
                    value = str(value).strip()  # Strip whitespace

                # Check if this is a boolean column
                if j in boolean_columns:
                    # Create checkbox item for boolean columns
                    item = QtWidgets.QTableWidgetItem()
                    item.setFlags(QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled)

                    # Set checkbox state based on value (case-insensitive)
                    # Support multiple boolean representations
                    value_upper = value.upper() if value else ""
                    # TRUE values: TRUE, 1, YES, Y, X
                    if value_upper in ['TRUE', '1', 'YES', 'Y', 'X']:
                        item.setCheckState(QtCore.Qt.Checked)
                        logger.info(f"Row {i}, Col '{df.columns[j]}': Setting CHECKED for value='{value}' (upper='{value_upper}')")
                        # Verify immediately after setting
                        actual_state = item.checkState()
                        logger.info(f"  -> Verified: checkState()={actual_state}, Checked={QtCore.Qt.Checked}, Unchecked={QtCore.Qt.Unchecked}")
                    else:
                        # FALSE values: FALSE, 0, NO, N, empty
                        item.setCheckState(QtCore.Qt.Unchecked)
                        logger.info(f"Row {i}, Col '{df.columns[j]}': Setting UNCHECKED for value='{value}' (upper='{value_upper}')")
                        # Verify immediately after setting
                        actual_state = item.checkState()
                        logger.info(f"  -> Verified: checkState()={actual_state}, Checked={QtCore.Qt.Checked}, Unchecked={QtCore.Qt.Unchecked}")

                    # Store the original value as text for export
                    item.setData(QtCore.Qt.UserRole, value)
                else:
                    # Create text item
                    item = QtWidgets.QTableWidgetItem(value)
                    item.setFlags(item.flags() & ~QtCore.Qt.ItemIsEditable)  # Make read-only

                table.setItem(i, j + 1, item)

                # Verify checkbox state after adding to table (for boolean columns only)
                if j in boolean_columns:
                    state_after_add = table.item(i, j + 1).checkState()
                    logger.info(f"  -> After table.setItem(): checkState()={state_after_add}")

        # Resize columns to content
        table.resizeColumnsToContents()
        # Make checkbox column narrower (with DPI scaling)
        app_settings = AppSettings()
        dpi_scale = app_settings.get_dpi_scale()
        table.setColumnWidth(0, int(60 * dpi_scale))

        # Unblock signals
        table.blockSignals(False)

    def _select_all_rows(self, tab_name, select):
        """Select or deselect all rows in a tab's table.

        Args:
            tab_name: Name of the tab
            select: True to select all, False to deselect all
        """
        widgets = self.tab_widgets.get(tab_name)
        if not widgets:
            return

        table = widgets["table"]
        table.blockSignals(True)

        check_state = QtCore.Qt.Checked if select else QtCore.Qt.Unchecked

        for row in range(table.rowCount()):
            checkbox_item = table.item(row, 0)
            if checkbox_item:
                checkbox_item.setCheckState(check_state)
                self._update_row_style(table, row, select)

        table.blockSignals(False)

    def _on_table_item_changed(self, tab_name, item):
        """Handle table item changes (checkbox state changes).

        Args:
            tab_name: Name of the tab
            item: The changed item
        """
        if item.column() != 0:  # Only handle checkbox column
            return

        widgets = self.tab_widgets.get(tab_name)
        if not widgets:
            return

        table = widgets["table"]
        row = item.row()
        is_checked = item.checkState() == QtCore.Qt.Checked

        self._update_row_style(table, row, is_checked)

    def _update_row_style(self, table, row, is_enabled):
        """Update the visual style of a row based on its enabled state.

        Args:
            table: The table widget
            row: Row index
            is_enabled: True if row is enabled, False if disabled
        """
        # Set color and font for all cells in the row
        for col in range(table.columnCount()):
            item = table.item(row, col)
            if item:
                if is_enabled:
                    # Default styling
                    item.setForeground(QtGui.QColor("#e0e0e0"))
                    font = item.font()
                    font.setStrikeOut(False)
                    item.setFont(font)
                else:
                    # Disabled styling - gray and strikethrough
                    item.setForeground(QtGui.QColor("#606060"))
                    font = item.font()
                    font.setStrikeOut(True)
                    item.setFont(font)

    def _set_status(self, message, is_error=False):
        """Set status message."""
        color = "#ff8080" if is_error else "#a0a0a0"
        self.status_label.setStyleSheet(f"color: {color}; padding: 5px;")
        self.status_label.setText(message)

    def _adjust_dialog_size(self):
        """Adjust dialog size to fit content when data is loaded."""
        # Get available screen size
        screen = QtWidgets.QApplication.primaryScreen()
        if screen:
            screen_geometry = screen.availableGeometry()

            # Set dialog to 85% of screen size
            new_width = int(screen_geometry.width() * 0.85)
            new_height = int(screen_geometry.height() * 0.85)

            # Respect minimum size
            new_width = max(new_width, self.minimumWidth())
            new_height = max(new_height, self.minimumHeight())

            self.resize(new_width, new_height)

            # Center the dialog on screen
            frame_geometry = self.frameGeometry()
            center_point = screen_geometry.center()
            frame_geometry.moveCenter(center_point)
            self.move(frame_geometry.topLeft())

    def get_imported_data(self):
        """Get the imported data from all tabs, filtered by selected rows.

        Returns:
            dict: Dictionary mapping tab names to DataFrames (only selected rows)
        """
        import pandas as pd

        result = {}
        for tab_name, (_, data) in self.tab_config.items():
            if data is None:
                continue

            # Get the table for this tab
            widgets = self.tab_widgets.get(tab_name)
            if not widgets:
                continue

            table = widgets["table"]

            # Find which rows are checked
            selected_indices = []
            for row in range(table.rowCount()):
                checkbox_item = table.item(row, 0)
                if checkbox_item and checkbox_item.checkState() == QtCore.Qt.Checked:
                    selected_indices.append(row)

            # Filter the DataFrame to only include selected rows
            if selected_indices:
                filtered_df = data.iloc[selected_indices].reset_index(drop=True)
                result[tab_name] = filtered_df
                logger.info(f"Tab '{tab_name}': {len(selected_indices)} of {len(data)} rows selected for import")

        return result


class DragDropArea(QtWidgets.QLabel):
    """A widget that accepts drag and drop of files and can be clicked to open a file dialog."""

    fileDropped = QtCore.Signal(str)  # Emits file path

    def __init__(self, parent=None):
        """Initialize the drag drop area."""
        super().__init__(parent)

        self.setAlignment(QtCore.Qt.AlignCenter)
        self.setText("Drag and drop an Excel file here\n\nor click to browse")
        self.setStyleSheet("""
            QLabel {
                border: 2px dashed #555555;
                border-radius: 8px;
                padding: 40px;
                background-color: #2b2b2b;
                color: #a0a0a0;
                font-size: 14px;
            }
            QLabel:hover {
                border-color: #777777;
                background-color: #333333;
                cursor: pointer;
            }
        """)

        # Enable drag and drop
        self.setAcceptDrops(True)
        self.setMinimumHeight(150)

        # Make it behave like a button
        self.setCursor(QtCore.Qt.PointingHandCursor)

    def mousePressEvent(self, event):
        """Handle mouse click to open file dialog."""
        if event.button() == QtCore.Qt.LeftButton:
            file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
                self,
                "Select Excel File",
                "",
                "Excel Files (*.xlsx *.xls);;All Files (*)"
            )
            if file_path:
                self.fileDropped.emit(file_path)

    def dragEnterEvent(self, event):
        """Handle drag enter event."""
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            self.setStyleSheet("""
                QLabel {
                    border: 2px dashed #4a9eff;
                    border-radius: 8px;
                    padding: 40px;
                    background-color: #353535;
                    color: #4a9eff;
                    font-size: 14px;
                }
            """)

    def dragLeaveEvent(self, event):
        """Handle drag leave event."""
        self.setStyleSheet("""
            QLabel {
                border: 2px dashed #555555;
                border-radius: 8px;
                padding: 40px;
                background-color: #2b2b2b;
                color: #a0a0a0;
                font-size: 14px;
            }
        """)

    def dropEvent(self, event):
        """Handle drop event."""
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            urls = event.mimeData().urls()
            if urls:
                file_path = urls[0].toLocalFile()
                self.fileDropped.emit(file_path)

        # Reset style
        self.setStyleSheet("""
            QLabel {
                border: 2px dashed #555555;
                border-radius: 8px;
                padding: 40px;
                background-color: #2b2b2b;
                color: #a0a0a0;
                font-size: 14px;
            }
            QLabel:hover {
                border-color: #777777;
                background-color: #333333;
                cursor: pointer;
            }
        """)


class BidSelectorWidget(QtWidgets.QWidget):
    """
    Reusable widget for selecting and managing Bids.
    Displays a selector bar with dropdown and action buttons.
    """

    # Signals
    bidChanged = QtCore.Signal(object)  # Emits selected bid data (dict or None)
    statusMessageChanged = QtCore.Signal(str, bool)  # message, is_error

    def __init__(self, sg_session, parent=None):
        """Initialize the Bid selector widget.

        Args:
            sg_session: ShotGrid session for API access
            parent: Parent widget
        """
        super().__init__(parent)
        self.sg_session = sg_session
        self.parent_app = parent

        # Current state
        self.current_rfq = None
        self.current_project_id = None

        # UI widgets
        self.group_box = None
        self.bid_combo = None
        self.bid_info_label = None
        self.set_current_btn = None
        self.add_btn = None
        self.remove_btn = None
        self.rename_btn = None
        self.refresh_btn = None
        self.import_btn = None
        self.status_label = None

        self._build_ui()

    def _build_ui(self):
        """Build the bid selector UI."""
        # Main collapsible group
        self.group_box = CollapsibleGroupBox("Bids")
        group = self.group_box

        # Selector row
        selector_row = QtWidgets.QHBoxLayout()
        selector_label = QtWidgets.QLabel("Bid:")
        selector_row.addWidget(selector_label)

        self.bid_combo = QtWidgets.QComboBox()
        self.bid_combo.setMinimumWidth(250)
        self.bid_combo.currentIndexChanged.connect(self._on_bid_changed)
        selector_row.addWidget(self.bid_combo, stretch=1)

        # Config Bid button (icon only) - match height to combobox
        self.config_bid_btn = QtWidgets.QPushButton()
        self.config_bid_btn.setToolTip("Configure Bid")
        self.config_bid_btn.setIcon(create_icon_from_svg_path(FILE_COG_SVG_PATH, size=16, color="#e0e0e0"))
        self.config_bid_btn.setIconSize(QtCore.QSize(16, 16))
        # Match size to the combobox height
        combo_height = self.bid_combo.sizeHint().height()
        self.config_bid_btn.setFixedSize(combo_height, combo_height)
        self.config_bid_btn.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                border: 1px solid #555555;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #3a3a3a;
                border: 1px solid #666666;
            }
            QPushButton:pressed {
                background-color: #2a2a2a;
            }
            QPushButton:disabled {
                opacity: 0.5;
            }
        """)
        self.config_bid_btn.setEnabled(False)
        self.config_bid_btn.clicked.connect(self._on_config_bid)
        selector_row.addWidget(self.config_bid_btn)

        self.set_current_btn = QtWidgets.QPushButton("Set as Current")
        self.set_current_btn.setEnabled(False)
        self.set_current_btn.clicked.connect(self._on_set_current_bid)
        self.set_current_btn.setToolTip("Set this Bid as the current one for the selected RFQ")
        selector_row.addWidget(self.set_current_btn)

        self.add_btn = QtWidgets.QPushButton("Create")
        self.add_btn.clicked.connect(self._on_add_bid)
        self.add_btn.setToolTip("Create a new Bid")
        selector_row.addWidget(self.add_btn)

        self.refresh_btn = QtWidgets.QPushButton("Refresh")
        self.refresh_btn.clicked.connect(self._on_refresh_bids)
        self.refresh_btn.setToolTip("Refresh the Bid list")
        selector_row.addWidget(self.refresh_btn)

        self.import_btn = QtWidgets.QPushButton("Import")
        self.import_btn.clicked.connect(self._on_import_bid)
        self.import_btn.setToolTip("Import bid data from an Excel file")
        selector_row.addWidget(self.import_btn)

        group.addLayout(selector_row)

        # Bid info label (breakdown and asset)
        self.bid_info_label = QtWidgets.QLabel("")
        self.bid_info_label.setStyleSheet("padding: 2px 0;")
        self.bid_info_label.setWordWrap(True)
        group.addWidget(self.bid_info_label)

        # Status label
        self.status_label = QtWidgets.QLabel("Select an RFQ to view Bids.")
        self.status_label.setStyleSheet("color: #a0a0a0; padding: 2px 0;")
        group.addWidget(self.status_label)

        # Add group to main layout
        main_layout = QtWidgets.QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.addWidget(group)

    def populate_bids(self, rfq=None, project_id=None, auto_select=True):
        """Populate the Bid combo box.

        Args:
            rfq: RFQ data dict (optional, used to auto-select linked bid)
            project_id: Project ID to load bids from
            auto_select: Whether to auto-select a bid (default: True)
        """
        # Store current RFQ and project for button handlers
        self.current_rfq = rfq
        self.current_project_id = project_id

        self.bid_combo.blockSignals(True)
        self.bid_combo.clear()
        self.bid_combo.addItem("-- Select Bid --", None)

        # If no project, clear everything and return
        if not project_id:
            self.bid_combo.blockSignals(False)
            self.set_current_btn.setEnabled(False)
            self.config_bid_btn.setEnabled(False)
            self._set_status("Select an RFQ to view Bids.")
            return

        bids = []
        try:
            # Get RFQ ID for filtering bids
            rfq_id = rfq.get("id") if rfq else None
            logger.info(f"Loading Bids for Project ID {project_id}, RFQ ID {rfq_id}")
            # Get bids filtered by RFQ (only bids linked to this RFQ via sg_parent_rfq)
            bids = self.sg_session.get_bids(
                project_id,
                fields=["id", "code", "name", "sg_bid_type", "sg_vfx_breakdown", "sg_bid_assets", "sg_price_list", "description"],
                rfq_id=rfq_id
            )
        except Exception as e:
            logger.error(f"Error loading Bids: {e}", exc_info=True)
            bids = []

        for bid in bids:
            # Format label: "Bid Name (Bid Type)"
            bid_name = bid.get("code") or f"Bid {bid.get('id', 'N/A')}"
            bid_type = bid.get("sg_bid_type", "Unknown")
            label = f"{bid_name} ({bid_type})"
            self.bid_combo.addItem(label, bid)

        self.bid_combo.blockSignals(False)

        # Enable Set and Config buttons only if there are bids and an RFQ is selected
        self.set_current_btn.setEnabled(len(bids) > 0 and rfq is not None)
        self.config_bid_btn.setEnabled(len(bids) > 0 and rfq is not None)

        # Status & selection
        if bids:
            self._set_status(f"Loaded {len(bids)} Bid(s) for this RFQ.")

            # Auto-select the current bid linked to the RFQ if present
            bid_was_selected = False
            if rfq and auto_select:
                linked_bid = rfq.get("sg_current_bid")

                linked_bid_id = None
                if isinstance(linked_bid, dict):
                    linked_bid_id = linked_bid.get("id")
                elif isinstance(linked_bid, list) and linked_bid:
                    linked_bid_id = linked_bid[0].get("id") if linked_bid[0] else None

                if linked_bid_id:
                    # Try to select the current bid
                    if self._select_bid_by_id(linked_bid_id):
                        bid_was_selected = True
                    else:
                        # Current bid not found in list, don't auto-select anything
                        logger.warning(f"Current bid {linked_bid_id} not found in RFQ bids")

            # If no bid was selected (either no linked bid or linked bid not found),
            # manually trigger bidChanged with None to reset downstream components
            if rfq and auto_select and not bid_was_selected:
                logger.info("No bid linked to RFQ - resetting bid selection")
                self._on_bid_changed(0)  # Trigger with index 0 (placeholder)
        else:
            self._set_status("No Bids found for this RFQ.")
            # No bids available - reset downstream components
            if auto_select:
                self._on_bid_changed(0)  # Trigger with index 0 to reset

    def _select_bid_by_id(self, bid_id):
        """Select a bid by its ID.

        Args:
            bid_id: Bid ID to select

        Returns:
            bool: True if found and selected, False otherwise
        """
        if not bid_id:
            return False

        for index in range(self.bid_combo.count()):
            bid = self.bid_combo.itemData(index)
            if isinstance(bid, dict) and bid.get("id") == bid_id:
                self.bid_combo.setCurrentIndex(index)
                return True
        return False

    def get_current_bid(self):
        """Get the currently selected bid.

        Returns:
            dict: Current bid data or None
        """
        return self.bid_combo.currentData()

    def _update_bid_info_label(self, bid):
        """Update the bid info label and group box title with breakdown, bid asset, price list and description info.

        Args:
            bid: Bid data dict or None
        """
        if not bid:
            # Clear label and group box title if no bid selected
            self.bid_info_label.setText("")
            self.group_box.setAdditionalInfo("")
            return

        # Colors for formatting
        label_color = "#a0a0a0"  # Light gray for labels (normal weight)
        value_color = "#6b9bd1"  # Blue for values (bold)

        info_parts = []

        # Get bid name for title bar
        bid_name = bid.get("code") or f"Bid {bid.get('id', 'N/A')}"
        # Start with "Current Bid: (bid name)"
        title_text = f"Current Bid: {bid_name}"

        # Add VFX Breakdown info
        breakdown = bid.get("sg_vfx_breakdown")
        if breakdown:
            # Extract breakdown name/code
            if isinstance(breakdown, dict):
                breakdown_name = breakdown.get("name") or breakdown.get("code") or f"ID {breakdown.get('id', 'N/A')}"
            elif isinstance(breakdown, list) and breakdown:
                breakdown_name = breakdown[0].get("name") or breakdown[0].get("code") or f"ID {breakdown[0].get('id', 'N/A')}"
            else:
                breakdown_name = str(breakdown)
        else:
            breakdown_name = "None"
        info_parts.append(f'<span style="color:{label_color};">Vfx Breakdown:</span> <span style="color:{value_color}; font-weight:bold;">{breakdown_name}</span>')

        # Add Bid Assets info
        bid_assets = bid.get("sg_bid_assets")
        if bid_assets:
            # Extract bid asset name/code
            if isinstance(bid_assets, dict):
                asset_name = bid_assets.get("name") or bid_assets.get("code") or f"ID {bid_assets.get('id', 'N/A')}"
            elif isinstance(bid_assets, list) and bid_assets:
                asset_name = bid_assets[0].get("name") or bid_assets[0].get("code") or f"ID {bid_assets[0].get('id', 'N/A')}"
            else:
                asset_name = str(bid_assets)
        else:
            asset_name = "None"
        info_parts.append(f'<span style="color:{label_color};">Bid Assets:</span> <span style="color:{value_color}; font-weight:bold;">{asset_name}</span>')

        # Add Price List info
        price_list = bid.get("sg_price_list")
        if price_list:
            # Extract price list name/code
            if isinstance(price_list, dict):
                price_list_name = price_list.get("name") or price_list.get("code") or f"ID {price_list.get('id', 'N/A')}"
            elif isinstance(price_list, list) and price_list:
                price_list_name = price_list[0].get("name") or price_list[0].get("code") or f"ID {price_list[0].get('id', 'N/A')}"
            else:
                price_list_name = str(price_list)
        else:
            price_list_name = "None"
        info_parts.append(f'<span style="color:{label_color};">Price List:</span> <span style="color:{value_color}; font-weight:bold;">{price_list_name}</span>')

        # Add Description info
        description = bid.get("description")
        if description:
            # Truncate long descriptions for display
            if len(description) > 50:
                description = description[:50] + "..."
        else:
            description = "None"
        info_parts.append(f'<span style="color:{label_color};">Description:</span> <span style="color:{value_color}; font-weight:bold;">{description}</span>')

        # Update the label with HTML formatted info (for display under dropdown)
        self.bid_info_label.setText(" · ".join(info_parts))

        # Update the group box title with bid name and info (for collapsed state)
        self.group_box.setAdditionalInfo(title_text)

    def clear(self):
        """Clear the bid selector."""
        self.bid_combo.blockSignals(True)
        self.bid_combo.clear()
        self.bid_combo.addItem("-- Select Bid --", None)
        self.bid_combo.blockSignals(False)

        # Clear the bid info label and group box title
        self.bid_info_label.setText("")
        self.group_box.setAdditionalInfo("")
        self.set_current_btn.setEnabled(False)
        self.config_bid_btn.setEnabled(False)
        self._set_status("Select an RFQ to view Bids.")

    def _on_bid_changed(self, index):
        """Handle bid selection change."""
        bid = self.bid_combo.itemData(index)

        if bid:
            bid_name = bid.get("code") or f"Bid {bid.get('id', 'N/A')}"
            logger.info(f"BID SELECTOR - Bid selected: {bid_name} (ID: {bid.get('id')})")
            logger.info(f"  Bid data keys: {list(bid.keys())}")
            logger.info(f"  sg_bid_assets: {bid.get('sg_bid_assets')}")
            logger.info(f"  sg_vfx_breakdown: {bid.get('sg_vfx_breakdown')}")
            logger.info(f"  sg_price_list: {bid.get('sg_price_list')}")
        else:
            if index == 0:
                self._set_status("Select a Bid to view its details.")

        # Update the bid info label with breakdown and bid asset info
        self._update_bid_info_label(bid)

        # Emit signal
        logger.info(f"BID SELECTOR - Emitting bidChanged signal with bid: {bid.get('code') if bid else None}")
        self.bidChanged.emit(bid)

    def _on_set_current_bid(self):
        """Handle Set as Current button click - link bid to RFQ."""
        bid = self.get_current_bid()
        if not bid:
            QtWidgets.QMessageBox.warning(self, "No Bid Selected", "Please select a Bid from the list.")
            return

        # Get current RFQ
        if not self.current_rfq:
            QtWidgets.QMessageBox.warning(self, "No RFQ Selected", "Please select an RFQ first.")
            return

        rfq = self.current_rfq

        bid_name = bid.get('code', f"Bid {bid.get('id')}")
        bid_type = bid.get('sg_bid_type', '')
        rfq_code = rfq.get('code', 'N/A')

        # Confirm with user
        reply = QtWidgets.QMessageBox.question(
            self,
            "Set Current Bid",
            f"Set '{bid_name}' ({bid_type}) as the current bid for RFQ '{rfq_code}'?",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No
        )

        if reply != QtWidgets.QMessageBox.Yes:
            return

        try:
            # Update RFQ to link this bid as the current bid
            rfq_id = rfq['id']
            update_data = {"sg_current_bid": {"type": "CustomEntity06", "id": bid['id']}}

            self.sg_session.sg.update("CustomEntity04", rfq_id, update_data)

            logger.info(f"DEBUG: Successfully updated ShotGrid - RFQ {rfq_id} sg_current_bid = Bid {bid['id']}")
            logger.info(f"DEBUG: self.parent_app = {self.parent_app}")
            logger.info(f"DEBUG: type(self.parent_app) = {type(self.parent_app)}")
            logger.info(f"DEBUG: hasattr(self.parent_app, 'parent_app') = {hasattr(self.parent_app, 'parent_app')}")
            logger.info(f"DEBUG: hasattr(self.parent_app, 'rfq_bid_label') = {hasattr(self.parent_app, 'rfq_bid_label')}")

            # Update the Current Bid label directly in main app
            # Need to go through parent_app.parent_app since:
            # BidSelectorWidget.parent_app = BiddingTab
            # BiddingTab.parent_app = MainApp (which has rfq_bid_label)
            main_app = None
            if hasattr(self.parent_app, 'parent_app'):
                main_app = self.parent_app.parent_app
                logger.info(f"DEBUG: main_app = {main_app}")
                logger.info(f"DEBUG: hasattr(main_app, 'rfq_bid_label') = {hasattr(main_app, 'rfq_bid_label')}")

            if main_app and hasattr(main_app, 'rfq_bid_label'):
                # Use the bid's name for display
                bid_display_name = bid.get('name') or bid.get('code') or bid_name
                label_text = f"{bid_display_name} ({bid_type})" if bid_type else bid_display_name
                main_app.rfq_bid_label.setText(label_text)
                logger.info(f"✓ Updated Current Bid label to: {label_text}")
            else:
                logger.warning("Could not update Current Bid label - main_app or rfq_bid_label not found")

            # Update the current_rfq's sg_current_bid field in memory to stay in sync
            if self.current_rfq:
                self.current_rfq['sg_current_bid'] = {"type": "CustomEntity06", "id": bid['id'], "name": bid_display_name}

            self.statusMessageChanged.emit(f"✓ Set '{bid_name}' as current bid for RFQ", False)
            QtWidgets.QMessageBox.information(self, "Success", f"'{bid_name}' is now the current bid for this RFQ.")

        except Exception as e:
            logger.error(f"Failed to set current bid: {e}", exc_info=True)
            QtWidgets.QMessageBox.critical(self, "Error", f"Failed to set current bid:\n{str(e)}")

    def _is_bid_name_unique(self, bid_name, project_id):
        """Check if the Bid name is unique within the project.

        Args:
            bid_name: Name to check
            project_id: Project ID to search in

        Returns:
            bool: True if name is unique, False otherwise
        """
        try:
            existing = self.sg_session.sg.find(
                "CustomEntity06",
                [
                    ["project", "is", {"type": "Project", "id": project_id}],
                    ["code", "is", bid_name]
                ],
                ["id"]
            )
            return len(existing) == 0
        except Exception as e:
            logger.error(f"Failed to check Bid name uniqueness: {e}")
            return True  # Allow creation if check fails

    def _get_next_version_number(self, base_name, entity_type, project_id):
        """Get the next version number for a child entity name.

        Args:
            base_name: Base name pattern (e.g., "MyBid-VFX Breakdown")
            entity_type: ShotGrid entity type (e.g., "CustomEntity01")
            project_id: Project ID

        Returns:
            int: Next version number (starting from 1)
        """
        try:
            # Query existing entities with similar names
            existing = self.sg_session.sg.find(
                entity_type,
                [
                    ["project", "is", {"type": "Project", "id": project_id}],
                    ["code", "starts_with", base_name + "-v"]
                ],
                ["code"]
            )

            if not existing:
                return 1

            # Extract version numbers
            import re
            max_version = 0
            pattern = re.compile(rf"^{re.escape(base_name)}-v(\d+)$")
            for entity in existing:
                match = pattern.match(entity.get("code", ""))
                if match:
                    version = int(match.group(1))
                    max_version = max(max_version, version)

            return max_version + 1
        except Exception as e:
            logger.error(f"Failed to get next version number: {e}")
            return 1

    def _create_child_entities_for_bid(self, bid_id, bid_name, project_id):
        """Create VFX Breakdown, Bid Assets, and Price List for a new Bid.

        Args:
            bid_id: ID of the newly created Bid
            bid_name: Name of the Bid (used for child entity naming)
            project_id: Project ID

        Returns:
            dict: Dictionary with created entity info
        """
        created = {"vfx_breakdown": None, "bid_assets": None, "price_list": None}

        try:
            # Create VFX Breakdown
            vfx_base_name = f"{bid_name}-VFX Breakdown"
            vfx_version = self._get_next_version_number(vfx_base_name, "CustomEntity01", project_id)
            vfx_name = f"{vfx_base_name}-v{vfx_version:03d}"

            vfx_breakdown = self.sg_session.sg.create("CustomEntity01", {
                "code": vfx_name,
                "project": {"type": "Project", "id": project_id},
                "sg_parent_bid": {"type": "CustomEntity06", "id": bid_id}
            })
            created["vfx_breakdown"] = vfx_breakdown
            logger.info(f"Created VFX Breakdown: {vfx_name} (ID: {vfx_breakdown['id']})")

            # Create initial Bidding Scene for VFX Breakdown
            self.sg_session.create_bidding_scene(project_id, vfx_breakdown['id'], code="New Bidding Scene")

            # Create Bid Assets
            assets_base_name = f"{bid_name}-Bid Assets"
            assets_version = self._get_next_version_number(assets_base_name, "CustomEntity08", project_id)
            assets_name = f"{assets_base_name}-v{assets_version:03d}"

            bid_assets = self.sg_session.sg.create("CustomEntity08", {
                "code": assets_name,
                "project": {"type": "Project", "id": project_id},
                "sg_parent_bid": {"type": "CustomEntity06", "id": bid_id}
            })
            created["bid_assets"] = bid_assets
            logger.info(f"Created Bid Assets: {assets_name} (ID: {bid_assets['id']})")

            # Create initial Asset Item for Bid Assets
            self.sg_session.create_asset_item(project_id, bid_assets['id'], code="New Asset")

            # Create Price List
            pricelist_base_name = f"{bid_name}-Price List"
            pricelist_version = self._get_next_version_number(pricelist_base_name, "CustomEntity10", project_id)
            pricelist_name = f"{pricelist_base_name}-v{pricelist_version:03d}"

            price_list = self.sg_session.sg.create("CustomEntity10", {
                "code": pricelist_name,
                "project": {"type": "Project", "id": project_id},
                "sg_parent_bid": {"type": "CustomEntity06", "id": bid_id}
            })
            created["price_list"] = price_list
            logger.info(f"Created Price List: {pricelist_name} (ID: {price_list['id']})")

            # Create initial Line Item for Price List
            self.sg_session.sg.create("CustomEntity03", {
                "code": "New Line Item",
                "project": {"type": "Project", "id": project_id},
                "sg_parent_pricelist": {"type": "CustomEntity10", "id": price_list['id']}
            })

            # Set Rate Card if available
            try:
                rate_cards = self.sg_session.sg.find(
                    "CustomNonProjectEntity01",
                    [],
                    ["id", "code"],
                    order=[{"field_name": "code", "direction": "asc"}]
                )
                if rate_cards:
                    self.sg_session.sg.update(
                        "CustomEntity10",
                        price_list['id'],
                        {"sg_rate_card": {"type": "CustomNonProjectEntity01", "id": rate_cards[0]['id']}}
                    )
                    logger.info(f"Set Rate Card on Price List: {rate_cards[0].get('code')}")
            except Exception as e:
                logger.warning(f"Could not set Rate Card: {e}")

            # Link all children to the Bid
            self.sg_session.sg.update(
                "CustomEntity06",
                bid_id,
                {
                    "sg_vfx_breakdown": {"type": "CustomEntity01", "id": vfx_breakdown['id']},
                    "sg_bid_assets": {"type": "CustomEntity08", "id": bid_assets['id']},
                    "sg_price_list": {"type": "CustomEntity10", "id": price_list['id']}
                }
            )
            logger.info(f"Linked all child entities to Bid {bid_id}")

        except Exception as e:
            logger.error(f"Failed to create child entities: {e}", exc_info=True)
            raise

        return created

    def _on_config_bid(self):
        """Handle Config Bid button click - show dialog to configure bid children."""
        bid = self.get_current_bid()
        if not bid:
            QtWidgets.QMessageBox.warning(self, "No Bid Selected", "Please select a Bid from the list.")
            return

        bid_id = bid.get("id")
        bid_name = bid.get("code", f"Bid {bid_id}")

        # Show config dialog
        dialog = ConfigBidDialog(
            self.sg_session,
            self.current_project_id,
            bid_id,
            bid,
            parent=self
        )

        result = dialog.exec_()

        # Check if bid was deleted
        if dialog.was_bid_deleted():
            # Refresh bids list after deletion
            self._refresh_bids()
            self.statusMessageChanged.emit(f"✓ Deleted bid '{bid_name}'", False)
            return

        if result != QtWidgets.QDialog.Accepted:
            return

        # Get values from dialog
        new_name = dialog.get_bid_name()
        description = dialog.get_description()
        vfx_breakdown_id = dialog.get_vfx_breakdown_id()
        bid_assets_id = dialog.get_bid_assets_id()
        price_list_id = dialog.get_price_list_id()
        rate_card_id = dialog.get_rate_card_id()

        try:
            # Update bid with name, description and selected children
            update_data = {}

            # Update name if changed
            if new_name and new_name != bid_name:
                update_data["code"] = new_name

            # Update description
            update_data["description"] = description if description else None

            if vfx_breakdown_id:
                update_data["sg_vfx_breakdown"] = {"type": "CustomEntity01", "id": vfx_breakdown_id}
            else:
                update_data["sg_vfx_breakdown"] = None

            if bid_assets_id:
                update_data["sg_bid_assets"] = {"type": "CustomEntity08", "id": bid_assets_id}
            else:
                update_data["sg_bid_assets"] = None

            if price_list_id:
                update_data["sg_price_list"] = {"type": "CustomEntity10", "id": price_list_id}
            else:
                update_data["sg_price_list"] = None

            # Update bid
            self.sg_session.sg.update("CustomEntity06", bid_id, update_data)
            logger.info(f"Updated Bid {bid_id} with config: {update_data}")

            # Update rate card on price list if both are selected
            if price_list_id and rate_card_id:
                self.sg_session.sg.update(
                    "CustomEntity10",
                    price_list_id,
                    {"sg_rate_card": {"type": "CustomNonProjectEntity01", "id": rate_card_id}}
                )
                logger.info(f"Updated Price List {price_list_id} with Rate Card {rate_card_id}")

            display_name = new_name if new_name else bid_name
            self._set_status(f"Bid '{display_name}' configuration saved.")

            # Refresh bids to update the info label
            self.populate_bids(self.current_rfq, self.current_project_id, auto_select=False)

            # Re-select the current bid
            self._select_bid_by_id(bid_id)

        except Exception as e:
            logger.error(f"Failed to update Bid configuration: {e}", exc_info=True)
            QtWidgets.QMessageBox.critical(
                self,
                "Error",
                f"Failed to save Bid configuration: {str(e)}"
            )

    def _on_add_bid(self):
        """Handle Add Bid button click."""
        # Get current project
        if not self.current_project_id:
            QtWidgets.QMessageBox.warning(self, "No Project Selected", "Please select a project first.")
            return

        # Require an RFQ to be selected
        if not self.current_rfq:
            QtWidgets.QMessageBox.warning(self, "No RFQ Selected", "Please select an RFQ first.")
            return

        rfq_id = self.current_rfq.get("id")

        # Show dialog to get bid name, type, and copy options
        dialog = AddBidDialog(self.sg_session, self.current_project_id, rfq_id, parent=self)
        if dialog.exec_() != QtWidgets.QDialog.Accepted:
            return

        bid_name = dialog.get_bid_name()
        bid_type = dialog.get_bid_type()
        description = dialog.get_description()

        if not bid_name:
            QtWidgets.QMessageBox.warning(self, "Invalid Input", "Please enter a bid name.")
            return

        # Check if Bid name is unique
        if not self._is_bid_name_unique(bid_name, self.current_project_id):
            QtWidgets.QMessageBox.warning(
                self,
                "Duplicate Name",
                f"A Bid with the name '{bid_name}' already exists in this project.\n"
                "Please choose a different name."
            )
            return

        try:
            # Create the bid linked to the current RFQ
            new_bid = self.sg_session.create_bid(
                self.current_project_id, bid_name, bid_type, parent_rfq_id=rfq_id, description=description
            )
            new_bid_id = new_bid['id']

            logger.info(f"Created new bid: {bid_name} (ID: {new_bid_id})")

            # Handle copy mode
            if dialog.is_copy_mode():
                source_bid_data = dialog.get_source_bid_data()
                if source_bid_data:
                    copied_items = []

                    # Copy VFX Breakdown if selected
                    if dialog.should_copy_vfx_breakdown():
                        source_vfx_breakdown = source_bid_data.get("sg_vfx_breakdown")
                        if source_vfx_breakdown:
                            source_vfx_id = source_vfx_breakdown.get("id") if isinstance(source_vfx_breakdown, dict) else None
                            if source_vfx_id:
                                vfx_name = f"{bid_name} - VFX Breakdown-v001"
                                new_vfx_breakdown = self._copy_vfx_breakdown(source_vfx_id, vfx_name, bid_id=new_bid_id)
                                if new_vfx_breakdown:
                                    # Link to the new Bid
                                    self.sg_session.sg.update(
                                        "CustomEntity06",
                                        new_bid_id,
                                        {"sg_vfx_breakdown": {"type": "CustomEntity01", "id": new_vfx_breakdown["id"]}}
                                    )
                                    copied_items.append("VFX Breakdown")
                                    logger.info(f"Linked VFX Breakdown {new_vfx_breakdown['id']} to Bid {new_bid_id}")

                    # Copy Bid Assets if selected
                    if dialog.should_copy_bid_assets():
                        source_bid_assets = source_bid_data.get("sg_bid_assets")
                        if source_bid_assets:
                            source_assets_id = source_bid_assets.get("id") if isinstance(source_bid_assets, dict) else None
                            if source_assets_id:
                                assets_name = f"{bid_name} - Assets-v001"
                                new_bid_assets = self._copy_bid_assets(source_assets_id, assets_name, bid_id=new_bid_id)
                                if new_bid_assets:
                                    # Link to the new Bid
                                    self.sg_session.sg.update(
                                        "CustomEntity06",
                                        new_bid_id,
                                        {"sg_bid_assets": {"type": "CustomEntity08", "id": new_bid_assets["id"]}}
                                    )
                                    copied_items.append("Bid Assets")
                                    logger.info(f"Linked Bid Assets {new_bid_assets['id']} to Bid {new_bid_id}")

                    # Copy Price List if selected
                    if dialog.should_copy_price_list():
                        source_price_list = source_bid_data.get("sg_price_list")
                        if source_price_list:
                            source_price_list_id = source_price_list.get("id") if isinstance(source_price_list, dict) else None
                            if source_price_list_id:
                                price_list_name = f"{bid_name} - Price List-v001"
                                new_price_list = self._copy_price_list(source_price_list_id, price_list_name, bid_id=new_bid_id)
                                if new_price_list:
                                    # Link to the new Bid
                                    self.sg_session.sg.update(
                                        "CustomEntity06",
                                        new_bid_id,
                                        {"sg_price_list": {"type": "CustomEntity10", "id": new_price_list["id"]}}
                                    )
                                    copied_items.append("Price List")
                                    logger.info(f"Linked Price List {new_price_list['id']} to Bid {new_bid_id}")

                    if copied_items:
                        self._set_status(f"Created bid '{bid_name}' with copied: {', '.join(copied_items)}")
                    else:
                        self._set_status(f"Created bid '{bid_name}' (nothing to copy from source)")
                else:
                    self._set_status(f"Created bid '{bid_name}'")
            else:
                # Not copy mode - create new child entities
                try:
                    created = self._create_child_entities_for_bid(new_bid_id, bid_name, self.current_project_id)
                    created_names = []
                    if created.get("vfx_breakdown"):
                        created_names.append("VFX Breakdown")
                    if created.get("bid_assets"):
                        created_names.append("Bid Assets")
                    if created.get("price_list"):
                        created_names.append("Price List")
                    self._set_status(f"Created bid '{bid_name}' with: {', '.join(created_names)}")
                except Exception as e:
                    logger.error(f"Failed to create child entities for bid: {e}", exc_info=True)
                    self._set_status(f"Created bid '{bid_name}' (warning: some child entities failed)")

            # Refresh the bid list
            self._refresh_bids()

            # Select the newly created bid
            self._select_bid_by_id(new_bid_id)

            self.statusMessageChanged.emit(f"✓ Created bid '{bid_name}'", False)
            QtWidgets.QMessageBox.information(self, "Success", f"Bid '{bid_name}' created successfully.")

        except Exception as e:
            logger.error(f"Failed to create bid: {e}", exc_info=True)
            QtWidgets.QMessageBox.critical(self, "Error", f"Failed to create bid:\n{str(e)}")

    def _copy_vfx_breakdown(self, source_id, new_name, bid_id=None):
        """Copy a VFX Breakdown with all its Bidding Scenes.

        Args:
            source_id: ID of the VFX Breakdown to copy from
            new_name: Name for the new VFX Breakdown
            bid_id: Optional Bid ID to link via sg_parent_bid

        Returns:
            dict: The created VFX Breakdown entity, or None on failure
        """
        try:
            entity_type = self.sg_session.get_vfx_breakdown_entity_type()

            # Create the new VFX Breakdown
            breakdown_data = {
                "code": new_name,
                "project": {"type": "Project", "id": self.current_project_id}
            }
            if bid_id:
                breakdown_data["sg_parent_bid"] = {"type": "CustomEntity06", "id": bid_id}
            new_breakdown = self.sg_session.sg.create(entity_type, breakdown_data)
            new_breakdown_id = new_breakdown["id"]

            logger.info(f"Created VFX Breakdown copy: {new_name} (ID: {new_breakdown_id})")

            # Fetch all bidding scenes from the source breakdown
            source_scenes = self.sg_session.get_bidding_scenes_for_vfx_breakdown(
                source_id,
                fields=[
                    "code", "sg_beat_id", "sg_vfx_breakdown_scene", "sg_page",
                    "sg_script_excerpt", "description", "sg_vfx_type", "sg_complexity",
                    "sg_category", "sg_vfx_description", "sg_number_of_shots",
                    "sg_sorting_priority", "sg_vfx_shot_work"
                ]
            )

            logger.info(f"Found {len(source_scenes)} bidding scenes to copy")

            # Copy each bidding scene
            copy_fields = [
                "code", "sg_beat_id", "sg_vfx_breakdown_scene", "sg_page",
                "sg_script_excerpt", "description", "sg_vfx_type", "sg_complexity",
                "sg_category", "sg_vfx_description", "sg_number_of_shots",
                "sg_sorting_priority", "sg_vfx_shot_work"
            ]

            for scene in source_scenes:
                new_scene_data = {
                    "project": {"type": "Project", "id": self.current_project_id},
                    "sg_parent": {"type": entity_type, "id": new_breakdown_id}
                }

                for field in copy_fields:
                    if field in scene and scene[field] is not None:
                        new_scene_data[field] = scene[field]

                self.sg_session.sg.create("CustomEntity02", new_scene_data)

            logger.info(f"Successfully copied {len(source_scenes)} bidding scenes to VFX Breakdown {new_breakdown_id}")

            return new_breakdown

        except Exception as e:
            logger.error(f"Failed to copy VFX Breakdown: {e}", exc_info=True)
            return None

    def _copy_bid_assets(self, source_id, new_name, bid_id=None):
        """Copy Bid Assets with all its Asset Items.

        Args:
            source_id: ID of the Bid Assets to copy from
            new_name: Name for the new Bid Assets
            bid_id: Optional Bid ID to link via sg_parent_bid

        Returns:
            dict: The created Bid Assets entity, or None on failure
        """
        try:
            # Create the new Bid Assets
            bid_assets_data = {
                "code": new_name,
                "project": {"type": "Project", "id": self.current_project_id}
            }
            if bid_id:
                bid_assets_data["sg_parent_bid"] = {"type": "CustomEntity06", "id": bid_id}
            new_bid_assets = self.sg_session.sg.create("CustomEntity08", bid_assets_data)
            new_bid_assets_id = new_bid_assets["id"]

            logger.info(f"Created Bid Assets copy: {new_name} (ID: {new_bid_assets_id})")

            # Fetch all Asset Items from the source Bid Assets
            source_items = self.sg_session.sg.find(
                "CustomEntity07",
                [["sg_bid_assets", "is", {"type": "CustomEntity08", "id": source_id}]],
                ["code", "sg_bid_asset_type", "sg_bidding_notes"]
            )

            logger.info(f"Found {len(source_items)} asset items to copy")

            # Copy each Asset Item
            copy_fields = ["code", "sg_bid_asset_type", "sg_bidding_notes"]

            for item in source_items:
                new_item_data = {
                    "project": {"type": "Project", "id": self.current_project_id},
                    "sg_bid_assets": {"type": "CustomEntity08", "id": new_bid_assets_id}
                }

                for field in copy_fields:
                    if field in item and item[field] is not None:
                        new_item_data[field] = item[field]

                self.sg_session.sg.create("CustomEntity07", new_item_data)

            logger.info(f"Successfully copied {len(source_items)} asset items to Bid Assets {new_bid_assets_id}")

            return new_bid_assets

        except Exception as e:
            logger.error(f"Failed to copy Bid Assets: {e}", exc_info=True)
            return None

    def _copy_price_list(self, source_id, new_name, bid_id=None):
        """Copy a Price List with all its Line Items.

        Args:
            source_id: ID of the Price List to copy from
            new_name: Name for the new Price List
            bid_id: Optional Bid ID to link via sg_parent_bid

        Returns:
            dict: The created Price List entity, or None on failure
        """
        try:
            # Get source Price List data including Rate Card
            source_data = self.sg_session.sg.find_one(
                "CustomEntity10",
                [["id", "is", source_id]],
                ["sg_line_items", "sg_rate_card"]
            )

            # Create the new Price List
            price_list_data = {
                "code": new_name,
                "project": {"type": "Project", "id": self.current_project_id}
            }
            if bid_id:
                price_list_data["sg_parent_bid"] = {"type": "CustomEntity06", "id": bid_id}
            new_price_list = self.sg_session.sg.create("CustomEntity10", price_list_data)
            new_price_list_id = new_price_list["id"]

            logger.info(f"Created Price List copy: {new_name} (ID: {new_price_list_id})")

            # Copy Rate Card link if exists (Rate Cards are shared, not copied)
            if source_data and source_data.get("sg_rate_card"):
                rate_card = source_data["sg_rate_card"]
                if isinstance(rate_card, dict) and rate_card.get("id"):
                    self.sg_session.sg.update(
                        "CustomEntity10",
                        new_price_list_id,
                        {"sg_rate_card": {"type": "CustomNonProjectEntity01", "id": rate_card["id"]}}
                    )
                    logger.info(f"Linked Rate Card {rate_card['id']} to new Price List")

            # Get source Line Items
            if not source_data or not source_data.get("sg_line_items"):
                logger.info(f"No Line Items to copy from Price List {source_id}")
                return new_price_list

            source_line_items = source_data["sg_line_items"]
            if not isinstance(source_line_items, list):
                source_line_items = [source_line_items] if source_line_items else []

            source_ids = [item["id"] for item in source_line_items if isinstance(item, dict) and item.get("id")]

            if not source_ids:
                return new_price_list

            # Query source Line Items with their fields
            source_items = self.sg_session.sg.find(
                "CustomEntity03",
                [["id", "in", source_ids]],
                ["code", "sg_price_static", "sg_price_formula"] + [f for f in self._get_line_item_mandays_fields()]
            )

            logger.info(f"Found {len(source_items)} line items to copy")

            # Copy each Line Item
            new_line_item_refs = []
            for item in source_items:
                new_item_data = {
                    "project": {"type": "Project", "id": self.current_project_id},
                    "sg_parent_pricelist": {"type": "CustomEntity10", "id": new_price_list_id}
                }

                # Copy all fields except id, sg_parent_pricelist (which we set above),
                # and read-only computed fields like sg_total_mandays
                skip_fields = ("id", "type", "sg_parent_pricelist", "sg_total_mandays")
                for field, value in item.items():
                    if field not in skip_fields and value is not None:
                        # Handle entity references
                        if isinstance(value, dict) and "type" in value and "id" in value:
                            new_item_data[field] = {"type": value["type"], "id": value["id"]}
                        elif isinstance(value, list):
                            new_item_data[field] = [
                                {"type": v["type"], "id": v["id"]}
                                for v in value
                                if isinstance(v, dict) and "type" in v and "id" in v
                            ]
                        else:
                            new_item_data[field] = value

                new_line_item = self.sg_session.sg.create("CustomEntity03", new_item_data)
                new_line_item_refs.append({"type": "CustomEntity03", "id": new_line_item["id"]})

            # Link all new Line Items to the Price List
            if new_line_item_refs:
                self.sg_session.sg.update(
                    "CustomEntity10",
                    new_price_list_id,
                    {"sg_line_items": new_line_item_refs}
                )

            logger.info(f"Successfully copied {len(new_line_item_refs)} line items to Price List {new_price_list_id}")

            return new_price_list

        except Exception as e:
            logger.error(f"Failed to copy Price List: {e}", exc_info=True)
            return None

    def _get_line_item_mandays_fields(self):
        """Get the list of mandays fields for Line Items.

        Returns:
            list: List of field names ending with '_mandays'
        """
        try:
            schema = self.sg_session.sg.schema_field_read("CustomEntity03")
            return [f for f in schema.keys() if f.endswith("_mandays")]
        except Exception as e:
            logger.error(f"Failed to get Line Item schema: {e}")
            return []

    def _on_remove_bid(self):
        """Handle Remove Bid button click."""
        bid = self.get_current_bid()
        if not bid:
            QtWidgets.QMessageBox.warning(self, "No Bid Selected", "Please select a Bid from the list.")
            return

        bid_name = bid.get('code', f"Bid {bid.get('id')}")

        # Confirm deletion
        reply = QtWidgets.QMessageBox.question(
            self,
            "Confirm Delete",
            f"Are you sure you want to delete bid '{bid_name}'?\n\nThis cannot be undone.",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No
        )

        if reply != QtWidgets.QMessageBox.Yes:
            return

        try:
            # Delete the bid
            self.sg_session.delete_bid(bid['id'])

            logger.info(f"Deleted bid: {bid_name} (ID: {bid['id']})")

            # Refresh the bid list
            self._refresh_bids()

            self.statusMessageChanged.emit(f"✓ Deleted bid '{bid_name}'", False)
            QtWidgets.QMessageBox.information(self, "Success", f"Bid '{bid_name}' deleted successfully.")

        except Exception as e:
            logger.error(f"Failed to delete bid: {e}", exc_info=True)
            QtWidgets.QMessageBox.critical(self, "Error", f"Failed to delete bid:\n{str(e)}")

    def _on_rename_bid(self):
        """Handle Rename Bid button click."""
        bid = self.get_current_bid()
        if not bid:
            QtWidgets.QMessageBox.warning(self, "No Bid Selected", "Please select a Bid from the list.")
            return

        current_name = bid.get('code', '')

        # Show dialog to get new name
        dialog = RenameBidDialog(current_name, parent=self)
        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            new_name = dialog.get_new_name()

            if not new_name or new_name == current_name:
                return

            try:
                # Update the bid name
                self.sg_session.update_bid(bid['id'], {"code": new_name})

                logger.info(f"Renamed bid from '{current_name}' to '{new_name}' (ID: {bid['id']})")

                # Refresh the bid list
                self._refresh_bids()

                # Select the renamed bid
                self._select_bid_by_id(bid['id'])

                self.statusMessageChanged.emit(f"✓ Renamed bid to '{new_name}'", False)
                QtWidgets.QMessageBox.information(self, "Success", f"Bid renamed to '{new_name}'.")

            except Exception as e:
                logger.error(f"Failed to rename bid: {e}", exc_info=True)
                QtWidgets.QMessageBox.critical(self, "Error", f"Failed to rename bid:\n{str(e)}")

    def _on_import_bid(self):
        """Handle Import button click with 3-dialog navigation flow."""
        # Check if we have required data
        if not self.current_project_id:
            QtWidgets.QMessageBox.warning(self, "No Project Selected", "Please select a project first.")
            return

        # Dialog navigation state
        current_step = 1  # 1=Import, 2=Mapping, 3=Select
        import_dialog = None
        mapping_dialog = None
        data = None
        excel_columns_dict = {}
        all_mappings = {}

        # Dialog navigation loop
        while True:
            if current_step == 1:
                # Step 1: ImportBidDialog - Load Excel file
                if import_dialog is None:
                    import_dialog = ImportBidDialog(self.sg_session, self.current_project_id, parent=self)

                result = import_dialog.exec_()

                if result == QtWidgets.QDialog.Accepted:
                    data = import_dialog.get_imported_data()
                    if not data:
                        QtWidgets.QMessageBox.warning(self, "No Data", "No data loaded from Excel file.")
                        return

                    # Extract columns from all DataFrames
                    excel_columns_dict = {}
                    if "VFX Breakdown" in data:
                        excel_columns_dict["breakdown"] = list(data["VFX Breakdown"].columns)
                    if "Assets" in data:
                        excel_columns_dict["assets"] = list(data["Assets"].columns)
                    if "Scene" in data:
                        excel_columns_dict["scenes"] = list(data["Scene"].columns)
                    if "Rates" in data:
                        excel_columns_dict["rates"] = list(data["Rates"].columns)

                    current_step = 2  # Move to mapping dialog
                else:
                    # User cancelled
                    return

            elif current_step == 2:
                # Step 2: ColumnMappingDialog - Map columns
                if mapping_dialog is None:
                    mapping_dialog = ColumnMappingDialog(
                        excel_columns_dict,
                        self.sg_session,
                        self.current_project_id,
                        parent=self
                    )

                result = mapping_dialog.exec_()

                if result == QtWidgets.QDialog.Accepted:
                    # Get all mappings
                    all_mappings = mapping_dialog.get_all_column_mappings()
                    logger.info(f"Received mappings for {len(all_mappings)} entity types")
                    current_step = 3  # Move to select dialog
                elif result == ColumnMappingDialog.BACK_CLICKED:
                    # Go back to import dialog
                    current_step = 1
                else:
                    # User cancelled
                    return

            elif current_step == 3:
                # Step 3: SelectBidDialog - Select Bid and entity types
                # Get existing bids
                existing_bids = self._get_existing_bids()

                # Determine which sheets have data
                available_sheets = {
                    "breakdown": "VFX Breakdown" in data,
                    "assets": "Assets" in data,
                    "scenes": "Scene" in data,
                    "rates": "Rates" in data
                }

                select_dialog = SelectBidDialog(existing_bids, available_sheets, parent=self)
                result = select_dialog.exec_()

                if result == QtWidgets.QDialog.Accepted:
                    # Perform import
                    self._perform_import(select_dialog, data, all_mappings)
                    return
                elif result == SelectBidDialog.BACK_CLICKED:
                    # Go back to mapping dialog
                    current_step = 2
                else:
                    # User cancelled
                    return

    def _get_existing_bids(self):
        """Get list of existing bids for the current project.

        Returns:
            list: List of bid dicts [{"id": 1, "code": "BidName"}, ...]
        """
        try:
            filters = [["project", "is", {"type": "Project", "id": self.current_project_id}]]
            bids = self.sg_session.sg.find(
                "CustomEntity06",
                filters,
                ["code"],
                order=[{"field_name": "code", "direction": "asc"}]
            )
            logger.info(f"Found {len(bids)} existing bids")
            return bids
        except Exception as e:
            logger.error(f"Failed to query existing bids: {e}", exc_info=True)
            return []

    def _perform_import(self, select_dialog, data, all_mappings):
        """Perform the actual import based on user selections.

        Args:
            select_dialog: SelectBidDialog with user's choices
            data: Dict of sheet data {sheet_name -> DataFrame}
            all_mappings: Dict of column mappings {entity_type -> {field -> column}}
        """
        # Get user selections
        is_new_bid = select_dialog.is_creating_new_bid()
        selected_entity_types = select_dialog.get_selected_entity_types()

        # Get or create Bid
        if is_new_bid:
            bid_name = select_dialog.get_new_bid_name()
            bid_type = select_dialog.get_new_bid_type()

            try:
                bid = self.sg_session.create_bid(
                    self.current_project_id,
                    bid_name,
                    bid_type
                )
                bid_id = bid["id"]
                logger.info(f"Created new Bid: {bid_id} - {bid_name}")
            except Exception as e:
                logger.error(f"Failed to create Bid: {e}", exc_info=True)
                QtWidgets.QMessageBox.critical(
                    self,
                    "Error",
                    f"Failed to create Bid:\n{str(e)}"
                )
                return
        else:
            bid_id = select_dialog.get_selected_bid_id()
            # Get bid name for versioning
            try:
                bid = self.sg_session.sg.find_one(
                    "CustomEntity06",
                    [["id", "is", bid_id]],
                    ["code"]
                )
                bid_name = bid.get("code", "Unknown")
                logger.info(f"Using existing Bid: {bid_id} - {bid_name}")
            except Exception as e:
                logger.error(f"Failed to query Bid: {e}", exc_info=True)
                QtWidgets.QMessageBox.critical(
                    self,
                    "Error",
                    f"Failed to query Bid:\n{str(e)}"
                )
                return

        # Import selected entity types
        results = {
            "breakdown": {"created": 0, "entity_id": None},
            "assets": {"created": 0, "entity_id": None},
            "scenes": {"created": 0, "entity_id": None},
            "rates": {"created": 0, "entity_id": None, "failed": 0}
        }

        try:
            # Import Assets first (so Breakdown can reference them)
            if selected_entity_types.get("assets") and "Assets" in data:
                assets_mapping = all_mappings.get("assets", {})
                created, entity_id = self._import_assets(
                    data["Assets"],
                    assets_mapping,
                    bid_id,
                    bid_name
                )
                results["assets"]["created"] = created
                results["assets"]["entity_id"] = entity_id

            # Import Breakdown (can now reference Assets)
            if selected_entity_types.get("breakdown") and "VFX Breakdown" in data:
                breakdown_mapping = all_mappings.get("breakdown", {})
                created, entity_id = self._import_vfx_breakdown(
                    data["VFX Breakdown"],
                    breakdown_mapping,
                    bid_id,
                    bid_name
                )
                results["breakdown"]["created"] = created
                results["breakdown"]["entity_id"] = entity_id

            # Import Scenes
            if selected_entity_types.get("scenes") and "Scene" in data:
                scenes_mapping = all_mappings.get("scenes", {})
                created, entity_id = self._import_scenes(data["Scene"], scenes_mapping, bid_id, bid_name)
                results["scenes"]["created"] = created
                results["scenes"]["entity_id"] = entity_id

            # Import Rates
            if selected_entity_types.get("rates") and "Rates" in data:
                rates_mapping = all_mappings.get("rates", {})
                created, entity_id, failed = self._import_rates(data["Rates"], rates_mapping, bid_id, bid_name)
                results["rates"]["created"] = created
                results["rates"]["entity_id"] = entity_id
                results["rates"]["failed"] = failed

            # Show success message
            self._show_import_success(results, data)

            # Refresh dropdowns and select new entities
            self._refresh_after_import(bid_id, results)

        except Exception as e:
            logger.error(f"Import failed: {e}", exc_info=True)
            QtWidgets.QMessageBox.critical(
                self,
                "Import Error",
                f"Import failed:\n{str(e)}"
            )

    def _show_import_success(self, results, data):
        """Show success message after import."""
        summary_lines = ["Successfully imported Excel data:\n"]
        has_failures = False

        if results["breakdown"]["created"] > 0:
            summary_lines.append(f"✓ Created {results['breakdown']['created']} VFX Breakdown items")
        if results["assets"]["created"] > 0:
            summary_lines.append(f"✓ Created {results['assets']['created']} Asset items")
        if results["scenes"]["created"] > 0:
            summary_lines.append(f"✓ Created {results['scenes']['created']} Scene items")
        if results["rates"]["created"] > 0:
            failed_count = results["rates"].get("failed", 0)
            if failed_count > 0:
                summary_lines.append(f"✓ Created {results['rates']['created']} Rate items ({failed_count} failed)")
                has_failures = True
            else:
                summary_lines.append(f"✓ Created {results['rates']['created']} Rate items")

        # Add warning if there were failures
        if has_failures:
            summary_lines.append("\n⚠ Some items failed to import. Check the logs for details.")

        QtWidgets.QMessageBox.information(
            self,
            "Import Successful" if not has_failures else "Import Completed with Errors",
            "\n".join(summary_lines)
        )

        logger.info(f"Import completed: {results}")

    def _refresh_after_import(self, bid_id, results):
        """Refresh dropdowns and select imported entities.

        Args:
            bid_id: ID of the Bid
            results: Dict of import results
        """
        # Refresh Bid dropdown
        self._refresh_bids()
        self._select_bid_by_id(bid_id)

        # Refresh VFX Breakdown dropdown if items were imported
        if results["breakdown"]["entity_id"] and self._refresh_vfx_breakdown_dropdown():
            logger.info(f"Refreshed VFX Breakdown dropdown")

        # Refresh Price Lists dropdown if rates were imported
        if results["rates"]["entity_id"] and self._refresh_price_lists_dropdown():
            logger.info(f"Refreshed Price Lists dropdown")

        logger.info("Completed post-import refresh")

    def _deduplicate_entity_refs(self, entity_refs):
        """
        Deduplicate entity references by ID, keeping only unique entries.

        Args:
            entity_refs (list): List of entity reference dicts [{"type": "...", "id": xxx}, ...]

        Returns:
            list: Deduplicated list of entity references
        """
        if not entity_refs:
            return []

        seen_ids = set()
        unique_refs = []

        for ref in entity_refs:
            if not isinstance(ref, dict):
                continue

            entity_id = ref.get("id")
            if entity_id and entity_id not in seen_ids:
                seen_ids.add(entity_id)
                unique_refs.append(ref)

        if len(unique_refs) < len(entity_refs):
            logger.info(f"Deduplicated entity references: {len(entity_refs)} -> {len(unique_refs)}")

        return unique_refs

    def _parse_and_lookup_assets(self, value, project_id):
        """
        Parse asset names from text and look up their ShotGrid entity references.

        Args:
            value: Text value containing asset names (possibly multiline)
            project_id: Project ID to search within

        Returns:
            list: List of unique entity references [{"type": "CustomEntity07", "id": xxx}, ...]
                  or None if no matches found
        """
        if not value or (isinstance(value, str) and not value.strip()):
            return None

        # Convert to string if not already
        text = str(value)

        # Parse asset names - split by newlines and strip whitespace
        asset_names_raw = [name.strip() for name in text.split('\n') if name.strip()]

        if not asset_names_raw:
            return None

        # Remove duplicates while preserving order
        seen = set()
        asset_names = []
        for name in asset_names_raw:
            if name not in seen:
                seen.add(name)
                asset_names.append(name)

        logger.info(f"Looking up assets (deduplicated): {asset_names}")

        # Query ShotGrid for CustomEntity07 records matching these names
        try:
            filters = [
                ["project", "is", {"type": "Project", "id": int(project_id)}],
                ["code", "in", asset_names]
            ]

            assets = self.sg_session.sg.find(
                "CustomEntity07",
                filters,
                ["id", "code"]
            )

            if not assets:
                logger.warning(f"No matching assets found for names: {asset_names}")
                return None

            # Deduplicate by code name first (in case of duplicate assets in SG)
            # Keep only the first asset found for each unique code
            seen_codes = set()
            unique_assets = []
            for asset in assets:
                asset_code = asset.get("code")
                if asset_code and asset_code not in seen_codes:
                    seen_codes.add(asset_code)
                    unique_assets.append(asset)
                elif asset_code and asset_code in seen_codes:
                    logger.warning(f"Duplicate asset found in ShotGrid: '{asset_code}' (ID: {asset['id']}), skipping")

            # Create entity references from unique assets
            entity_refs = []
            for asset in unique_assets:
                entity_refs.append({"type": "CustomEntity07", "id": asset["id"]})

            # Log matches
            found_names = [asset["code"] for asset in unique_assets]
            logger.info(f"Found {len(entity_refs)} unique asset(s): {found_names}")

            # Warn about any missing assets
            missing = set(asset_names) - set(found_names)
            if missing:
                logger.warning(f"Could not find assets: {missing}")

            return entity_refs

        except Exception as e:
            logger.error(f"Error looking up assets: {e}", exc_info=True)
            return None

    def _import_vfx_breakdown(self, df, column_mapping, bid_id, bid_name):
        """Import VFX Breakdown data to ShotGrid.

        Args:
            df: DataFrame containing VFX Breakdown data
            column_mapping: Dict mapping ShotGrid fields to Excel column names
            bid_id: ID of the Bid to link to
            bid_name: Name of the Bid for versioning

        Returns:
            tuple: (Number of records created, VFX Breakdown ID)
        """
        import pandas as pd

        if df is None or len(df) == 0:
            return 0, None

        logger.info(f"Starting VFX Breakdown import with {len(df)} rows for Bid: {bid_name} ({bid_id})")
        logger.info(f"Column mapping: {column_mapping}")

        # Determine version number for VFX Breakdown
        # Query existing VFX Breakdowns with pattern: {bid_name}-Breakdown-v*
        try:
            filters = [
                ["project", "is", {"type": "Project", "id": self.current_project_id}],
                ["code", "starts_with", f"{bid_name}-Breakdown-v"]
            ]
            existing_breakdowns = self.sg_session.sg.find(
                "CustomEntity01",
                filters,
                ["code"]
            )

            # Find the highest version number
            max_version = 0
            for breakdown in existing_breakdowns:
                code = breakdown.get("code", "")
                # Extract version number from code (e.g., "MyBid-Breakdown-v003" -> 3)
                if code.startswith(f"{bid_name}-Breakdown-v"):
                    version_str = code.split("-v")[-1]
                    try:
                        version = int(version_str)
                        max_version = max(max_version, version)
                    except ValueError:
                        continue

            # Increment to get next version
            next_version = max_version + 1
            breakdown_name = f"{bid_name}-Breakdown-v{next_version:03d}"

            logger.info(f"Generated VFX Breakdown name: {breakdown_name} (version {next_version})")

        except Exception as e:
            logger.error(f"Failed to query existing VFX Breakdowns: {e}", exc_info=True)
            # Default to v001 if query fails
            breakdown_name = f"{bid_name}-Breakdown-v001"
            logger.warning(f"Defaulting to version 001: {breakdown_name}")

        # Create progress dialog for import
        total_steps = len(df) + 2  # +2 for VFX Breakdown and Bid creation
        progress = QtWidgets.QProgressDialog("Importing data to ShotGrid...", None, 0, total_steps, self)
        progress.setWindowModality(QtCore.Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.setValue(0)
        QtWidgets.QApplication.processEvents()

        # Step 5: Now that mapping is confirmed, create VFX Breakdown in ShotGrid
        progress.setLabelText(f"Creating VFX Breakdown: {breakdown_name}...")
        QtWidgets.QApplication.processEvents()

        try:
            breakdown_data = {
                "code": breakdown_name,
                "project": {"type": "Project", "id": self.current_project_id},
                "sg_parent_bid": {"type": "CustomEntity06", "id": bid_id}
            }
            vfx_breakdown = self.sg_session.sg.create("CustomEntity01", breakdown_data)
            breakdown_id = vfx_breakdown["id"]
            logger.info(f"Created VFX Breakdown (CustomEntity01): {breakdown_id} - {breakdown_name}")
        except Exception as e:
            logger.error(f"Failed to create VFX Breakdown: {e}", exc_info=True)
            progress.close()
            QtWidgets.QMessageBox.critical(
                self,
                "Error",
                f"Failed to create VFX Breakdown:\n{str(e)}"
            )
            return 0, None

        progress.setValue(1)
        QtWidgets.QApplication.processEvents()

        # Link VFX Breakdown to Bid
        progress.setLabelText(f"Linking VFX Breakdown to Bid...")
        QtWidgets.QApplication.processEvents()

        try:
            self.sg_session.sg.update(
                "CustomEntity06",
                bid_id,
                {"sg_vfx_breakdown": {"type": "CustomEntity01", "id": breakdown_id}}
            )
            logger.info(f"Linked VFX Breakdown {breakdown_id} to Bid {bid_id}")
        except Exception as e:
            logger.error(f"Failed to link VFX Breakdown to Bid: {e}", exc_info=True)
            # Non-fatal, continue with import

        progress.setValue(2)
        QtWidgets.QApplication.processEvents()

        # Create ShotGrid records (CustomEntity02) linked to VFX Breakdown
        created_count = 0
        failed_count = 0

        for index, row in df.iterrows():
            # Update progress
            progress.setLabelText(f"Creating breakdown item {created_count + 1} of {len(df)}...")
            progress.setValue(2 + index + 1)
            QtWidgets.QApplication.processEvents()
            try:
                # Build SG data from mapping
                sg_data = {
                    "project": {"type": "Project", "id": self.current_project_id},
                    "sg_parent": {"type": "CustomEntity01", "id": breakdown_id}  # Link to VFX Breakdown
                }

                for sg_field, excel_col in column_mapping.items():
                    if excel_col is None:
                        continue

                    # Get value from DataFrame
                    value = row[excel_col]

                    # Skip empty values
                    if pd.isna(value) or value == "":
                        continue

                    # Convert value based on field type
                    field_type = ColumnMappingDialog.BREAKDOWN_ITEM_REQUIRED_FIELDS.get(sg_field)

                    if field_type == "number":
                        try:
                            sg_data[sg_field] = int(value)
                        except (ValueError, TypeError):
                            logger.warning(f"Could not convert '{value}' to number for field '{sg_field}'")
                            continue
                    elif field_type == "float":
                        try:
                            sg_data[sg_field] = float(value)
                        except (ValueError, TypeError):
                            logger.warning(f"Could not convert '{value}' to float for field '{sg_field}'")
                            continue
                    elif field_type == "checkbox":
                        # Convert to boolean
                        if isinstance(value, bool):
                            sg_data[sg_field] = value
                        elif isinstance(value, str):
                            sg_data[sg_field] = value.lower() in ["true", "yes", "1", "x"]
                        else:
                            sg_data[sg_field] = bool(value)
                    elif field_type == "entity":
                        # Handle entity references (e.g., sg_bid_assets)
                        # Parse the text to extract asset names and look them up in ShotGrid
                        if sg_field == "sg_bid_assets":
                            entity_refs = self._parse_and_lookup_assets(value, self.current_project_id)
                            if entity_refs:
                                # Deduplicate entity references before saving to ShotGrid
                                entity_refs = self._deduplicate_entity_refs(entity_refs)
                                sg_data[sg_field] = entity_refs
                                logger.info(f"Row {index}: Setting {len(entity_refs)} deduplicated asset reference(s)")
                            # If no matches found, skip this field (don't set it)
                        else:
                            # For other entity fields, store as text for now
                            if isinstance(value, str):
                                sg_data[sg_field] = value
                            else:
                                sg_data[sg_field] = str(value)
                    else:
                        # Text and list fields - import as-is without modification
                        # If it's already a string, use it directly to preserve exact formatting
                        if isinstance(value, str):
                            sg_data[sg_field] = value
                        else:
                            # Only convert non-string values
                            sg_data[sg_field] = str(value)

                # Create the record
                result = self.sg_session.sg.create("CustomEntity02", sg_data)
                created_count += 1
                logger.info(f"Created CustomEntity02: {result['id']} with code '{sg_data.get('code', 'N/A')}' linked to VFX Breakdown {breakdown_id}")

            except Exception as e:
                failed_count += 1
                logger.error(f"Failed to create CustomEntity02 for row {index}: {e}", exc_info=True)

        # Close progress dialog
        progress.setValue(total_steps)
        progress.close()

        # Log summary (detailed message shown at end of all imports)
        if failed_count > 0:
            logger.warning(f"Created VFX Breakdown '{breakdown_name}' with {created_count} items ({failed_count} failed)")
        else:
            logger.info(f"Successfully created VFX Breakdown '{breakdown_name}' with {created_count} items")

        return created_count, breakdown_id

    def _import_assets(self, df, column_mapping, bid_id, bid_name):
        """Import Assets data to ShotGrid.

        Args:
            df: DataFrame containing Assets data
            column_mapping: Dict mapping ShotGrid fields to Excel column names
            bid_id: ID of the Bid to link to
            bid_name: Name of the Bid for versioning

        Returns:
            tuple: (Number of records created, Bid Assets ID)
        """
        import pandas as pd

        if df is None or len(df) == 0:
            return 0, None

        logger.info(f"Starting Assets import with {len(df)} rows")
        logger.info(f"Column mapping: {column_mapping}")
        logger.info(f"Importing Assets for Bid: {bid_name} (ID: {bid_id})")

        # Check for duplicate asset codes in the Excel sheet
        code_column = column_mapping.get("code")
        if code_column and code_column in df.columns:
            # Get all non-empty codes
            codes = df[code_column].dropna()
            codes = codes[codes != ""]

            # Check for duplicates
            duplicates = codes[codes.duplicated()].unique()

            if len(duplicates) > 0:
                duplicate_list = "\n".join([f"  - {code}" for code in duplicates])
                error_msg = (
                    f"Duplicate asset names found in the Assets sheet:\n\n"
                    f"{duplicate_list}\n\n"
                    f"Please fix the Excel file to ensure all asset names are unique before importing."
                )
                logger.error(f"Duplicate assets found: {list(duplicates)}")
                QtWidgets.QMessageBox.critical(
                    self,
                    "Duplicate Assets Found",
                    error_msg
                )
                return 0, None

        # Step 2: Determine version number for Bid Assets
        # Query existing Bid Assets with pattern: {bid_name}-Bid Assets-v*
        try:
            filters = [
                ["project", "is", {"type": "Project", "id": self.current_project_id}],
                ["code", "starts_with", f"{bid_name}-Bid Assets-v"]
            ]
            existing_bid_assets = self.sg_session.sg.find(
                "CustomEntity08",
                filters,
                ["code"]
            )

            # Find the highest version number
            max_version = 0
            for bid_asset in existing_bid_assets:
                code = bid_asset.get("code", "")
                # Extract version number from code (e.g., "MyBid-Bid Assets-v003" -> 3)
                if code.startswith(f"{bid_name}-Bid Assets-v"):
                    version_str = code.split("-v")[-1]
                    try:
                        version = int(version_str)
                        max_version = max(max_version, version)
                    except ValueError:
                        continue

            # Increment to get next version
            next_version = max_version + 1
            bid_assets_name = f"{bid_name}-Bid Assets-v{next_version:03d}"

            logger.info(f"Generated Bid Assets name: {bid_assets_name} (version {next_version})")

        except Exception as e:
            logger.error(f"Failed to query existing Bid Assets: {e}", exc_info=True)
            # Default to v001 if query fails
            bid_assets_name = f"{bid_name}-Bid Assets-v001"
            logger.warning(f"Defaulting to version 001: {bid_assets_name}")

        # Create progress dialog for import
        total_steps = len(df) + 1  # +1 for Bid Assets creation
        progress = QtWidgets.QProgressDialog("Importing Assets to ShotGrid...", None, 0, total_steps, self)
        progress.setWindowModality(QtCore.Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.setValue(0)
        QtWidgets.QApplication.processEvents()

        # Step 5: Create Bid Assets in ShotGrid
        progress.setLabelText(f"Creating Bid Assets: {bid_assets_name}...")
        QtWidgets.QApplication.processEvents()

        try:
            bid_assets_data = {
                "code": bid_assets_name,
                "project": {"type": "Project", "id": self.current_project_id},
                "sg_parent_bid": {"type": "CustomEntity06", "id": bid_id}
            }
            bid_assets = self.sg_session.sg.create("CustomEntity08", bid_assets_data)
            bid_assets_id = bid_assets["id"]
            logger.info(f"Created Bid Assets (CustomEntity08): {bid_assets_id} - {bid_assets_name}")
        except Exception as e:
            logger.error(f"Failed to create Bid Assets: {e}", exc_info=True)
            progress.close()
            QtWidgets.QMessageBox.critical(
                self,
                "Error",
                f"Failed to create Bid Assets:\n{str(e)}"
            )
            return 0, None

        progress.setValue(1)
        QtWidgets.QApplication.processEvents()

        # Step 6: Link Bid Assets to current Bid
        try:
            self.sg_session.sg.update(
                "CustomEntity06",
                bid_id,
                {"sg_bid_assets": {"type": "CustomEntity08", "id": bid_assets_id}}
            )
            logger.info(f"Linked Bid Assets {bid_assets_id} to Bid {bid_id}")
        except Exception as e:
            logger.error(f"Failed to link Bid Assets to Bid: {e}", exc_info=True)
            QtWidgets.QMessageBox.warning(
                self,
                "Warning",
                f"Bid Assets created but failed to link to Bid:\n{str(e)}"
            )

        # Step 7: Create Asset items (CustomEntity07) linked to Bid Assets
        created_count = 0
        failed_count = 0

        for index, row in df.iterrows():
            # Update progress
            progress.setLabelText(f"Processing Asset item {created_count + 1} of {len(df)}...")
            progress.setValue(1 + index + 1)
            QtWidgets.QApplication.processEvents()

            try:
                # Build SG data from mapping
                sg_data = {
                    "project": {"type": "Project", "id": self.current_project_id},
                    "sg_bid_assets": {"type": "CustomEntity08", "id": bid_assets_id}  # Link to Bid Assets
                }

                for sg_field, excel_col in column_mapping.items():
                    if excel_col is None:
                        continue

                    # Get value from DataFrame
                    value = row[excel_col]

                    # Skip empty values
                    if pd.isna(value) or value == "":
                        continue

                    # Convert value based on field type
                    field_type = ColumnMappingDialog.ASSET_ITEM_REQUIRED_FIELDS.get(sg_field)

                    if field_type == "number":
                        try:
                            sg_data[sg_field] = int(value)
                        except (ValueError, TypeError):
                            logger.warning(f"Could not convert '{value}' to number for field '{sg_field}'")
                            continue
                    elif field_type == "float":
                        try:
                            sg_data[sg_field] = float(value)
                        except (ValueError, TypeError):
                            logger.warning(f"Could not convert '{value}' to float for field '{sg_field}'")
                            continue
                    elif field_type == "checkbox":
                        # Convert to boolean
                        if isinstance(value, bool):
                            sg_data[sg_field] = value
                        elif isinstance(value, str):
                            sg_data[sg_field] = value.lower() in ["true", "yes", "1", "x"]
                        else:
                            sg_data[sg_field] = bool(value)
                    else:
                        # Text and list fields - import as-is without modification
                        # If it's already a string, use it directly to preserve exact formatting
                        if isinstance(value, str):
                            sg_data[sg_field] = value
                        else:
                            # Only convert non-string values
                            sg_data[sg_field] = str(value)

                # Validate that asset has a code
                asset_code = sg_data.get("code")
                if not asset_code:
                    logger.warning(f"Row {index}: No code specified, skipping")
                    failed_count += 1
                    continue

                # Always create new asset - each import creates fresh Asset items
                try:
                    result = self.sg_session.sg.create("CustomEntity07", sg_data)
                    created_count += 1
                    logger.info(f"Created new CustomEntity07 (Asset): {result['id']} with code '{asset_code}' linked to Bid Assets {bid_assets_id}")
                except Exception as e:
                    logger.error(f"Failed to create asset for row {index}: {e}", exc_info=True)
                    failed_count += 1

            except Exception as e:
                failed_count += 1
                logger.error(f"Failed to process Asset item for row {index}: {e}", exc_info=True)

        # Close progress dialog
        progress.setValue(total_steps)
        progress.close()

        # Log summary (detailed message shown at end of all imports)
        if failed_count > 0:
            logger.warning(f"Created Bid Assets '{bid_assets_name}' with {created_count} Asset items ({failed_count} failed)")
        else:
            logger.info(f"Successfully created Bid Assets '{bid_assets_name}' with {created_count} Asset items")

        return created_count, bid_assets_id

    def _import_scenes(self, df, column_mapping, bid_id, bid_name):
        """Import Scenes data to ShotGrid.

        TODO: Needs clarification:
        - Correct entity type (currently "Scene" in ENTITY_CONFIGS - verify if correct)
        - Parent container entity type and naming convention (e.g., "Bid Scenes-v###")
        - Field name on Bid for linking
        - Complete field mappings

        Args:
            df: DataFrame containing Scenes data
            column_mapping: Dict mapping ShotGrid fields to Excel column names
            bid_id: ID of the Bid to link to
            bid_name: Name of the Bid for versioning

        Returns:
            tuple: (Number of records created, Parent entity ID)
        """
        logger.warning("Scene import not yet fully implemented - needs entity type and linking details")
        return 0, None

    def _import_rates(self, df, column_mapping, bid_id, bid_name):
        """Import Rates data (Line Items) to ShotGrid.

        Creates a Price List (CustomEntity10) container and populates it with Line Items (CustomEntity03)
        from the Excel Rate tab. Auto-discovers all fields ending with '_mandays' from the schema.

        Args:
            df: DataFrame containing Rates data
            column_mapping: Dict mapping ShotGrid fields to Excel column names
            bid_id: ID of the Bid to link to
            bid_name: Name of the Bid for versioning

        Returns:
            tuple: (Number of Line Items created, Price List ID, Number of Line Items failed)
        """
        import pandas as pd

        if df is None or len(df) == 0:
            return 0, None, 0

        logger.info(f"Starting Rates import with {len(df)} rows for Bid: {bid_name} ({bid_id})")
        logger.info(f"Column mapping: {column_mapping}")

        # Step 1: Fetch schema for CustomEntity03 (Line Items) to auto-discover fields
        try:
            schema = self.sg_session.sg.schema_field_read("CustomEntity03")
            logger.info(f"Fetched schema for CustomEntity03 (Line Items)")

            # Build field allowlist: find all fields ending with "_mandays"
            line_items_fields = ["id", "code"]
            mandays_fields = []
            for field_name in schema.keys():
                if field_name.endswith("_mandays"):
                    mandays_fields.append(field_name)

            # Sort mandays fields alphabetically for consistent display
            mandays_fields.sort()
            line_items_fields.extend(mandays_fields)

            # Build field schema dictionary for type conversion
            field_schema = {}
            for field_name in line_items_fields:
                if field_name not in schema:
                    continue
                field_info = schema[field_name]
                field_schema[field_name] = {
                    "data_type": field_info.get("data_type", {}).get("value"),
                    "properties": field_info.get("properties", {}),
                }

            logger.info(f"Auto-discovered {len(mandays_fields)} mandays fields: {mandays_fields}")

        except Exception as e:
            logger.error(f"Failed to fetch schema for CustomEntity03: {e}", exc_info=True)
            QtWidgets.QMessageBox.critical(
                self,
                "Error",
                f"Failed to fetch Line Items schema:\n{str(e)}"
            )
            return 0, None, 0

        # Step 2: Determine version number for Price List
        # Query existing Price Lists with pattern: {bid_name}-Rates-v*
        try:
            filters = [
                ["project", "is", {"type": "Project", "id": self.current_project_id}],
                ["code", "starts_with", f"{bid_name}-Rates-v"]
            ]
            existing_price_lists = self.sg_session.sg.find(
                "CustomEntity10",
                filters,
                ["code"]
            )

            # Find the highest version number
            max_version = 0
            for price_list in existing_price_lists:
                code = price_list.get("code", "")
                # Extract version number from code (e.g., "MyBid-Rates-v003" -> 3)
                if code.startswith(f"{bid_name}-Rates-v"):
                    version_str = code.split("-v")[-1]
                    try:
                        version = int(version_str)
                        max_version = max(max_version, version)
                    except ValueError:
                        continue

            # Increment to get next version
            next_version = max_version + 1
            price_list_name = f"{bid_name}-Rates-v{next_version:03d}"

            logger.info(f"Generated Price List name: {price_list_name} (version {next_version})")

        except Exception as e:
            logger.error(f"Failed to query existing Price Lists: {e}", exc_info=True)
            # Default to v001 if query fails
            price_list_name = f"{bid_name}-Rates-v001"
            logger.warning(f"Defaulting to version 001: {price_list_name}")

        # Create progress dialog for import
        total_steps = len(df) + 2  # +2 for Price List creation and Bid linking
        progress = QtWidgets.QProgressDialog("Importing data to ShotGrid...", None, 0, total_steps, self)
        progress.setWindowModality(QtCore.Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.setValue(0)
        QtWidgets.QApplication.processEvents()

        # Step 3: Create Price List in ShotGrid
        progress.setLabelText(f"Creating Price List: {price_list_name}...")
        QtWidgets.QApplication.processEvents()

        try:
            price_list_data = {
                "code": price_list_name,
                "project": {"type": "Project", "id": self.current_project_id},
                "sg_parent_bid": {"type": "CustomEntity06", "id": bid_id}
            }
            price_list = self.sg_session.sg.create("CustomEntity10", price_list_data)
            price_list_id = price_list["id"]
            logger.info(f"Created Price List (CustomEntity10): {price_list_id} - {price_list_name}")
        except Exception as e:
            logger.error(f"Failed to create Price List: {e}", exc_info=True)
            progress.close()
            QtWidgets.QMessageBox.critical(
                self,
                "Error",
                f"Failed to create Price List:\n{str(e)}"
            )
            return 0, None, 0

        progress.setValue(1)
        QtWidgets.QApplication.processEvents()

        # Step 4: Link Price List to Bid
        progress.setLabelText(f"Linking Price List to Bid...")
        QtWidgets.QApplication.processEvents()

        try:
            self.sg_session.sg.update(
                "CustomEntity06",
                bid_id,
                {"sg_price_list": {"type": "CustomEntity10", "id": price_list_id}}
            )
            logger.info(f"Linked Price List {price_list_id} to Bid {bid_id}")
        except Exception as e:
            logger.error(f"Failed to link Price List to Bid: {e}", exc_info=True)
            # Non-fatal, continue with import

        progress.setValue(2)
        QtWidgets.QApplication.processEvents()

        # Step 5: Create Line Items (CustomEntity03) linked to Price List
        created_count = 0
        failed_count = 0
        created_line_item_ids = []

        for index, row in df.iterrows():
            # Update progress
            progress.setLabelText(f"Creating Line Item {created_count + 1} of {len(df)}...")
            progress.setValue(2 + index + 1)
            QtWidgets.QApplication.processEvents()

            try:
                # Build SG data from mapping
                sg_data = {
                    "project": {"type": "Project", "id": self.current_project_id},
                    "sg_parent_pricelist": {"type": "CustomEntity10", "id": price_list_id},
                }

                for sg_field, excel_col in column_mapping.items():
                    if excel_col is None:
                        continue

                    # Skip fields that should not be imported
                    # SG ID is auto-generated, Total Mandays and Price are calculated fields
                    if sg_field in ["id", "_calc_price"]:
                        continue

                    # Get value from DataFrame
                    value = row[excel_col]

                    # Skip empty values
                    if pd.isna(value) or value == "":
                        continue

                    # Get field type from schema
                    field_type = None
                    if sg_field in field_schema:
                        field_type = field_schema[sg_field]["data_type"]

                    # Convert value based on field type
                    if field_type == "number":
                        try:
                            sg_data[sg_field] = int(value)
                        except (ValueError, TypeError):
                            logger.warning(f"Could not convert '{value}' to number for field '{sg_field}'")
                            continue
                    elif field_type == "float":
                        try:
                            sg_data[sg_field] = float(value)
                        except (ValueError, TypeError):
                            logger.warning(f"Could not convert '{value}' to float for field '{sg_field}'")
                            continue
                    elif field_type == "checkbox":
                        # Convert to boolean
                        if isinstance(value, bool):
                            sg_data[sg_field] = value
                        elif isinstance(value, str):
                            sg_data[sg_field] = value.lower() in ["true", "yes", "1", "x"]
                        else:
                            sg_data[sg_field] = bool(value)
                    else:
                        # Text and other fields - import as-is
                        if isinstance(value, str):
                            sg_data[sg_field] = value
                        else:
                            sg_data[sg_field] = str(value)

                # Create the Line Item record
                result = self.sg_session.sg.create("CustomEntity03", sg_data)
                created_line_item_ids.append(result["id"])
                created_count += 1
                logger.info(f"Created Line Item (CustomEntity03): {result['id']} with code '{sg_data.get('code', 'N/A')}'")

            except Exception as e:
                failed_count += 1
                logger.error(f"Failed to create Line Item for row {index}: {e}", exc_info=True)

        # Step 6: Link all created Line Items to the Price List via sg_line_items field
        if created_line_item_ids:
            progress.setLabelText(f"Linking {len(created_line_item_ids)} Line Items to Price List...")
            QtWidgets.QApplication.processEvents()

            try:
                # Build entity references for all created Line Items
                line_item_refs = [{"type": "CustomEntity03", "id": item_id} for item_id in created_line_item_ids]

                self.sg_session.sg.update(
                    "CustomEntity10",
                    price_list_id,
                    {"sg_line_items": line_item_refs}
                )
                logger.info(f"Linked {len(created_line_item_ids)} Line Items to Price List {price_list_id}")
            except Exception as e:
                logger.error(f"Failed to link Line Items to Price List: {e}", exc_info=True)
                # Non-fatal, Line Items were still created

        # Close progress dialog
        progress.setValue(total_steps)
        progress.close()

        # Log summary
        if failed_count > 0:
            logger.warning(f"Created Price List '{price_list_name}' with {created_count} Line Items ({failed_count} failed)")
        else:
            logger.info(f"Successfully created Price List '{price_list_name}' with {created_count} Line Items")

        return created_count, price_list_id, failed_count

    def _refresh_vfx_breakdown_dropdown(self):
        """Refresh the VFX Breakdown dropdown in the VFX Breakdown tab.

        Returns:
            bool: True if refresh was successful, False otherwise
        """
        try:
            # Navigate to VFX Breakdown tab via parent
            if hasattr(self.parent_app, 'vfx_breakdown_tab'):
                vfx_tab = self.parent_app.vfx_breakdown_tab

                # Call populate_vfx_breakdown_combo to refresh
                if hasattr(vfx_tab, 'populate_vfx_breakdown_combo'):
                    # Get current RFQ if available
                    rfq = getattr(vfx_tab, 'current_rfq', None) or self.current_rfq
                    vfx_tab.populate_vfx_breakdown_combo(rfq, auto_select=False)
                    logger.info("VFX Breakdown dropdown refreshed successfully")
                    return True
                else:
                    logger.warning("VFX Breakdown tab does not have populate_vfx_breakdown_combo method")
            else:
                logger.warning("Parent app does not have vfx_breakdown_tab attribute")
        except Exception as e:
            logger.error(f"Failed to refresh VFX Breakdown dropdown: {e}", exc_info=True)

        return False

    def _refresh_price_lists_dropdown(self):
        """Refresh the Price Lists dropdown in the Rates tab.

        Returns:
            bool: True if refresh was successful, False otherwise
        """
        try:
            # Navigate to Rates tab via parent
            if hasattr(self.parent_app, 'rates_tab'):
                rates_tab = self.parent_app.rates_tab

                # Call _refresh_price_lists to refresh
                if hasattr(rates_tab, '_refresh_price_lists'):
                    rates_tab._refresh_price_lists()
                    logger.info("Price Lists dropdown refreshed successfully")
                    return True
                else:
                    logger.warning("Rates tab does not have _refresh_price_lists method")
            else:
                logger.warning("Parent app does not have rates_tab attribute")
        except Exception as e:
            logger.error(f"Failed to refresh Price Lists dropdown: {e}", exc_info=True)

        return False

    def _on_refresh_bids(self):
        """Handle Refresh button click."""
        self._refresh_bids()

    def _refresh_bids(self):
        """Refresh the bid list from ShotGrid."""
        # Get current selections to restore after refresh
        current_bid_id = None
        current_bid = self.get_current_bid()
        if current_bid:
            current_bid_id = current_bid.get('id')

        # Repopulate bids using stored RFQ and project
        self.populate_bids(self.current_rfq, self.current_project_id, auto_select=False)

        # Restore selection if possible
        if current_bid_id:
            self._select_bid_by_id(current_bid_id)

        self.statusMessageChanged.emit("✓ Bid list refreshed", False)
        logger.info("Bid list refreshed")

    def _set_status(self, message, is_error=False):
        """Set the status message.

        Args:
            message: Status message to display
            is_error: Whether this is an error message (changes color)
        """
        color = "#ff8080" if is_error else "#a0a0a0"
        self.status_label.setStyleSheet(f"color: {color}; padding: 2px 0;")
        self.status_label.setText(message)
        self.statusMessageChanged.emit(message, is_error)
