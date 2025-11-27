from PySide6 import QtWidgets, QtCore, QtGui
import logging
import os
import tempfile

try:
    from .logger import logger
    from .sliding_overlay_panel import SlidingOverlayPanel
    from .bid_selector_widget import CollapsibleGroupBox
    from .thumbnail_cache import ThumbnailCache
    from .settings import AppSettings
    from .document_folder_pane_widget import DocumentFolderPaneWidget
except (ImportError, ValueError, SystemError):
    logger = logging.getLogger("FFPackageManager")
    from sliding_overlay_panel import SlidingOverlayPanel
    from bid_selector_widget import CollapsibleGroupBox
    from thumbnail_cache import ThumbnailCache
    from settings import AppSettings
    from document_folder_pane_widget import DocumentFolderPaneWidget

# Global document cache instance (shared across all widgets)
_document_cache = None


def get_document_cache():
    """Get or create the global document cache instance."""
    global _document_cache
    if _document_cache is None:
        settings = AppSettings()
        cache_path = settings.get_thumbnail_cache_path()
        max_age = settings.get_thumbnail_cache_max_age_days()
        _document_cache = ThumbnailCache(cache_path, max_age)
    return _document_cache


def reset_document_cache():
    """Reset the global document cache instance."""
    global _document_cache
    _document_cache = None


class UploadDocumentTypeDialog(QtWidgets.QDialog):
    """Dialog for selecting the type of document to upload."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.selected_type = None
        self.setWindowTitle("Select Document Type")
        self.setModal(True)
        self._setup_ui()

    def _setup_ui(self):
        """Setup the dialog UI."""
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(10)

        # Title
        title_label = QtWidgets.QLabel("Select the type of document you're uploading:")
        title_label.setStyleSheet("font-weight: bold; font-size: 12px;")
        layout.addWidget(title_label)

        # Type buttons
        button_layout = QtWidgets.QVBoxLayout()
        button_layout.setSpacing(5)

        types = [
            ("Script", "script"),
            ("Documents", "document"),
        ]

        for label, type_value in types:
            btn = QtWidgets.QPushButton(label)
            btn.setMinimumHeight(40)
            btn.setStyleSheet("""
                QPushButton {
                    background-color: #3a3a3a;
                    color: white;
                    border: 1px solid #555;
                    border-radius: 4px;
                    padding: 8px;
                    font-size: 11px;
                }
                QPushButton:hover {
                    background-color: #4a9eff;
                    border: 1px solid #5eb3ff;
                }
            """)
            btn.clicked.connect(lambda checked, t=type_value: self._on_type_selected(t))
            button_layout.addWidget(btn)

        layout.addLayout(button_layout)

        # Cancel button
        cancel_btn = QtWidgets.QPushButton("Cancel")
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #555;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px;
            }
            QPushButton:hover {
                background-color: #666;
            }
        """)
        layout.addWidget(cancel_btn)

        self.setMinimumWidth(300)

    def _on_type_selected(self, type_value):
        """Handle type selection."""
        self.selected_type = type_value
        self.accept()

    def get_selected_type(self):
        """Get the selected type."""
        return self.selected_type


class DocumentPoller(QtCore.QObject):
    """Background poller that checks ShotGrid for document availability and updates widgets."""

    documentReady = QtCore.Signal(int, dict)  # (version_id, updated_version_data)

    def __init__(self, sg_session, viewer_widget, parent=None):
        super().__init__(parent)
        self.sg_session = sg_session
        self.viewer_widget = viewer_widget
        self.pending_versions = {}  # version_id -> attempt_count
        self.poll_timer = QtCore.QTimer(self)
        self.poll_timer.timeout.connect(self._check_pending_documents)
        self.poll_timer.setInterval(2000)

    def add_version_to_poll(self, version_id):
        """Add a version to poll for document availability."""
        self.pending_versions[version_id] = 0
        if not self.poll_timer.isActive():
            self.poll_timer.start()

    def _check_pending_documents(self):
        """Check all pending versions for document availability."""
        if not self.pending_versions:
            self.poll_timer.stop()
            return

        versions_to_remove = []

        for version_id, attempt_count in list(self.pending_versions.items()):
            if attempt_count >= 30:
                logger.warning(f"Giving up on document for version {version_id} after 30 attempts")
                versions_to_remove.append(version_id)
                continue

            try:
                version_data = self.sg_session.sg.find_one(
                    'Version',
                    [['id', 'is', version_id]],
                    ['code', 'image', 'sg_version_type', 'created_at', 'project', 'sg_uploaded_movie']
                )

                if version_data:
                    uploaded_movie = version_data.get('sg_uploaded_movie')
                    has_file = False

                    if isinstance(uploaded_movie, dict):
                        has_file = bool(uploaded_movie.get('url') or uploaded_movie.get('name'))
                    elif isinstance(uploaded_movie, str):
                        has_file = bool(uploaded_movie)

                    if has_file:
                        thumbnail_widget = self.viewer_widget.find_thumbnail_by_version_id(version_id)
                        if thumbnail_widget:
                            thumbnail_widget.refresh_thumbnail(version_data)
                        self.documentReady.emit(version_id, version_data)
                        versions_to_remove.append(version_id)
                    else:
                        self.pending_versions[version_id] = attempt_count + 1

            except Exception as e:
                logger.error(f"Error checking document for version {version_id}: {e}", exc_info=True)
                self.pending_versions[version_id] = attempt_count + 1

        for version_id in versions_to_remove:
            self.pending_versions.pop(version_id, None)

        if not self.pending_versions:
            self.poll_timer.stop()


class DocumentViewerDialog(QtWidgets.QDialog):
    """Dialog for viewing a document with zoom and pan capabilities."""

    def __init__(self, version_data, sg_session, parent=None):
        super().__init__(parent)
        self.version_data = version_data
        self.sg_session = sg_session
        self.current_zoom = 1.0
        self.document_pixmap = None
        self.current_page = 0
        self.total_pages = 1
        self.document_path = None

        version_code = version_data.get('code', 'Document Viewer')
        self.setWindowTitle(f"Document Viewer - {version_code}")
        self.resize(1200, 800)

        self._setup_ui()
        self._load_document()

    def _setup_ui(self):
        """Setup the dialog UI."""
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(5, 5, 5, 5)

        # Toolbar
        toolbar = QtWidgets.QHBoxLayout()

        # Page navigation (for multi-page documents)
        self.prev_page_btn = QtWidgets.QPushButton("< Prev")
        self.prev_page_btn.clicked.connect(self._prev_page)
        self.prev_page_btn.setEnabled(False)
        toolbar.addWidget(self.prev_page_btn)

        # Page slider for quick navigation
        self.page_slider = QtWidgets.QSlider(QtCore.Qt.Horizontal)
        self.page_slider.setMinimum(1)
        self.page_slider.setMaximum(1)
        self.page_slider.setValue(1)
        self.page_slider.setFixedWidth(150)
        self.page_slider.setEnabled(False)
        self.page_slider.valueChanged.connect(self._on_page_slider_changed)
        toolbar.addWidget(self.page_slider)

        self.page_label = QtWidgets.QLabel("Page 1 of 1")
        toolbar.addWidget(self.page_label)

        self.next_page_btn = QtWidgets.QPushButton("Next >")
        self.next_page_btn.clicked.connect(self._next_page)
        self.next_page_btn.setEnabled(False)
        toolbar.addWidget(self.next_page_btn)

        toolbar.addSpacing(20)

        # Zoom controls
        zoom_in_btn = QtWidgets.QPushButton("Zoom In (+)")
        zoom_in_btn.clicked.connect(self._zoom_in)
        toolbar.addWidget(zoom_in_btn)

        zoom_out_btn = QtWidgets.QPushButton("Zoom Out (-)")
        zoom_out_btn.clicked.connect(self._zoom_out)
        toolbar.addWidget(zoom_out_btn)

        zoom_reset_btn = QtWidgets.QPushButton("Reset (0)")
        zoom_reset_btn.clicked.connect(self._zoom_reset)
        toolbar.addWidget(zoom_reset_btn)

        fit_btn = QtWidgets.QPushButton("Fit to Window")
        fit_btn.clicked.connect(self._fit_to_window)
        toolbar.addWidget(fit_btn)

        toolbar.addStretch()

        # Zoom level label
        self.zoom_label = QtWidgets.QLabel("100%")
        self.zoom_label.setStyleSheet("font-weight: bold;")
        toolbar.addWidget(self.zoom_label)

        # Toggle details pane button
        self.toggle_details_btn = QtWidgets.QPushButton("Details >")
        self.toggle_details_btn.setToolTip("Show/Hide Details Panel")
        self.toggle_details_btn.clicked.connect(self._toggle_details_pane)
        toolbar.addWidget(self.toggle_details_btn)

        layout.addLayout(toolbar)

        # Container widget for document view
        self.view_container = QtWidgets.QWidget()
        view_container_layout = QtWidgets.QVBoxLayout(self.view_container)
        view_container_layout.setContentsMargins(0, 0, 0, 0)

        # Graphics view for displaying document
        self.graphics_view = QtWidgets.QGraphicsView()
        self.graphics_scene = QtWidgets.QGraphicsScene()
        self.graphics_view.setScene(self.graphics_scene)
        self.graphics_view.setRenderHint(QtGui.QPainter.Antialiasing)
        self.graphics_view.setRenderHint(QtGui.QPainter.SmoothPixmapTransform)
        self.graphics_view.setDragMode(QtWidgets.QGraphicsView.ScrollHandDrag)
        self.graphics_view.setBackgroundBrush(QtGui.QBrush(QtGui.QColor(30, 30, 30)))

        self.graphics_view.wheelEvent = self._wheel_event

        view_container_layout.addWidget(self.graphics_view)
        layout.addWidget(self.view_container)

        # Create sliding overlay panel for details
        self.details_panel = SlidingOverlayPanel(
            parent=self.view_container,
            panel_width=350,
            animation_duration=250
        )
        self.details_panel.set_title("Document Details")

        details_content = self._create_details_content()
        self.details_panel.set_content(details_content)

        self.details_panel.panel_shown.connect(self._on_details_shown)
        self.details_panel.panel_hidden.connect(self._on_details_hidden)

        # Status bar
        self.status_label = QtWidgets.QLabel("Loading...")
        self.status_label.setStyleSheet("padding: 5px; background-color: #2b2b2b;")
        layout.addWidget(self.status_label)

        # Close button
        button_layout = QtWidgets.QHBoxLayout()
        button_layout.addStretch()
        close_btn = QtWidgets.QPushButton("Close")
        close_btn.clicked.connect(self.accept)
        button_layout.addWidget(close_btn)
        layout.addLayout(button_layout)

    def _create_details_content(self):
        """Create the details content for the sliding panel."""
        details_widget = QtWidgets.QWidget()
        details_layout = QtWidgets.QVBoxLayout(details_widget)
        details_layout.setContentsMargins(0, 0, 0, 0)

        self.details_text = QtWidgets.QTextEdit()
        self.details_text.setReadOnly(True)
        self.details_text.setStyleSheet("""
            QTextEdit {
                background-color: #2b2b2b;
                border: none;
                color: #ddd;
            }
        """)
        details_layout.addWidget(self.details_text)

        self._update_details()

        return details_widget

    def _toggle_details_pane(self):
        """Toggle the visibility of the details pane."""
        self.details_panel.toggle()

    def _on_details_shown(self):
        """Handle details panel shown event."""
        self.toggle_details_btn.setText("Details <")

    def _on_details_hidden(self):
        """Handle details panel hidden event."""
        self.toggle_details_btn.setText("Details >")

    def _update_details(self):
        """Update the details panel with version information."""
        details = f"<b>Document:</b> {self.version_data.get('code', 'Unknown')}<br>"
        details += f"<b>ID:</b> {self.version_data.get('id', 'N/A')}<br>"

        version_type = self.version_data.get('sg_version_type', '')
        if isinstance(version_type, dict):
            version_type = version_type.get('name', 'Unknown')
        details += f"<b>Type:</b> {version_type}<br>"

        status = self.version_data.get('sg_status_list', 'N/A')
        details += f"<b>Status:</b> {status}<br>"

        entity = self.version_data.get('entity', {})
        if entity:
            details += f"<b>Asset:</b> {entity.get('name', 'Unknown')}<br>"

        task = self.version_data.get('sg_task', {})
        if task:
            details += f"<b>Task:</b> {task.get('name', 'Unknown')}<br>"

        user = self.version_data.get('user', {})
        if user:
            details += f"<b>Created By:</b> {user.get('name', 'Unknown')}<br>"

        created_at = self.version_data.get('created_at', 'N/A')
        details += f"<b>Created:</b> {created_at}<br>"

        updated_at = self.version_data.get('updated_at', 'N/A')
        details += f"<b>Updated:</b> {updated_at}<br>"

        description = self.version_data.get('description', '')
        if description:
            details += f"<br><b>Description:</b><br>{description}"

        self.details_text.setHtml(details)

    def _load_document(self):
        """Load the document from ShotGrid."""
        try:
            document_url = None
            document_name = None

            uploaded_movie = self.version_data.get('sg_uploaded_movie')
            if uploaded_movie:
                if isinstance(uploaded_movie, dict):
                    document_url = uploaded_movie.get('url')
                    document_name = uploaded_movie.get('name', '')
                elif isinstance(uploaded_movie, str):
                    document_url = uploaded_movie

            if not document_url:
                self.status_label.setText("No document available")
                return

            # Download document to temp file
            from PySide6.QtCore import QThread, QObject, Signal
            import requests

            class DocumentLoader(QObject):
                finished = Signal(str, bytes)  # (filename, data)
                error = Signal(str)

                def __init__(self, url, filename):
                    super().__init__()
                    self.url = url
                    self.filename = filename

                def run(self):
                    try:
                        response = requests.get(self.url, timeout=60)
                        if response.status_code == 200:
                            self.finished.emit(self.filename, response.content)
                        else:
                            self.error.emit(f"HTTP {response.status_code}")
                    except Exception as e:
                        logger.error(f"Failed to download document: {e}")
                        self.error.emit(str(e))

            self.loader = DocumentLoader(document_url, document_name or 'document')
            self.loader_thread = QThread()
            self.loader.moveToThread(self.loader_thread)
            self.loader_thread.started.connect(self.loader.run)
            self.loader.finished.connect(self._on_document_loaded)
            self.loader.error.connect(self._on_document_error)
            self.loader_thread.start()

        except Exception as e:
            logger.error(f"Error loading document: {e}")
            self.status_label.setText(f"Error: {str(e)}")

    def _on_document_loaded(self, filename, document_data):
        """Handle document loaded."""
        try:
            # Save to temp file
            ext = os.path.splitext(filename)[1].lower() if filename else '.pdf'
            temp_file = tempfile.NamedTemporaryFile(suffix=ext, delete=False)
            temp_file.write(document_data)
            temp_file.close()
            self.document_path = temp_file.name

            # Determine file type and render
            if ext in ['.pdf']:
                self._render_pdf()
            elif ext in ['.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff', '.tif']:
                self._render_image(document_data)
            elif ext in ['.txt', '.md', '.json', '.xml', '.html', '.css', '.js', '.py']:
                self._render_text(document_data)
            elif ext in ['.xls', '.xlsx']:
                self._render_excel(filename)
            else:
                # For other documents (doc, docx), show file info
                self._render_document_placeholder(filename, len(document_data))

            # Update status with hints
            size_kb = len(document_data) / 1024
            hint = ""
            if self.total_pages > 1:
                hint = " | Scroll to change pages, Ctrl+Scroll to zoom"
            else:
                hint = " | Ctrl+Scroll to zoom"
            self.status_label.setText(
                f"Document: {filename} | Size: {size_kb:.1f} KB | "
                f"Version: {self.version_data.get('code', 'Unknown')}{hint}"
            )

            if hasattr(self, 'loader_thread'):
                self.loader_thread.quit()
                self.loader_thread.wait()

        except Exception as e:
            logger.error(f"Error displaying document: {e}")
            self.status_label.setText(f"Error: {str(e)}")

    def _render_pdf(self):
        """Render PDF document using PyMuPDF if available, otherwise show placeholder."""
        try:
            import fitz  # PyMuPDF
            doc = fitz.open(self.document_path)
            self.total_pages = len(doc)
            self.current_page = 0

            self._update_page_controls()
            self._render_pdf_page(doc, 0)
            doc.close()

        except ImportError:
            logger.warning("PyMuPDF not installed, showing PDF placeholder")
            self._render_document_placeholder("PDF Document", 0, is_pdf=True)
        except Exception as e:
            logger.error(f"Error rendering PDF: {e}")
            self._render_document_placeholder("PDF Document", 0, is_pdf=True)

    def _render_pdf_page(self, doc, page_num, preserve_zoom=False):
        """Render a specific page of the PDF.

        Args:
            doc: PyMuPDF document object
            page_num: Page number to render (0-indexed)
            preserve_zoom: If True, maintain current zoom level instead of fitting to window
        """
        try:
            import fitz  # PyMuPDF
            page = doc.load_page(page_num)
            # Render at 2x resolution for better quality
            mat = fitz.Matrix(2.0, 2.0)
            pix = page.get_pixmap(matrix=mat)

            # Convert to QPixmap
            img_data = pix.tobytes("ppm")
            self.document_pixmap = QtGui.QPixmap()
            self.document_pixmap.loadFromData(img_data)

            if not self.document_pixmap.isNull():
                # Save current transform if preserving zoom
                current_transform = self.graphics_view.transform() if preserve_zoom else None

                self.graphics_scene.clear()
                self.pixmap_item = self.graphics_scene.addPixmap(self.document_pixmap)

                if preserve_zoom and current_transform is not None:
                    # Restore the zoom level
                    self.graphics_view.setTransform(current_transform)
                    # Center on the new page
                    self.graphics_view.centerOn(self.pixmap_item)
                else:
                    self._fit_to_window()

        except Exception as e:
            logger.error(f"Error rendering PDF page {page_num}: {e}")

    def _render_image(self, image_data):
        """Render image document."""
        self.document_pixmap = QtGui.QPixmap()
        self.document_pixmap.loadFromData(image_data)

        if not self.document_pixmap.isNull():
            self.graphics_scene.clear()
            self.pixmap_item = self.graphics_scene.addPixmap(self.document_pixmap)
            self._fit_to_window()

    def _render_text(self, text_data):
        """Render text document."""
        try:
            text = text_data.decode('utf-8', errors='replace')

            # Create a text display widget
            text_edit = QtWidgets.QTextEdit()
            text_edit.setReadOnly(True)
            text_edit.setPlainText(text)
            text_edit.setStyleSheet("""
                QTextEdit {
                    background-color: #1e1e1e;
                    color: #ddd;
                    font-family: 'Courier New', monospace;
                    font-size: 12px;
                    border: none;
                }
            """)

            # Use a QGraphicsProxyWidget to display in the scene
            self.graphics_scene.clear()
            proxy = self.graphics_scene.addWidget(text_edit)
            proxy.setMinimumSize(800, 600)

            self.total_pages = 1
            self._update_page_controls()

        except Exception as e:
            logger.error(f"Error rendering text: {e}")

    def _render_excel(self, filename):
        """Render Excel spreadsheet document.

        Supports both .xls (xlrd) and .xlsx (openpyxl) formats.
        Displays sheets in tabs with full spreadsheet content.
        """
        ext = os.path.splitext(filename)[1].lower() if filename else ''

        try:
            if ext == '.xlsx':
                self._render_excel_xlsx()
            else:  # .xls
                self._render_excel_xls()
        except ImportError as e:
            logger.warning(f"Excel library not available: {e}")
            self._render_excel_placeholder(filename, ext)
        except Exception as e:
            logger.error(f"Error rendering Excel file: {e}")
            self._render_excel_placeholder(filename, ext)

    def _render_excel_xlsx(self):
        """Render .xlsx file using openpyxl."""
        import openpyxl

        workbook = openpyxl.load_workbook(self.document_path, data_only=True)
        sheet_names = workbook.sheetnames

        self.total_pages = len(sheet_names)
        self.current_page = 0
        self._excel_workbook = workbook
        self._excel_sheet_names = sheet_names

        self._update_page_controls()
        self._render_excel_sheet(0)

    def _render_excel_xls(self):
        """Render .xls file using xlrd."""
        import xlrd

        workbook = xlrd.open_workbook(self.document_path)
        sheet_names = workbook.sheet_names()

        self.total_pages = len(sheet_names)
        self.current_page = 0
        self._excel_workbook = workbook
        self._excel_sheet_names = sheet_names
        self._excel_is_xlrd = True

        self._update_page_controls()
        self._render_excel_sheet(0)

    def _render_excel_sheet(self, sheet_index):
        """Render a specific sheet from the Excel workbook.

        Args:
            sheet_index: Index of the sheet to render (0-based)
        """
        self.graphics_scene.clear()

        # Create container widget for the spreadsheet
        container = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(container)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(5)

        # Sheet name header
        sheet_name = self._excel_sheet_names[sheet_index]
        header = QtWidgets.QLabel(f"Sheet: {sheet_name}")
        header.setStyleSheet("""
            QLabel {
                color: #27ae60;
                font-size: 14px;
                font-weight: bold;
                padding: 5px;
                background-color: #1e3a1e;
                border-radius: 4px;
            }
        """)
        layout.addWidget(header)

        # Create table widget for spreadsheet data
        table = QtWidgets.QTableWidget()
        table.setStyleSheet("""
            QTableWidget {
                background-color: #1e1e1e;
                color: #ddd;
                gridline-color: #444;
                font-family: 'Segoe UI', sans-serif;
                font-size: 11px;
                border: 1px solid #444;
            }
            QTableWidget::item {
                padding: 4px 8px;
                border-bottom: 1px solid #333;
            }
            QTableWidget::item:selected {
                background-color: #264f78;
            }
            QHeaderView::section {
                background-color: #2d2d2d;
                color: #ddd;
                padding: 6px;
                border: 1px solid #444;
                font-weight: bold;
            }
            QTableCornerButton::section {
                background-color: #2d2d2d;
                border: 1px solid #444;
            }
        """)

        # Get sheet data based on format type
        if hasattr(self, '_excel_is_xlrd') and self._excel_is_xlrd:
            # xlrd for .xls files
            sheet = self._excel_workbook.sheet_by_index(sheet_index)
            rows = sheet.nrows
            cols = sheet.ncols

            table.setRowCount(rows)
            table.setColumnCount(cols)

            for row in range(rows):
                for col in range(cols):
                    cell_value = sheet.cell_value(row, col)
                    cell_type = sheet.cell_type(row, col)

                    # Format cell value based on type
                    if cell_type == 3:  # Date
                        try:
                            import xlrd
                            date_tuple = xlrd.xldate_as_tuple(cell_value, self._excel_workbook.datemode)
                            display_value = f"{date_tuple[0]:04d}-{date_tuple[1]:02d}-{date_tuple[2]:02d}"
                        except:
                            display_value = str(cell_value)
                    elif isinstance(cell_value, float) and cell_value == int(cell_value):
                        display_value = str(int(cell_value))
                    else:
                        display_value = str(cell_value) if cell_value != '' else ''

                    item = QtWidgets.QTableWidgetItem(display_value)
                    item.setFlags(item.flags() & ~QtCore.Qt.ItemIsEditable)
                    table.setItem(row, col, item)
        else:
            # openpyxl for .xlsx files
            sheet = self._excel_workbook[sheet_name]

            # Get dimensions
            rows = sheet.max_row or 1
            cols = sheet.max_column or 1

            table.setRowCount(rows)
            table.setColumnCount(cols)

            for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=rows, max_col=cols)):
                for col_idx, cell in enumerate(row):
                    cell_value = cell.value
                    if cell_value is None:
                        display_value = ''
                    elif isinstance(cell_value, float) and cell_value == int(cell_value):
                        display_value = str(int(cell_value))
                    else:
                        display_value = str(cell_value)

                    item = QtWidgets.QTableWidgetItem(display_value)
                    item.setFlags(item.flags() & ~QtCore.Qt.ItemIsEditable)
                    table.setItem(row_idx, col_idx, item)

        # Set column headers (A, B, C, ...)
        headers = []
        for i in range(table.columnCount()):
            if i < 26:
                headers.append(chr(65 + i))  # A-Z
            else:
                # AA, AB, etc. for columns beyond Z
                headers.append(chr(65 + i // 26 - 1) + chr(65 + i % 26))
        table.setHorizontalHeaderLabels(headers)

        # Resize columns to fit content
        table.resizeColumnsToContents()

        # Set minimum column width
        for col in range(table.columnCount()):
            if table.columnWidth(col) < 60:
                table.setColumnWidth(col, 60)
            elif table.columnWidth(col) > 300:
                table.setColumnWidth(col, 300)

        # Add table with stretch to fill available space
        layout.addWidget(table, 1)  # stretch factor 1

        # Add sheet info
        info_label = QtWidgets.QLabel(f"Rows: {table.rowCount()} | Columns: {table.columnCount()}")
        info_label.setStyleSheet("color: #888; font-size: 10px; padding: 2px;")
        layout.addWidget(info_label)

        container.setStyleSheet("background-color: #2b2b2b;")

        # Add to scene with size matching the dialog
        proxy = self.graphics_scene.addWidget(container)

        # Get the dialog's view size for proper sizing
        view_size = self.graphics_view.viewport().size()
        # Use dialog size or fallback to reasonable defaults
        width = max(view_size.width(), 1100) if view_size.width() > 100 else 1100
        height = max(view_size.height(), 700) if view_size.height() > 100 else 700

        proxy.setMinimumSize(width, height)
        proxy.resize(width, height)

        # Set scene rect to match
        self.graphics_scene.setSceneRect(0, 0, width, height)

        # Update page label to show sheet name
        self.page_label.setText(f"Sheet {sheet_index + 1} of {self.total_pages}: {sheet_name}")

    def _render_excel_placeholder(self, filename, ext):
        """Render placeholder when Excel libraries are not available."""
        self.graphics_scene.clear()

        placeholder = QtWidgets.QWidget()
        placeholder.setMinimumSize(600, 400)
        layout = QtWidgets.QVBoxLayout(placeholder)
        layout.setAlignment(QtCore.Qt.AlignCenter)

        # Excel icon
        icon_label = QtWidgets.QLabel("XLS")
        icon_label.setAlignment(QtCore.Qt.AlignCenter)
        icon_label.setStyleSheet("""
            QLabel {
                background-color: #27ae60;
                color: white;
                font-size: 48px;
                font-weight: bold;
                padding: 30px 40px;
                border-radius: 10px;
            }
        """)
        layout.addWidget(icon_label)

        # Filename
        name_label = QtWidgets.QLabel(filename or "Excel Document")
        name_label.setStyleSheet("font-size: 16px; color: #ddd; margin-top: 20px;")
        name_label.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(name_label)

        # Install instructions
        if ext == '.xlsx':
            install_msg = "Install openpyxl to preview: pip install openpyxl"
        else:
            install_msg = "Install xlrd to preview: pip install xlrd"

        info_label = QtWidgets.QLabel(install_msg)
        info_label.setStyleSheet("font-size: 11px; color: #888; margin-top: 10px;")
        info_label.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(info_label)

        placeholder.setStyleSheet("background-color: #2b2b2b;")
        proxy = self.graphics_scene.addWidget(placeholder)
        self.graphics_view.fitInView(proxy, QtCore.Qt.KeepAspectRatio)

    def _render_document_placeholder(self, filename, file_size, is_pdf=False):
        """Render a placeholder for unsupported document types."""
        self.graphics_scene.clear()

        # Create placeholder widget
        placeholder = QtWidgets.QWidget()
        placeholder.setMinimumSize(600, 400)
        layout = QtWidgets.QVBoxLayout(placeholder)
        layout.setAlignment(QtCore.Qt.AlignCenter)

        # Document icon
        icon_label = QtWidgets.QLabel()
        icon_label.setAlignment(QtCore.Qt.AlignCenter)

        # Get appropriate icon based on file extension
        ext = os.path.splitext(filename)[1].lower() if filename else ''
        if ext in ['.pdf']:
            icon_text = "PDF"
            icon_color = "#e74c3c"
        elif ext in ['.doc', '.docx']:
            icon_text = "DOC"
            icon_color = "#3498db"
        elif ext in ['.xls', '.xlsx']:
            icon_text = "XLS"
            icon_color = "#27ae60"
        elif ext in ['.txt']:
            icon_text = "TXT"
            icon_color = "#95a5a6"
        else:
            icon_text = "FILE"
            icon_color = "#9b59b6"

        icon_label.setText(icon_text)
        icon_label.setStyleSheet(f"""
            QLabel {{
                background-color: {icon_color};
                color: white;
                font-size: 48px;
                font-weight: bold;
                padding: 30px 40px;
                border-radius: 10px;
            }}
        """)
        layout.addWidget(icon_label)

        # Filename label
        name_label = QtWidgets.QLabel(filename or "Document")
        name_label.setStyleSheet("font-size: 16px; color: #ddd; margin-top: 20px;")
        name_label.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(name_label)

        # File size if available
        if file_size > 0:
            size_label = QtWidgets.QLabel(f"Size: {file_size / 1024:.1f} KB")
            size_label.setStyleSheet("font-size: 12px; color: #888;")
            size_label.setAlignment(QtCore.Qt.AlignCenter)
            layout.addWidget(size_label)

        # Info message
        if is_pdf:
            info_label = QtWidgets.QLabel("Install PyMuPDF (fitz) to preview PDFs")
        else:
            info_label = QtWidgets.QLabel("Preview not available for this file type")
        info_label.setStyleSheet("font-size: 11px; color: #666; margin-top: 10px;")
        info_label.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(info_label)

        placeholder.setStyleSheet("background-color: #2b2b2b;")

        proxy = self.graphics_scene.addWidget(placeholder)
        self.graphics_view.fitInView(proxy, QtCore.Qt.KeepAspectRatio)

    def _update_page_controls(self):
        """Update page navigation controls."""
        self.page_label.setText(f"Page {self.current_page + 1} of {self.total_pages}")
        self.prev_page_btn.setEnabled(self.current_page > 0)
        self.next_page_btn.setEnabled(self.current_page < self.total_pages - 1)

        # Update slider
        self.page_slider.blockSignals(True)
        self.page_slider.setMaximum(self.total_pages)
        self.page_slider.setValue(self.current_page + 1)
        self.page_slider.setEnabled(self.total_pages > 1)
        self.page_slider.blockSignals(False)

    def _on_page_slider_changed(self, value):
        """Handle page slider value change."""
        new_page = value - 1  # Slider is 1-based, pages are 0-based
        if new_page != self.current_page and 0 <= new_page < self.total_pages:
            self.current_page = new_page
            self._reload_current_page()

    def _prev_page(self):
        """Go to previous page."""
        if self.current_page > 0:
            self.current_page -= 1
            self._reload_current_page()

    def _next_page(self):
        """Go to next page."""
        if self.current_page < self.total_pages - 1:
            self.current_page += 1
            self._reload_current_page()

    def _reload_current_page(self):
        """Reload the current page while preserving zoom level."""
        self._update_page_controls()
        if self.document_path:
            ext = os.path.splitext(self.document_path)[1].lower()
            if ext == '.pdf':
                try:
                    import fitz
                    doc = fitz.open(self.document_path)
                    self._render_pdf_page(doc, self.current_page, preserve_zoom=True)
                    doc.close()
                except Exception as e:
                    logger.error(f"Error reloading page: {e}")
            elif ext in ['.xls', '.xlsx']:
                # Reload Excel sheet
                if hasattr(self, '_excel_workbook') and self._excel_workbook:
                    try:
                        self._render_excel_sheet(self.current_page)
                    except Exception as e:
                        logger.error(f"Error reloading Excel sheet: {e}")

    def _on_document_error(self, error_msg):
        """Handle document load error."""
        self.status_label.setText(f"Failed to load document: {error_msg}")
        if hasattr(self, 'loader_thread'):
            self.loader_thread.quit()
            self.loader_thread.wait()

    def _wheel_event(self, event):
        """Handle mouse wheel for page navigation and zooming.

        - Regular scroll: Navigate pages (for multi-page PDFs)
        - Ctrl+scroll: Zoom in/out
        """
        # Check if Ctrl is pressed for zooming
        if event.modifiers() & QtCore.Qt.ControlModifier:
            zoom_factor = 1.15

            if event.angleDelta().y() > 0:
                self.graphics_view.scale(zoom_factor, zoom_factor)
                self.current_zoom *= zoom_factor
            else:
                self.graphics_view.scale(1 / zoom_factor, 1 / zoom_factor)
                self.current_zoom /= zoom_factor

            self._update_zoom_label()
        else:
            # Regular scroll for page navigation (multi-page PDFs)
            if self.total_pages > 1:
                if event.angleDelta().y() < 0:
                    # Scroll down - next page
                    self._next_page()
                elif event.angleDelta().y() > 0:
                    # Scroll up - previous page
                    self._prev_page()
            else:
                # Single page - allow normal scrolling/panning
                # Pass to default handler
                pass

    def _zoom_in(self):
        """Zoom in the document."""
        zoom_factor = 1.25
        self.graphics_view.scale(zoom_factor, zoom_factor)
        self.current_zoom *= zoom_factor
        self._update_zoom_label()

    def _zoom_out(self):
        """Zoom out the document."""
        zoom_factor = 1.25
        self.graphics_view.scale(1 / zoom_factor, 1 / zoom_factor)
        self.current_zoom /= zoom_factor
        self._update_zoom_label()

    def _zoom_reset(self):
        """Reset zoom to 100%."""
        self.graphics_view.resetTransform()
        self.current_zoom = 1.0
        self._update_zoom_label()

    def _fit_to_window(self):
        """Fit document to window."""
        if self.document_pixmap and not self.document_pixmap.isNull():
            self.graphics_view.fitInView(
                self.graphics_scene.sceneRect(),
                QtCore.Qt.KeepAspectRatio
            )
            view_rect = self.graphics_view.viewport().rect()
            scene_rect = self.graphics_scene.sceneRect()

            x_ratio = view_rect.width() / scene_rect.width() if scene_rect.width() > 0 else 1
            y_ratio = view_rect.height() / scene_rect.height() if scene_rect.height() > 0 else 1
            self.current_zoom = min(x_ratio, y_ratio)
            self._update_zoom_label()

    def _update_zoom_label(self):
        """Update the zoom level label."""
        zoom_percent = int(self.current_zoom * 100)
        self.zoom_label.setText(f"{zoom_percent}%")

    def showEvent(self, event):
        """Handle dialog show event."""
        super().showEvent(event)
        if self.document_pixmap and not self.document_pixmap.isNull():
            QtCore.QTimer.singleShot(0, self._fit_to_window)

    def keyPressEvent(self, event):
        """Handle keyboard shortcuts."""
        if event.key() == QtCore.Qt.Key_Plus or event.key() == QtCore.Qt.Key_Equal:
            self._zoom_in()
        elif event.key() == QtCore.Qt.Key_Minus:
            self._zoom_out()
        elif event.key() == QtCore.Qt.Key_0:
            self._zoom_reset()
        elif event.key() == QtCore.Qt.Key_Escape:
            self.accept()
        elif event.key() == QtCore.Qt.Key_Left:
            self._prev_page()
        elif event.key() == QtCore.Qt.Key_Right:
            self._next_page()
        else:
            super().keyPressEvent(event)

    def closeEvent(self, event):
        """Clean up temporary file and resources on close."""
        # Clean up Excel workbook if loaded
        if hasattr(self, '_excel_workbook') and self._excel_workbook:
            try:
                if hasattr(self._excel_workbook, 'close'):
                    self._excel_workbook.close()
            except Exception as e:
                logger.warning(f"Could not close Excel workbook: {e}")
            self._excel_workbook = None
            self._excel_sheet_names = None
            if hasattr(self, '_excel_is_xlrd'):
                delattr(self, '_excel_is_xlrd')

        # Clean up temporary file
        if self.document_path and os.path.exists(self.document_path):
            try:
                os.unlink(self.document_path)
            except Exception as e:
                logger.warning(f"Could not delete temp file: {e}")
        super().closeEvent(event)


class DocumentThumbnailWidget(QtWidgets.QWidget):
    """Widget for displaying a single document thumbnail."""

    clicked = QtCore.Signal(dict)
    deleteRequested = QtCore.Signal(dict)

    def __init__(self, version_data, sg_session=None, parent=None):
        super().__init__(parent)
        self.version_data = version_data
        self.sg_session = sg_session
        self.selected = False
        self.folders_containing_this = []

        self.setFixedSize(180, 200)
        self.setCursor(QtCore.Qt.PointingHandCursor)

        self.drag_start_position = None

        self.tooltip_timer = QtCore.QTimer()
        self.tooltip_timer.setSingleShot(True)
        self.tooltip_timer.timeout.connect(self._show_tooltip)

        self._setup_ui()
        self._load_thumbnail()

    def _setup_ui(self):
        """Setup the thumbnail UI."""
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(5, 5, 5, 5)
        layout.setSpacing(5)

        # Thumbnail placeholder
        self.thumbnail_label = QtWidgets.QLabel()
        self.thumbnail_label.setFixedSize(170, 140)
        self.thumbnail_label.setAlignment(QtCore.Qt.AlignCenter)
        self.thumbnail_label.setStyleSheet("""
            QLabel {
                background-color: #2b2b2b;
                border: 2px solid #444;
                border-radius: 4px;
            }
        """)

        self.thumbnail_label.setText("Loading...")
        self.thumbnail_label.setStyleSheet(self.thumbnail_label.styleSheet() + "color: #888;")

        layout.addWidget(self.thumbnail_label)

        # Document name
        version_code = self.version_data.get('code', 'Unknown')
        name_label = QtWidgets.QLabel(version_code)
        name_label.setWordWrap(True)
        name_label.setMaximumWidth(170)
        name_label.setAlignment(QtCore.Qt.AlignCenter)
        name_label.setStyleSheet("font-size: 11px; color: #ddd;")
        layout.addWidget(name_label)

        # Document type
        version_type = self.version_data.get('sg_version_type', '')
        if isinstance(version_type, dict):
            version_type = version_type.get('name', '')
        type_label = QtWidgets.QLabel(version_type or 'Unknown Type')
        type_label.setAlignment(QtCore.Qt.AlignCenter)
        type_label.setStyleSheet("font-size: 9px; color: #888;")
        layout.addWidget(type_label)

    def _load_thumbnail(self):
        """Load thumbnail or show document type icon."""
        # Get file extension from uploaded movie
        uploaded_movie = self.version_data.get('sg_uploaded_movie')
        filename = ''
        if uploaded_movie:
            if isinstance(uploaded_movie, dict):
                filename = uploaded_movie.get('name', '')
            elif isinstance(uploaded_movie, str):
                filename = uploaded_movie

        ext = os.path.splitext(filename)[1].lower() if filename else ''

        # Check if there's an image thumbnail
        image_data = self.version_data.get('image')
        thumbnail_url = None

        if image_data:
            if isinstance(image_data, str):
                thumbnail_url = image_data
            elif isinstance(image_data, dict):
                thumbnail_url = image_data.get('url')

        if thumbnail_url:
            # Load actual thumbnail
            self._load_thumbnail_from_url(thumbnail_url)
        else:
            # Show document type icon
            self._show_document_icon(ext)

    def _load_thumbnail_from_url(self, url):
        """Load thumbnail from URL."""
        from PySide6.QtCore import QThread, QObject, Signal
        import requests

        class ThumbnailLoader(QObject):
            finished = Signal(bytes)
            error = Signal()

            def __init__(self, url):
                super().__init__()
                self.url = url

            def run(self):
                try:
                    response = requests.get(self.url, timeout=10)
                    if response.status_code == 200:
                        self.finished.emit(response.content)
                    else:
                        self.error.emit()
                except Exception as e:
                    logger.error(f"Failed to download thumbnail: {e}")
                    self.error.emit()

        self.loader = ThumbnailLoader(url)
        self.loader_thread = QThread()
        self.loader.moveToThread(self.loader_thread)
        self.loader_thread.started.connect(self.loader.run)
        self.loader.finished.connect(self._on_thumbnail_loaded)
        self.loader.error.connect(self._on_thumbnail_error)
        self.loader_thread.start()

    def _on_thumbnail_loaded(self, image_data):
        """Handle thumbnail loaded."""
        try:
            pixmap = QtGui.QPixmap()
            pixmap.loadFromData(image_data)

            if not pixmap.isNull():
                scaled_pixmap = pixmap.scaled(
                    170, 140,
                    QtCore.Qt.KeepAspectRatio,
                    QtCore.Qt.SmoothTransformation
                )
                self.thumbnail_label.setPixmap(scaled_pixmap)
                self.thumbnail_label.setText("")

            if hasattr(self, 'loader_thread'):
                self.loader_thread.quit()
                self.loader_thread.wait()

        except Exception as e:
            logger.error(f"Error displaying thumbnail: {e}")
            self._show_document_icon('')

    def _on_thumbnail_error(self):
        """Handle thumbnail load error."""
        # Get file extension and show icon instead
        uploaded_movie = self.version_data.get('sg_uploaded_movie')
        filename = ''
        if uploaded_movie:
            if isinstance(uploaded_movie, dict):
                filename = uploaded_movie.get('name', '')
        ext = os.path.splitext(filename)[1].lower() if filename else ''
        self._show_document_icon(ext)

        if hasattr(self, 'loader_thread'):
            self.loader_thread.quit()
            self.loader_thread.wait()

    def _show_document_icon(self, ext):
        """Show document type icon."""
        # Determine icon based on extension
        if ext in ['.pdf']:
            icon_text = "PDF"
            icon_color = "#e74c3c"
        elif ext in ['.doc', '.docx']:
            icon_text = "DOC"
            icon_color = "#3498db"
        elif ext in ['.xls', '.xlsx']:
            icon_text = "XLS"
            icon_color = "#27ae60"
        elif ext in ['.txt', '.md']:
            icon_text = "TXT"
            icon_color = "#95a5a6"
        else:
            icon_text = "DOC"
            icon_color = "#9b59b6"

        self.thumbnail_label.setText(icon_text)
        self.thumbnail_label.setStyleSheet(f"""
            QLabel {{
                background-color: {icon_color};
                color: white;
                font-size: 32px;
                font-weight: bold;
                border: 2px solid #444;
                border-radius: 4px;
            }}
        """)

    def refresh_thumbnail(self, updated_version_data):
        """Refresh thumbnail with updated version data."""
        self.version_data = updated_version_data
        self._load_thumbnail()

    def mousePressEvent(self, event):
        """Handle mouse press to select thumbnail."""
        if event.button() == QtCore.Qt.LeftButton:
            self.drag_start_position = event.pos()
            self.clicked.emit(self.version_data)

    def mouseMoveEvent(self, event):
        """Handle mouse move to start drag operation."""
        if not (event.buttons() & QtCore.Qt.LeftButton):
            return
        if not self.drag_start_position:
            return

        if (event.pos() - self.drag_start_position).manhattanLength() < QtWidgets.QApplication.startDragDistance():
            return

        drag = QtGui.QDrag(self)
        mime_data = QtCore.QMimeData()

        version_id = self.version_data.get('id')
        if version_id:
            mime_data.setData("application/x-document-version-id", str(version_id).encode())
            drag.setMimeData(mime_data)

            # Set drag pixmap
            if self.thumbnail_label.pixmap():
                pixmap = self.thumbnail_label.pixmap().scaled(
                    80, 80,
                    QtCore.Qt.KeepAspectRatio,
                    QtCore.Qt.SmoothTransformation
                )
                drag.setPixmap(pixmap)
                drag.setHotSpot(pixmap.rect().center())

            drag.exec(QtCore.Qt.CopyAction)

    def mouseDoubleClickEvent(self, event):
        """Handle double-click to open document viewer."""
        if event.button() == QtCore.Qt.LeftButton:
            self._open_document_view()

    def contextMenuEvent(self, event):
        """Handle right-click context menu."""
        menu = QtWidgets.QMenu(self)

        view_action = menu.addAction("View Document")
        view_action.triggered.connect(self._open_document_view)

        menu.addSeparator()

        delete_action = menu.addAction("Delete from ShotGrid")
        delete_action.triggered.connect(self._request_delete)

        menu.exec(event.globalPos())

    def _request_delete(self):
        """Request deletion of this version."""
        self.deleteRequested.emit(self.version_data)

    def _open_document_view(self):
        """Open the document viewer dialog."""
        dialog = DocumentViewerDialog(self.version_data, self.sg_session, self)
        dialog.exec()

    def set_selected(self, selected):
        """Set the selected state."""
        self.selected = selected
        self._update_border()

    def set_folders_containing(self, folder_names):
        """Set which folders contain this document."""
        self.folders_containing_this = folder_names
        self._update_border()

    def _update_border(self):
        """Update the border based on selection and folder states."""
        if self.selected:
            border_color = "#4a9eff"
        elif len(self.folders_containing_this) >= 2:
            border_color = "#ff8c00"
        elif len(self.folders_containing_this) == 1:
            border_color = "#00cc00"
        else:
            border_color = "#444"

        # Get current background color
        current_style = self.thumbnail_label.styleSheet()
        if "background-color:" in current_style:
            # Extract background color and preserve icon style
            import re
            bg_match = re.search(r'background-color:\s*([^;]+);', current_style)
            if bg_match:
                bg_color = bg_match.group(1)
                # Check if it's an icon (colored background)
                if bg_color not in ['#2b2b2b']:
                    self.thumbnail_label.setStyleSheet(f"""
                        QLabel {{
                            background-color: {bg_color};
                            color: white;
                            font-size: 32px;
                            font-weight: bold;
                            border: 2px solid {border_color};
                            border-radius: 4px;
                        }}
                    """)
                    return

        self.thumbnail_label.setStyleSheet(f"""
            QLabel {{
                background-color: #2b2b2b;
                border: 2px solid {border_color};
                border-radius: 4px;
            }}
        """)

    def _show_tooltip(self):
        """Show tooltip with folder information."""
        if self.folders_containing_this:
            folder_list = "\n".join([f"  - {name}" for name in self.folders_containing_this])
            tooltip_text = f"Dropped to {len(self.folders_containing_this)} folder(s):\n{folder_list}"
            QtWidgets.QToolTip.showText(QtGui.QCursor.pos(), tooltip_text, self)

    def enterEvent(self, event):
        """Handle mouse enter."""
        super().enterEvent(event)
        if self.folders_containing_this:
            self.tooltip_timer.start(1000)

    def leaveEvent(self, event):
        """Handle mouse leave."""
        super().leaveEvent(event)
        self.tooltip_timer.stop()
        QtWidgets.QToolTip.hideText()


class DroppableDocumentCategoryContainer(QtWidgets.QWidget):
    """Container widget that accepts document drops for a specific category."""

    documentDroppedToCategory = QtCore.Signal(int, str)  # (document_id, target_category)

    def __init__(self, category_name, parent=None):
        super().__init__(parent)
        self.category_name = category_name
        self.setAcceptDrops(True)
        self._is_drag_over = False

    def dragEnterEvent(self, event):
        """Handle drag enter event."""
        if event.mimeData().hasFormat("application/x-document-version-id"):
            event.acceptProposedAction()
            self._is_drag_over = True
            self.setStyleSheet("background-color: rgba(74, 159, 255, 50);")
        else:
            event.ignore()

    def dragLeaveEvent(self, event):
        """Handle drag leave event."""
        self._is_drag_over = False
        self.setStyleSheet("")

    def dropEvent(self, event):
        """Handle drop event."""
        if event.mimeData().hasFormat("application/x-document-version-id"):
            document_id_bytes = event.mimeData().data("application/x-document-version-id")
            document_id = int(bytes(document_id_bytes).decode())

            event.acceptProposedAction()
            logger.info(f"Dropped document {document_id} into category {self.category_name}")

            self.documentDroppedToCategory.emit(document_id, self.category_name)

            self._is_drag_over = False
            self.setStyleSheet("")
        else:
            event.ignore()


class DocumentViewerWidget(QtWidgets.QWidget):
    """Widget for viewing document versions with thumbnails."""

    def __init__(self, sg_session, parent=None, packages_tab=None):
        super().__init__(parent)
        self.sg_session = sg_session
        self.packages_tab = packages_tab
        self.current_project_id = None
        self.all_versions = []
        self.filtered_versions = []
        self.thumbnail_widgets = []
        self.selected_thumbnail = None

        # Filter states - categories for documents
        self.filter_states = {
            'Script': True,
            'Documents': True
        }

        self.category_groups = {}

        self._is_rebuilding = False
        self._is_rearranging = False

        self.resize_timer = QtCore.QTimer()
        self.resize_timer.setSingleShot(True)
        self.resize_timer.timeout.connect(self._rebuild_thumbnails_delayed)

        self.document_poller = DocumentPoller(sg_session, self, self)
        self.document_poller.documentReady.connect(self._on_document_ready)

        self._setup_ui()

    def _setup_ui(self):
        """Setup the UI with thumbnails and filters."""
        main_layout = QtWidgets.QHBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)

        self.splitter = QtWidgets.QSplitter(QtCore.Qt.Horizontal)
        self.splitter.setChildrenCollapsible(False)

        # Left: Thumbnails with filters
        self.thumbnail_dock = self._create_thumbnail_dock()
        self.splitter.addWidget(self.thumbnail_dock)

        # Right: Folder pane
        self.folder_pane = DocumentFolderPaneWidget(self)
        self.folder_pane.document_viewer = self
        self.splitter.addWidget(self.folder_pane)

        # Connect folder signals
        self.folder_pane.documentDropped.connect(self.update_thumbnail_states)
        self.folder_pane.documentDropped.connect(self._deselect_current_thumbnail)
        self.folder_pane.documentRemoved.connect(self._unlink_document_from_package)

        self.splitter.setSizes([500, 500])
        self.splitter.setStretchFactor(0, 1)
        self.splitter.setStretchFactor(1, 1)

        self.splitter.splitterMoved.connect(self._on_splitter_moved)

        main_layout.addWidget(self.splitter)

    def _on_splitter_moved(self, pos, index):
        """Handle splitter moved."""
        self.resize_timer.stop()
        self.resize_timer.start(500)

    def _rebuild_thumbnails_delayed(self):
        """Rearrange thumbnails after resize."""
        if self.thumbnail_widgets:
            self._rearrange_thumbnails()

    def _create_thumbnail_dock(self):
        """Create the thumbnail view with filters and category groups."""
        dock_widget = QtWidgets.QWidget()
        dock_layout = QtWidgets.QVBoxLayout(dock_widget)
        dock_layout.setContentsMargins(5, 5, 5, 5)

        # Header
        header_layout = QtWidgets.QHBoxLayout()

        title_label = QtWidgets.QLabel("Documents")
        title_font = title_label.font()
        title_font.setPointSize(12)
        title_font.setBold(True)
        title_label.setFont(title_font)
        header_layout.addWidget(title_label)

        header_layout.addStretch()

        # Add Document button
        add_doc_btn = QtWidgets.QPushButton("Add Document")
        add_doc_btn.setToolTip("Browse for a document to upload to ShotGrid")
        add_doc_btn.clicked.connect(self._browse_for_document)
        header_layout.addWidget(add_doc_btn)

        # Refresh button
        refresh_btn = QtWidgets.QPushButton("Refresh")
        refresh_btn.setToolTip("Reload all documents from ShotGrid")
        refresh_btn.clicked.connect(self._refresh_documents)
        header_layout.addWidget(refresh_btn)

        # Separator
        separator = QtWidgets.QFrame()
        separator.setFrameShape(QtWidgets.QFrame.VLine)
        separator.setFrameShadow(QtWidgets.QFrame.Sunken)
        header_layout.addWidget(separator)

        # Filter label
        filter_label = QtWidgets.QLabel("Filter:")
        header_layout.addWidget(filter_label)

        # Filter checkboxes
        self.filter_checkboxes = {}
        for filter_type in ['Script', 'Documents']:
            checkbox = QtWidgets.QCheckBox(filter_type)
            checkbox.setChecked(True)
            checkbox.stateChanged.connect(lambda state, ft=filter_type: self._on_filter_changed(ft, state))
            header_layout.addWidget(checkbox)
            self.filter_checkboxes[filter_type] = checkbox

        dock_layout.addLayout(header_layout)

        # Thumbnail scroll area
        self.scroll_area = QtWidgets.QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        self.scroll_area.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        self.scroll_area.setAcceptDrops(True)

        # Container for category groups
        self.thumbnail_container = QtWidgets.QWidget()
        self.thumbnail_container.setAcceptDrops(True)
        self.thumbnail_container_layout = QtWidgets.QVBoxLayout(self.thumbnail_container)
        self.thumbnail_container_layout.setContentsMargins(5, 5, 5, 5)
        self.thumbnail_container_layout.setSpacing(10)

        # Install event filter for file uploads
        self.thumbnail_container.dragEnterEvent = self._container_drag_enter
        self.thumbnail_container.dragLeaveEvent = self._container_drag_leave
        self.thumbnail_container.dropEvent = self._container_drop

        # Create collapsible groups for each category
        self.category_groups = {}
        for category in ['Script', 'Documents']:
            group = CollapsibleGroupBox(f"{category} (0 items)")

            category_container = DroppableDocumentCategoryContainer(category)
            category_container.documentDroppedToCategory.connect(self._on_document_dropped_to_category)

            grid_layout = QtWidgets.QGridLayout(category_container)
            grid_layout.setSpacing(10)
            grid_layout.setAlignment(QtCore.Qt.AlignTop | QtCore.Qt.AlignLeft)

            group.addWidget(category_container)
            self.thumbnail_container_layout.addWidget(group)

            self.category_groups[category] = {
                'group': group,
                'container': category_container,
                'layout': grid_layout
            }

        self.thumbnail_container_layout.addStretch()

        self.scroll_area.setWidget(self.thumbnail_container)
        dock_layout.addWidget(self.scroll_area)

        return dock_widget

    def _on_filter_changed(self, filter_type, state):
        """Handle filter checkbox changes."""
        self.filter_states[filter_type] = (state == QtCore.Qt.Checked)
        self._apply_filters()

    def _apply_filters(self):
        """Apply current filters and rebuild thumbnails."""
        self.filtered_versions = []
        for version in self.all_versions:
            version_type = self._get_version_type(version)
            if self.filter_states.get(version_type, True):
                self.filtered_versions.append(version)

        self._rebuild_thumbnails()

    def _get_version_type(self, version):
        """Get the type category for a version."""
        sg_version_type = version.get('sg_version_type', '')
        if isinstance(sg_version_type, dict):
            version_type = sg_version_type.get('name', '').lower()
        else:
            version_type = str(sg_version_type).lower()

        # Map to filter categories for documents
        if 'script' in version_type:
            return 'Script'
        else:
            return 'Documents'

    def _rearrange_thumbnails(self):
        """Rearrange existing thumbnails in grid without recreating them."""
        if self._is_rearranging or self._is_rebuilding:
            return

        try:
            self._is_rearranging = True

            if not hasattr(self, 'thumbnail_container') or not self.thumbnail_container:
                return

            if not self.thumbnail_widgets:
                return

            if not hasattr(self, 'category_groups') or not self.category_groups:
                return

            thumbnail_width = 180 + 10
            available_width = self.thumbnail_container.width()

            if available_width <= 0:
                available_width = 800

            columns = max(1, available_width // thumbnail_width)

            if columns == 0:
                columns = 2

            categorized_thumbnails = {
                'Script': [],
                'Documents': []
            }

            for thumbnail in self.thumbnail_widgets:
                category = self._get_version_type(thumbnail.version_data)
                if category in categorized_thumbnails:
                    categorized_thumbnails[category].append(thumbnail)

            for category, thumbnails in categorized_thumbnails.items():
                if not thumbnails:
                    continue

                group_data = self.category_groups.get(category)
                if not group_data:
                    continue

                layout = group_data['layout']

                for idx, thumbnail in enumerate(thumbnails):
                    row = idx // columns
                    col = idx % columns

                    layout.removeWidget(thumbnail)
                    layout.addWidget(thumbnail, row, col)

        except Exception as e:
            logger.error(f"Error rearranging thumbnails: {e}", exc_info=True)
        finally:
            self._is_rearranging = False

    def _rebuild_thumbnails(self):
        """Rebuild the thumbnail grid."""
        if self._is_rebuilding or self._is_rearranging:
            return

        try:
            self._is_rebuilding = True

            self.selected_thumbnail = None

            if self.folder_pane:
                self.folder_pane.highlight_folders_for_document(None)

            for thumbnail in self.thumbnail_widgets:
                try:
                    thumbnail.clicked.disconnect()
                    thumbnail.deleteRequested.disconnect()
                except:
                    pass
                thumbnail.deleteLater()
            self.thumbnail_widgets.clear()

            for category, group_data in self.category_groups.items():
                layout = group_data['layout']
                while layout.count():
                    item = layout.takeAt(0)
                    if item.widget():
                        widget = item.widget()
                        widget.setParent(None)
                        widget.deleteLater()

            QtCore.QCoreApplication.processEvents()

            self._add_thumbnails_grouped()
        except Exception as e:
            logger.error(f"Error rebuilding thumbnails: {e}", exc_info=True)
        finally:
            self._is_rebuilding = False

    def _add_thumbnails_grouped(self):
        """Add thumbnails grouped by category."""
        if not hasattr(self, 'category_groups') or not self.category_groups:
            logger.warning("Category groups not available")
            return

        grouped_versions = {
            'Script': [],
            'Documents': []
        }

        for version in self.filtered_versions:
            category = self._get_version_type(version)
            if category in grouped_versions:
                grouped_versions[category].append(version)

        thumbnail_width = 180 + 10
        available_width = self.thumbnail_container.width()
        if available_width <= 0:
            available_width = 800
        columns = max(1, available_width // thumbnail_width)
        if columns == 0:
            columns = 2

        for category, versions in grouped_versions.items():
            group_data = self.category_groups.get(category)
            if not group_data:
                continue

            group = group_data['group']
            layout = group_data['layout']

            count = len(versions)
            group.setTitle(f"{category} ({count} item{'s' if count != 1 else ''})")

            for idx, version in enumerate(versions):
                row = idx // columns
                col = idx % columns

                thumbnail = DocumentThumbnailWidget(version, self.sg_session, self)
                thumbnail.clicked.connect(self._on_thumbnail_clicked)
                thumbnail.deleteRequested.connect(self._on_delete_requested)
                layout.addWidget(thumbnail, row, col)
                self.thumbnail_widgets.append(thumbnail)

                # Check if needs polling
                uploaded_movie = version.get('sg_uploaded_movie')
                has_file = False
                if isinstance(uploaded_movie, dict):
                    has_file = bool(uploaded_movie.get('url') or uploaded_movie.get('name'))
                elif isinstance(uploaded_movie, str):
                    has_file = bool(uploaded_movie)

                if not has_file:
                    self.document_poller.add_version_to_poll(version.get('id'))

    def find_thumbnail_by_version_id(self, version_id):
        """Find a thumbnail widget by version ID."""
        for thumbnail in self.thumbnail_widgets:
            if thumbnail.version_data.get('id') == version_id:
                return thumbnail
        return None

    def _on_thumbnail_clicked(self, version_data):
        """Handle thumbnail click."""
        if self.selected_thumbnail:
            try:
                if self.selected_thumbnail in self.thumbnail_widgets:
                    self.selected_thumbnail.set_selected(False)
            except RuntimeError:
                pass
            self.selected_thumbnail = None

        for thumbnail in self.thumbnail_widgets:
            if thumbnail.version_data.get('id') == version_data.get('id'):
                thumbnail.set_selected(True)
                self.selected_thumbnail = thumbnail
                break

        if self.folder_pane:
            document_id = version_data.get('id')
            self.folder_pane.highlight_folders_for_document(document_id)

    def _deselect_current_thumbnail(self, document_id=None, folder_name=None, folder_type=None):
        """Deselect the currently selected thumbnail."""
        if self.selected_thumbnail:
            try:
                if self.selected_thumbnail in self.thumbnail_widgets:
                    self.selected_thumbnail.set_selected(False)
            except RuntimeError:
                pass
            self.selected_thumbnail = None

        if self.folder_pane:
            self.folder_pane.highlight_folders_for_document(None)

    def _on_document_dropped_to_category(self, document_id, target_category):
        """Handle document dropped into a category group."""
        category_to_sg_type = {
            'Script': 'Script',
            'Documents': 'Document'
        }

        sg_version_type = category_to_sg_type.get(target_category, 'Document')

        try:
            if self.sg_session:
                self.sg_session.sg.update(
                    'Version',
                    document_id,
                    {'sg_version_type': sg_version_type}
                )

            for version in self.all_versions:
                if version.get('id') == document_id:
                    version['sg_version_type'] = sg_version_type
                    break

            for version in self.filtered_versions:
                if version.get('id') == document_id:
                    version['sg_version_type'] = sg_version_type
                    break

            # Link to selected package with appropriate folder
            self._link_document_to_package_category(document_id, target_category)

            self._rebuild_thumbnails()
            self.update_thumbnail_states()

        except Exception as e:
            logger.error(f"Failed to update document category: {e}", exc_info=True)
            QtWidgets.QMessageBox.critical(
                self,
                "Update Failed",
                f"Failed to update document category:\n{str(e)}"
            )

    def _link_document_to_package_category(self, document_id, category):
        """Link a document to the currently selected package in a category folder.

        Args:
            document_id: ID of the document version
            category: Category name ('Script' or 'Documents')
        """
        logger.warning(f"DEBUG _link_document_to_package_category called: doc={document_id}, category={category}")

        # Get selected package from folder pane
        if not self.folder_pane:
            logger.warning("DEBUG No folder_pane, skipping package link")
            return

        selected_package = self.folder_pane.get_selected_package()
        logger.warning(f"DEBUG Selected package from folder_pane: {selected_package}")

        if not selected_package:
            logger.warning("DEBUG No package selected, skipping package link")
            return

        try:
            # Get package ID from packages_tab
            if not self.packages_tab:
                logger.warning("DEBUG No packages_tab reference, skipping package link")
                return

            sg_package_id = None
            packages = getattr(self.packages_tab, 'packages', {})
            logger.warning(f"DEBUG packages dict has {len(packages)} packages: {list(packages.keys())}")

            for pkg_name, pkg_data in packages.items():
                if pkg_name == selected_package:
                    sg_package_id = pkg_data.get('sg_package_id')
                    logger.warning(f"DEBUG Found package '{pkg_name}' with sg_package_id={sg_package_id}")
                    break

            if not sg_package_id:
                logger.warning(f"DEBUG Could not find package ID for '{selected_package}' in packages: {list(packages.keys())}")
                return

            # Determine folder path based on category
            folder_path = f"/{category}"

            # Link version to package with folder
            logger.warning(f"DEBUG Calling link_version_to_package_with_folder: version={document_id}, package={sg_package_id}, folder={folder_path}")
            self.sg_session.link_version_to_package_with_folder(
                version_id=document_id,
                package_id=sg_package_id,
                folder_name=folder_path
            )

            logger.warning(f"DEBUG Successfully linked document {document_id} to package '{selected_package}' in {folder_path}")

            # Refresh package data tree if available
            if hasattr(self.packages_tab, 'package_data_tree') and self.packages_tab.package_data_tree:
                self.packages_tab.package_data_tree.load_package_versions(sg_package_id)

        except Exception as e:
            logger.error(f"Failed to link document to package: {e}", exc_info=True)

    def _on_delete_requested(self, version_data):
        """Handle delete request from thumbnail context menu."""
        version_code = version_data.get('code', 'Unknown')
        version_id = version_data.get('id')

        if not version_id:
            logger.error("Cannot delete version: no ID found")
            return

        reply = QtWidgets.QMessageBox.question(
            self,
            "Confirm Deletion",
            f"Are you sure you want to delete '{version_code}' from ShotGrid?\n\n"
            "This action cannot be undone.",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No
        )

        if reply != QtWidgets.QMessageBox.Yes:
            return

        try:
            QtWidgets.QApplication.setOverrideCursor(QtCore.Qt.WaitCursor)

            self.sg_session.sg.delete('Version', version_id)

            self.all_versions = [v for v in self.all_versions if v.get('id') != version_id]
            self.filtered_versions = [v for v in self.filtered_versions if v.get('id') != version_id]

            if self.selected_thumbnail and self.selected_thumbnail.version_data.get('id') == version_id:
                self.selected_thumbnail = None

            if version_id in self.document_poller.pending_versions:
                self.document_poller.pending_versions.pop(version_id, None)

            self._rebuild_thumbnails()

            QtWidgets.QApplication.restoreOverrideCursor()

            QtWidgets.QMessageBox.information(
                self,
                "Deleted",
                f"'{version_code}' has been deleted from ShotGrid."
            )

        except Exception as e:
            QtWidgets.QApplication.restoreOverrideCursor()
            logger.error(f"Failed to delete version {version_id}: {e}", exc_info=True)
            QtWidgets.QMessageBox.critical(
                self,
                "Delete Failed",
                f"Failed to delete '{version_code}' from ShotGrid:\n{str(e)}"
            )

    def _on_document_ready(self, version_id, updated_version_data):
        """Handle document becoming ready from the poller."""
        for i, version in enumerate(self.all_versions):
            if version.get('id') == version_id:
                self.all_versions[i] = updated_version_data
                break

        for i, version in enumerate(self.filtered_versions):
            if version.get('id') == version_id:
                self.filtered_versions[i] = updated_version_data
                break

    def _on_document_uploaded(self, version_data):
        """Handle successful document upload."""
        if version_data not in self.all_versions:
            self.all_versions.append(version_data)

        QtCore.QTimer.singleShot(500, self._rebuild_after_upload)

    def _rebuild_after_upload(self):
        """Rebuild UI after document upload."""
        if hasattr(self, 'folder_pane') and self.folder_pane:
            if hasattr(self.folder_pane, 'shared_document_loader'):
                loader = self.folder_pane.shared_document_loader

                alive_threads = [t for t in loader.active_threads if t.is_alive()]

                if alive_threads or loader.active_loads > 0:
                    QtCore.QTimer.singleShot(300, self._rebuild_after_upload)
                    return

        for _ in range(5):
            QtCore.QCoreApplication.processEvents()
            QtCore.QThread.msleep(30)

        self.filtered_versions = []
        for version in self.all_versions:
            version_type = self._get_version_type(version)
            if self.filter_states.get(version_type, True):
                self.filtered_versions.append(version)

        self._rebuild_thumbnails()
        self.update_folder_pane()
        self.update_thumbnail_states()

    def load_project_versions(self, project_id):
        """Load all document versions for the given project."""
        self.current_project_id = project_id

        if not project_id or not self.sg_session:
            self.all_versions = []
            self._rebuild_thumbnails()
            return

        # Get all document versions for the project
        self.all_versions = self.sg_session.get_all_document_versions_for_project(project_id)

        self._apply_filters()

    def _refresh_documents(self):
        """Refresh all documents by reloading from ShotGrid."""
        if not self.current_project_id:
            return

        try:
            QtWidgets.QApplication.setOverrideCursor(QtCore.Qt.WaitCursor)

            self.all_versions = []
            self.filtered_versions = []

            self.all_versions = self.sg_session.get_all_document_versions_for_project(self.current_project_id)

            self._apply_filters()
            self.update_folder_pane()
            self.update_thumbnail_states()

            QtWidgets.QApplication.restoreOverrideCursor()

        except Exception as e:
            QtWidgets.QApplication.restoreOverrideCursor()
            logger.error(f"Failed to refresh documents: {e}", exc_info=True)
            QtWidgets.QMessageBox.critical(
                self,
                "Refresh Failed",
                f"Failed to refresh documents from ShotGrid:\n{str(e)}"
            )

    def _browse_for_document(self):
        """Open file browser to select a document to upload."""
        if not self.current_project_id or not self.sg_session:
            QtWidgets.QMessageBox.warning(
                self,
                "No Project Selected",
                "Please select a project before uploading documents."
            )
            return

        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            "Select Document to Upload",
            "",
            "Documents (*.pdf *.doc *.docx *.xls *.xlsx *.txt *.md);;All Files (*)"
        )

        if file_path:
            self._upload_document(file_path)

    def _container_drag_enter(self, event):
        """Handle drag enter on the thumbnail container."""
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            self.thumbnail_container.setStyleSheet("background-color: rgba(74, 159, 255, 30);")
        else:
            event.ignore()

    def _container_drag_leave(self, event):
        """Handle drag leave on the thumbnail container."""
        self.thumbnail_container.setStyleSheet("")

    def _container_drop(self, event):
        """Handle drop on the thumbnail container."""
        self.thumbnail_container.setStyleSheet("")

        if not self.current_project_id or not self.sg_session:
            QtWidgets.QMessageBox.warning(
                self,
                "No Project Selected",
                "Please select a project before uploading documents."
            )
            event.ignore()
            return

        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            if urls:
                file_path = urls[0].toLocalFile()
                if self._is_valid_document_file(file_path):
                    event.acceptProposedAction()
                    self._upload_document(file_path)
                else:
                    QtWidgets.QMessageBox.warning(
                        self,
                        "Invalid File",
                        "Please drop a valid document file (PDF, DOC, DOCX, XLS, XLSX, TXT)"
                    )
        else:
            event.ignore()

    def _is_valid_document_file(self, file_path):
        """Check if the file is a valid document."""
        valid_extensions = ['.pdf', '.doc', '.docx', '.xls', '.xlsx', '.txt', '.md']
        _, ext = os.path.splitext(file_path.lower())
        return ext in valid_extensions

    def _upload_document(self, file_path):
        """Upload a document to ShotGrid."""
        dialog = UploadDocumentTypeDialog(self)
        if dialog.exec() != QtWidgets.QDialog.Accepted:
            return

        selected_type = dialog.get_selected_type()
        if not selected_type:
            return

        progress = QtWidgets.QProgressDialog(
            "Uploading document to ShotGrid...",
            "Cancel",
            0, 0,
            self
        )
        progress.setWindowModality(QtCore.Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.setValue(0)

        try:
            type_mapping = {
                'script': 'Script',
                'document': 'Document'
            }
            sg_version_type = type_mapping.get(selected_type, 'Document')

            filename = os.path.basename(file_path)
            version_code = os.path.splitext(filename)[0]

            version_data = {
                'project': {'type': 'Project', 'id': self.current_project_id},
                'code': version_code,
                'sg_version_type': sg_version_type,
                'description': f'Uploaded via FF Bidding App'
            }

            version = self.sg_session.sg.create('Version', version_data)

            self.sg_session.sg.upload(
                'Version',
                version['id'],
                file_path,
                field_name='sg_uploaded_movie'
            )

            version = self.sg_session.sg.find_one(
                'Version',
                [['id', 'is', version['id']]],
                ['code', 'image', 'sg_version_type', 'created_at', 'project', 'sg_uploaded_movie']
            )

            progress.close()

            QtWidgets.QMessageBox.information(
                self,
                "Upload Successful",
                f"Document uploaded successfully as '{version_code}'"
            )

            self._on_document_uploaded(version)

        except Exception as e:
            progress.close()
            logger.error(f"Failed to upload document: {e}", exc_info=True)
            QtWidgets.QMessageBox.critical(
                self,
                "Upload Failed",
                f"Failed to upload document to ShotGrid:\n{str(e)}"
            )

    def clear(self):
        """Clear all thumbnails."""
        self.all_versions = []
        self.filtered_versions = []
        self._rebuild_thumbnails()

    def update_folder_pane(self):
        """Update the folder pane with assets and scenes from VFX Breakdown."""
        if not self.packages_tab:
            logger.info("No packages_tab reference, skipping folder pane update")
            return

        breakdown_widget = getattr(self.packages_tab, 'breakdown_widget', None)
        if not breakdown_widget:
            logger.info("No breakdown_widget found")
            return

        model = getattr(breakdown_widget, 'model', None)
        if not model:
            logger.info("No model found in breakdown_widget")
            return

        all_scenes = getattr(model, 'all_bidding_scenes_data', [])
        if not all_scenes:
            logger.info("No bidding scenes data available")
            return

        unique_assets = set()
        unique_scenes = set()

        for scene in all_scenes:
            bid_assets = scene.get('sg_bid_assets', [])
            if bid_assets:
                if isinstance(bid_assets, list):
                    for asset in bid_assets:
                        if isinstance(asset, dict):
                            asset_name = asset.get('name')
                            if asset_name:
                                unique_assets.add(asset_name)
                        elif isinstance(asset, str):
                            unique_assets.add(asset)

            scene_code = scene.get('sg_sequence_code')
            if scene_code:
                if isinstance(scene_code, str):
                    unique_scenes.add(scene_code)
                elif isinstance(scene_code, dict):
                    scene_name = scene_code.get('name')
                    if scene_name:
                        unique_scenes.add(scene_name)

        self.folder_pane.set_assets(list(unique_assets))
        self.folder_pane.set_scenes(list(unique_scenes))

        self._load_folder_mappings()

    def _load_folder_mappings(self):
        """Load saved document-to-folder mappings from settings."""
        if not self.packages_tab:
            return

        current_package = getattr(self.packages_tab, 'current_package_name', None)
        if not current_package:
            return

        packages = getattr(self.packages_tab, 'packages', {})
        package_data = packages.get(current_package, {})
        folder_mappings = package_data.get('document_folder_mappings', None)

        if folder_mappings:
            self.folder_pane.load_folder_mappings(folder_mappings)
            self.update_thumbnail_states()

    def save_folder_mappings(self):
        """Save document-to-folder mappings to current package."""
        if not self.packages_tab:
            return

        current_package = getattr(self.packages_tab, 'current_package_name', None)
        if not current_package:
            return

        packages = getattr(self.packages_tab, 'packages', {})
        if current_package not in packages:
            return

        mappings = self.folder_pane.get_folder_mappings()
        packages[current_package]['document_folder_mappings'] = mappings

    def update_thumbnail_states(self, dropped_document_id=None, folder_name=None, folder_type=None):
        """Update all thumbnail border states based on folder mappings."""
        if not self.folder_pane:
            return

        # Only link for asset/scene folder drops, not for section drops
        # Section drops (Script/Documents) are already handled by document_folder_pane_widget
        if dropped_document_id and folder_name and folder_type and folder_type not in ("section",):
            self._link_document_to_package(dropped_document_id, folder_name, folder_type)

        mappings = self.folder_pane.get_folder_mappings()

        document_to_folders = {}

        for asset_name, document_ids in mappings.get('assets', {}).items():
            for document_id in document_ids:
                if document_id not in document_to_folders:
                    document_to_folders[document_id] = []
                document_to_folders[document_id].append(f"Asset: {asset_name}")

        for scene_code, document_ids in mappings.get('scenes', {}).items():
            for document_id in document_ids:
                if document_id not in document_to_folders:
                    document_to_folders[document_id] = []
                document_to_folders[document_id].append(f"Scene: {scene_code}")

        for thumbnail in self.thumbnail_widgets:
            document_id = thumbnail.version_data.get('id')
            folders = document_to_folders.get(document_id, [])
            thumbnail.set_folders_containing(folders)

    def _link_document_to_package(self, document_id, folder_name, folder_type):
        """Link a document version to the currently selected package with folder info."""
        logger.warning(f"DEBUG _link_document_to_package (3-param): doc={document_id}, folder={folder_name}, type={folder_type}")

        selected_package = self.folder_pane.get_selected_package()
        if not selected_package:
            logger.warning("DEBUG No package selected (3-param)")
            return

        if not self.packages_tab:
            logger.warning("DEBUG No packages_tab reference, cannot link to ShotGrid")
            return

        sg_package_id = None
        packages = getattr(self.packages_tab, 'packages', {})
        for pkg_name, pkg_data in packages.items():
            if pkg_name == selected_package:
                sg_package_id = pkg_data.get('sg_package_id')
                break

        if not sg_package_id:
            logger.warning(f"DEBUG No ShotGrid package ID found for package '{selected_package}'")
            return

        version_data = None
        for thumbnail in self.thumbnail_widgets:
            if thumbnail.version_data.get('id') == document_id:
                version_data = thumbnail.version_data
                break

        if not version_data:
            logger.warning(f"DEBUG Could not find version data for document {document_id}")
            return

        sg_version_type = version_data.get('sg_version_type', '')
        if isinstance(sg_version_type, dict):
            category = sg_version_type.get('name', 'Document')
        else:
            category = str(sg_version_type) if sg_version_type else 'Document'

        folder_type_plural = 'assets' if folder_type == 'asset' else 'scenes'
        folder_path = f"/{folder_type_plural}/{folder_name}/{category}"

        logger.warning(f"DEBUG Creating folder path: {folder_path} (folder_type={folder_type}, folder_name={folder_name}, category={category})")

        try:
            self.sg_session.link_version_to_package_with_folder(
                version_id=document_id,
                package_id=sg_package_id,
                folder_name=folder_path
            )

            logger.warning(f"DEBUG Successfully linked document {document_id} to {folder_path}")

            if hasattr(self.packages_tab, 'package_data_tree') and self.packages_tab.package_data_tree:
                self.packages_tab.package_data_tree.load_package_versions(sg_package_id)

        except Exception as e:
            logger.error(f"Failed to link document to package: {e}", exc_info=True)

    def _unlink_document_from_package(self, document_id, folder_name, folder_type):
        """Unlink a document version from the currently selected package."""
        selected_package = self.folder_pane.get_selected_package()
        if not selected_package:
            return

        if not self.packages_tab:
            return

        sg_package_id = None
        packages = getattr(self.packages_tab, 'packages', {})
        for pkg_name, pkg_data in packages.items():
            if pkg_name == selected_package:
                sg_package_id = pkg_data.get('sg_package_id')
                break

        if not sg_package_id:
            logger.warning(f"No ShotGrid package ID found for package '{selected_package}'")
            return

        version_data = None
        for thumbnail in self.thumbnail_widgets:
            if thumbnail.version_data.get('id') == document_id:
                version_data = thumbnail.version_data
                break

        if not version_data:
            logger.warning(f"Could not find version data for document {document_id}")
            return

        sg_version_type = version_data.get('sg_version_type', '')
        if isinstance(sg_version_type, dict):
            category = sg_version_type.get('name', 'Document')
        else:
            category = str(sg_version_type) if sg_version_type else 'Document'

        folder_type_plural = 'assets' if folder_type == 'asset' else 'scenes'
        folder_path = f"/{folder_type_plural}/{folder_name}/{category}"

        try:
            self.sg_session.remove_folder_reference_from_package(
                version_id=document_id,
                package_id=sg_package_id,
                folder_path=folder_path
            )

            if hasattr(self.packages_tab, 'package_data_tree') and self.packages_tab.package_data_tree:
                self.packages_tab.package_data_tree.load_package_versions(sg_package_id)

        except Exception as e:
            logger.error(f"Failed to unlink document from package: {e}", exc_info=True)
